"""
サブグループデータ管理モジュール
ファイル読み込み、ユーザーデータ、設定データの管理を集約
"""
import os
import json
from config.common import SUBGROUP_JSON_PATH, OUTPUT_RDE_DIR, INPUT_DIR

import logging

# ロガー設定
logger = logging.getLogger(__name__)


class SubgroupDataManager:
    """サブグループデータの読み込み・管理を担当"""
    
    @staticmethod
    def load_user_entries():
        """
        subGroup.jsonからユーザーリストを取得
        
        Returns:
            list[dict] | str | None: ユーザーリスト、エラーメッセージ、またはNone
        """
        if not os.path.exists(SUBGROUP_JSON_PATH):
            return None
        try:
            with open(SUBGROUP_JSON_PATH, encoding="utf-8") as f:
                data = json.load(f)
            
            # includedセクションからユーザーを取得し、sourceとsource_formatを追加
            user_list = []
            for item in data.get("included", []):
                if item.get("type") == "user":
                    # 既存のitem構造を保持しつつ、attributesにsource情報を追加
                    user_entry = item.copy()
                    if "attributes" not in user_entry:
                        user_entry["attributes"] = {}
                    
                    # source情報を追加（フィルタ機能用）
                    user_entry["source"] = "subGroup.json"
                    user_entry["source_format"] = "included"
                    
                    user_list.append(user_entry)
            
            return user_list
        except Exception as e:
            return f"ユーザーリスト取得エラー: {e}"
    
    @staticmethod
    def load_subgroups_data():
        """
        subGroup.jsonからサブグループデータを取得
        
        Returns:
            dict | str | None: サブグループデータ、エラーメッセージ、またはNone
        """
        if not os.path.exists(SUBGROUP_JSON_PATH):
            return None
        try:
            with open(SUBGROUP_JSON_PATH, encoding="utf-8") as f:
                data = json.load(f)
            return data
        except Exception as e:
            return f"サブグループデータ取得エラー: {e}"
    
    @staticmethod
    def get_subgroups_list():
        """
        サブグループのリストを取得
        
        Returns:
            list[dict] | str | None: グループリスト、エラーメッセージ、またはNone
        """
        data = SubgroupDataManager.load_subgroups_data()
        if isinstance(data, str) or data is None:
            return data
        return [item for item in data.get("data", []) if item.get("type") == "group"]
    
    @staticmethod
    def create_user_map(user_entries):
        """
        ユーザーリストからIDマッピングを作成
        
        Args:
            user_entries (list): ユーザーエントリリスト
            
        Returns:
            dict: user_id -> user_info のマッピング
        """
        user_map = {}
        if user_entries and not isinstance(user_entries, str):
            for user in user_entries:
                user_id = user.get("id", "")
                attr = user.get("attributes", {})
                user_map[user_id] = {
                    "userName": attr.get("userName", user_id),
                    "emailAddress": attr.get("emailAddress", "")
                }
        return user_map


class SubgroupConfigManager:
    """サブグループ設定ファイルの管理を担当"""
    
    @staticmethod
    def check_required_files():
        """
        必要なファイルの存在チェックとパス返却
        
        Returns:
            dict: ファイルパス情報と不足ファイルリスト
        """
        try:
            # openpyxlの可用性チェック
            import openpyxl
            openpyxl_available = True
        except ImportError:
            openpyxl_available = False
        
        try:
            from config.common import OUTPUT_RDE_DATA_DIR, INPUT_DIR
            output_dir = OUTPUT_RDE_DATA_DIR
            input_dir = INPUT_DIR
        except Exception:
            # フォールバック: 新しいパス管理システムを使用
            from config.common import get_dynamic_file_path
            output_dir = get_dynamic_file_path('output/rde/data')
            input_dir = get_dynamic_file_path('input')
        
        # ファイルパス定義
        info_path = os.path.join(output_dir, 'info.json')
        config_json_path = os.path.join(input_dir, 'group_config.json')
        config_csv_path = os.path.join(input_dir, 'group_config.csv')
        config_xlsx_path = os.path.join(input_dir, 'group_config.xlsx')
        member_path = os.path.join(input_dir, 'rde-member.txt')
        
        # 使用可能な設定ファイル判定
        config_file_used = None
        if openpyxl_available and os.path.exists(config_xlsx_path):
            config_file_used = 'xlsx'
        elif os.path.exists(config_csv_path):
            config_file_used = 'csv'
        elif os.path.exists(config_json_path):
            config_file_used = 'json'
        
        # 不足ファイルチェック
        missing = []
        for p, label in zip([info_path, member_path], ["info.json", "rde-member.txt"]):
            if not os.path.exists(p):
                missing.append(label)
        
        if not config_file_used:
            missing.append("group_config.xlsx or group_config.csv or group_config.json")
        
        return {
            "info_path": info_path,
            "config_json_path": config_json_path,
            "config_csv_path": config_csv_path,
            "config_xlsx_path": config_xlsx_path,
            "member_path": member_path,
            "config_file_used": config_file_used,
            "missing": missing,
            "output_dir": output_dir,
            "input_dir": input_dir
        }
    
    @staticmethod
    def load_config_files(paths):
        """
        設定ファイル・メンバーリストの読み込み
        
        Args:
            paths (dict): check_required_files()の戻り値
            
        Returns:
            tuple: (info, group_config, member_lines) または (error_message, None, None)
        """
        try:
            # info.json読み込み
            with open(paths["info_path"], encoding="utf-8") as f:
                info = json.load(f)
            
            # グループ設定読み込み
            config_file_used = paths["config_file_used"]
            if config_file_used == 'xlsx':
                group_config = SubgroupConfigManager._load_xlsx_config(paths["config_xlsx_path"])
            elif config_file_used == 'csv':
                group_config = SubgroupConfigManager._load_csv_config(paths["config_csv_path"])
            else:
                with open(paths["config_json_path"], encoding="utf-8") as f:
                    group_config = json.load(f)
            
            # メンバーリスト読み込み
            with open(paths["member_path"], encoding="utf-8") as f:
                member_lines = f.read().split(';')
            
            return info, group_config, member_lines
        except Exception as e:
            return f"ファイル読み込みエラー: {e}", None, None
    
    @staticmethod
    def _load_xlsx_config(xlsx_path):
        """Excel設定ファイルの読み込み"""
        import openpyxl
        group_config = []
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            config_item = dict(zip(headers, row))
            group_config.append(config_item)
        return group_config
    
    @staticmethod
    def _load_csv_config(csv_path):
        """CSV設定ファイルの読み込み"""
        import csv
        group_config = []
        with open(csv_path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                group_config.append(row)
        return group_config


class MemberDataProcessor:
    """メンバーデータ処理専用クラス"""
    
    @staticmethod
    def load_member_info(member_path):
        """
        rde-member.txtからメンバー情報を読み込み
        
        Args:
            member_path (str): メンバーファイルパス
            
        Returns:
            dict: user_id -> member_info のマッピング
        """
        member_info = {}
        if os.path.exists(member_path):
            try:
                with open(member_path, encoding="utf-8") as f:
                    lines = f.read().split(';')
                for line in lines:
                    parts = line.strip().split(';')
                    if len(parts) >= 3:
                        user_id, name, email = parts[0], parts[1], parts[2]
                        member_info[user_id] = {"name": name, "email": email}
            except Exception as e:
                logger.warning("メンバー情報読み込みエラー: %s", e)
        return member_info
    
    @staticmethod
    def create_initial_role_mapping(user_entries, member_info):
        """
        初期ロールマッピングの作成
        
        Args:
            user_entries (list): ユーザーエントリリスト
            member_info (dict): メンバー情報
            
        Returns:
            tuple: (prechecked_user_ids, initial_roles)
        """
        prechecked_user_ids = set()
        initial_roles = {}
        
        if user_entries and not isinstance(user_entries, str):
            for user in user_entries:
                user_id = user.get("id", "")
                if user_id in member_info:
                    prechecked_user_ids.add(user_id)
                    # デフォルトはASSISTANTロール
                    initial_roles[user_id] = "ASSISTANT"
        
        return prechecked_user_ids, initial_roles
    
    @staticmethod
    def extract_subjects_for_display(subjects_data):
        """
        課題データの表示用抽出
        
        Args:
            subjects_data (list): 課題データリスト
            
        Returns:
            str: 表示用文字列
        """
        if not subjects_data:
            return ""
        
        subject_texts = []
        for subject in subjects_data:
            if isinstance(subject, dict):
                grant_number = subject.get("grantNumber", "")
                title = subject.get("title", "")
                if grant_number and title:
                    subject_texts.append(f"{grant_number}:{title}")
                else:
                    subject_texts.append(grant_number or title)
            else:
                subject_texts.append(str(subject))
        
        return ",".join(subject_texts)
    
    @staticmethod
    def extract_funds_for_display(funds_data):
        """
        研究資金データの表示用抽出
        
        Args:
            funds_data (list): 研究資金データリスト
            
        Returns:
            str: 表示用文字列
        """
        if not funds_data:
            return ""
        
        fund_numbers = []
        for fund in funds_data:
            if isinstance(fund, dict):
                fund_number = fund.get("fundNumber", "")
                if fund_number:
                    fund_numbers.append(fund_number)
            else:
                fund_numbers.append(str(fund))
        
        return ",".join(fund_numbers)
