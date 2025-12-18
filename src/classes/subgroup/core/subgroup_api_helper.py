"""
サブグループAPI関連の共通処理モジュール（移行版）
新規作成・修正タブで共有するAPI処理を提供
新しいSubgroupApiClientに段階的移行中
"""
import os
import json
import datetime
import traceback
from typing import Dict
from qt_compat.widgets import QMessageBox
from qt_compat.core import QTimer
from .subgroup_api_client import SubgroupApiClient, SubgroupPayloadBuilder

import logging
from config.common import SUBGROUP_DETAILS_DIR, SUBGROUP_REL_DETAILS_DIR

# ロガー設定
logger = logging.getLogger(__name__)

# サブグループ詳細ファイルのユーザー属性キャッシュ
_DETAIL_USER_CACHE: Dict[str, Dict[str, dict]] = {}


def fetch_user_details_by_id(user_id, bearer_token=None):
    """
    userIdからユーザー詳細情報をAPI経由で取得
    
    Args:
        user_id (str): 取得するユーザーのID
        bearer_token (str): 認証トークン（Noneの場合は自動取得を試行）
    
    Returns:
        dict: ユーザー詳細情報 {"id": "", "userName": "", "emailAddress": "", ...} 
              取得失敗時は None
    """
    if not user_id:
        logger.warning("fetch_user_details_by_id: user_idが空です")
        return None
    
    if not bearer_token:
        logger.warning("fetch_user_details_by_id: bearer_tokenが提供されていません")
        return None
    
    try:
        from classes.utils.api_request_helper import api_request
        
        # API URL構築
        api_url = f"https://rde-user-api.nims.go.jp/users/{user_id}"
        
        # ヘッダー準備（AuthorizationはBearerTokenManagerで自動設定されるため除外）
        headers = {
            'Accept': 'application/vnd.api+json',
            'Accept-Encoding': 'gzip, deflate, br, zstd',
            'Accept-Language': 'ja,en-US;q=0.9,en;q=0.8',
            'Connection': 'keep-alive',
            'Host': 'rde-user-api.nims.go.jp',
            'Origin': 'https://rde-user.nims.go.jp',
            'Referer': 'https://rde-user.nims.go.jp/',
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-site',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36',
            'sec-ch-ua': '"Not;A=Brand";v="99", "Google Chrome";v="139", "Chromium";v="139"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"'
        }
        
        logger.debug("ユーザー詳細取得API呼び出し: %s", api_url)
        
        # API呼び出し（統一管理システム使用）
        response = api_request("GET", api_url, bearer_token=bearer_token, headers=headers, timeout=10)
        response.raise_for_status()
        
        # レスポンス解析
        data = response.json()
        user_data = data.get('data', {})
        
        if user_data.get('type') != 'user':
            logger.warning("期待されるユーザータイプではありません: %s", user_data.get('type'))
            return None
        
        # 詳細情報を抽出
        attributes = user_data.get('attributes', {})
        user_details = {
            'id': user_data.get('id', user_id),
            'userName': attributes.get('userName', '').strip(),
            'emailAddress': attributes.get('emailAddress', ''),
            'familyName': attributes.get('familyName', ''),
            'givenName': attributes.get('givenName', ''),
            'organizationName': attributes.get('organizationName', ''),
            'isDeleted': attributes.get('isDeleted', False),
            'source': 'api_fetch'
        }
        
        logger.debug("ユーザー詳細取得成功: %s (%s)", user_details.get('userName', 'Unknown'), user_id)
        return user_details
        
    except Exception as e:
        logger.error("ユーザー詳細取得エラー (ID: %s): %s", user_id, e)
        return None


def _load_detail_user_attributes(subgroup_id: str) -> dict:
    """サブグループ個別ファイルからユーザー属性を取得（キャッシュ付き）。"""
    if not subgroup_id:
        return {}
    if subgroup_id in _DETAIL_USER_CACHE:
        return _DETAIL_USER_CACHE[subgroup_id]

    attr_map: dict[str, dict] = {}
    candidate_paths = [
        os.path.join(SUBGROUP_DETAILS_DIR, f"{subgroup_id}.json"),
        os.path.join(SUBGROUP_REL_DETAILS_DIR, f"{subgroup_id}.json"),
    ]

    for path in candidate_paths:
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            for item in (data or {}).get("included", []) or []:
                if item.get("type") == "user":
                    uid = item.get("id")
                    if uid:
                        attr_map[str(uid)] = item.get("attributes", {}) or {}
            if attr_map:
                break
        except FileNotFoundError:
            logger.debug("サブグループ詳細ファイル未検出: %s", path)
        except Exception as e:
            logger.debug("サブグループ詳細読込失敗 (%s): %s", path, e)

    _DETAIL_USER_CACHE[subgroup_id] = attr_map
    return attr_map


def check_subgroup_files():
    """必要なファイルの存在チェックとパス返却"""
    try:
        import openpyxl
        openpyxl_available = True
    except ImportError:
        openpyxl_available = False
    try:
        from config.common import OUTPUT_RDE_DATA_DIR, INPUT_DIR
        output_dir = OUTPUT_RDE_DATA_DIR
        input_dir = INPUT_DIR
    except Exception:
        # フォールバック: 新しいパス管理システムを使用
        from config.common import get_dynamic_file_path
        output_dir = get_dynamic_file_path('output/rde/data')
        input_dir = get_dynamic_file_path('input')
    info_path = os.path.join(output_dir, 'info.json')
    config_json_path = os.path.join(input_dir, 'group_config.json')
    config_csv_path = os.path.join(input_dir, 'group_config.csv')
    config_xlsx_path = os.path.join(input_dir, 'group_config.xlsx')
    member_path = os.path.join(input_dir, 'rde-member.txt')
    config_file_used = None
    if openpyxl_available and os.path.exists(config_xlsx_path):
        config_file_used = 'xlsx'
    elif os.path.exists(config_csv_path):
        config_file_used = 'csv'
    elif os.path.exists(config_json_path):
        config_file_used = 'json'
    missing = []
    # info.jsonのみ必須（rde-member.txtは任意）
    if not os.path.exists(info_path):
        missing.append("info.json")
    
    if not config_file_used:
        pass
        #missing.append("group_config.xlsx or group_config.csv or group_config.json")
    return {
        "info_path": info_path,
        "config_json_path": config_json_path,
        "config_csv_path": config_csv_path,
        "config_xlsx_path": config_xlsx_path,
        "member_path": member_path,
        "config_file_used": config_file_used,
        "missing": missing,
        "output_dir": output_dir,
        "input_dir": input_dir
    }


def load_subgroup_config(paths):
    """設定ファイル・メンバーリストの読み込み"""
    import csv
    try:
        with open(paths["info_path"], encoding="utf-8") as f:
            info = json.load(f)
        config_file_used = paths["config_file_used"]
        if config_file_used == 'xlsx':
            import openpyxl
            group_config = []
            wb = openpyxl.load_workbook(paths["config_xlsx_path"])
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                subjects = [s.strip() for s in str(row_dict.get('subjects', '') or '').split(',') if s.strip()]
                funds = [f.strip() for f in str(row_dict.get('funds', '') or '').split(',') if f.strip()]
                group_config.append({
                    "group_name": str(row_dict.get("group_name", "") or ""),
                    "description": str(row_dict.get("description", "") or ""),
                    "subjects": subjects,
                    "funds": funds
                })
        elif config_file_used == 'csv':
            group_config = []
            with open(paths["config_csv_path"], encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    subjects = [s.strip() for s in row.get('subjects', '').split(',') if s.strip()]
                    funds = [f.strip() for f in row.get('funds', '').split(',') if f.strip()]
                    group_config.append({
                        "group_name": row.get("group_name", ""),
                        "description": row.get("description", ""),
                        "subjects": subjects,
                        "funds": funds
                    })
        else:
            with open(paths["config_json_path"], encoding="utf-8") as f:
                group_config = json.load(f)
        with open(paths["member_path"], encoding="utf-8") as f:
            member_lines = f.read().split(';')
        return info, group_config, member_lines
    except Exception as e:
        return f"ファイル読み込みエラー: {e}", None, None


def load_unified_member_list(subgroup_id=None, dynamic_users=None, bearer_token=None):
    """
    統合メンバーリスト読み込み（rde-member.txt + subGroup.json + 動的追加ユーザー + API補完）
    
    **修正タブ専用モード**: subgroup_idが指定されている場合は、そのサブグループのメンバーのみを返す
    （rde-member.txtや動的追加ユーザーは含めない）
    
    Args:
        subgroup_id: 修正対象のサブグループID（新規作成時はNone、修正時は指定）
        dynamic_users: 動的追加されたユーザーリスト [{"id": "", "userName": "", "emailAddress": ""}, ...]
        bearer_token: API補完用の認証トークン
    
    Returns:
        tuple: (unified_users, member_info)
            unified_users: 統合されたユーザーリスト
            member_info: ユーザーID -> ユーザー詳細情報のマッピング
    """
    logger.debug("load_unified_member_list 開始: subgroup_id=%s (修正タブモード=%s), bearer_token=%s", 
                 subgroup_id, subgroup_id is not None, 'あり' if bearer_token else 'なし')
    
    # 0. subGroup.jsonを最初に読み込み（複数箇所で使用するため）
    subgroup_data = None
    detail_attr_map: Dict[str, dict] = _load_detail_user_attributes(subgroup_id) if subgroup_id else {}
    email_to_user_map = {}
    try:
        from config.common import get_dynamic_file_path
        subgroup_json_path = get_dynamic_file_path("output/rde/data/subGroup.json")
        if os.path.exists(subgroup_json_path):
            with open(subgroup_json_path, 'r', encoding='utf-8') as f:
                subgroup_data = json.load(f)
            
            # includedセクションからメールアドレス→ユーザー情報マッピングを作成
            included_items = subgroup_data.get("included", [])
            for item in included_items:
                if item.get("type") == "user":
                    attr = item.get("attributes", {})
                    email = attr.get('emailAddress', '').strip()
                    if email:
                        email_to_user_map[email] = {
                            'id': item.get('id', ''),
                            'userName': attr.get('userName', ''),
                            'emailAddress': email,
                            'organizationName': attr.get('organizationName', ''),
                            'familyName': attr.get('familyName', ''),
                            'givenName': attr.get('givenName', '')
                        }
            logger.debug("subGroup.json読み込み完了: %s名のメールマッピング作成", len(email_to_user_map))
    except Exception as e:
        logger.warning("subGroup.json事前読み込みエラー: %s", e)

    # 1. rde-member.txtからメンバー読み込み（修正タブモードではスキップ）
    rde_members = {}
    if subgroup_id is None:  # 新規作成時のみrde-member.txtを読み込む
        try:
            paths = check_subgroup_files()
            if os.path.exists(paths["member_path"]):
                with open(paths["member_path"], 'r', encoding='utf-8') as f:
                    content = f.read()
                
                # JSON形式（セミコロン区切り）とCSV形式（カンマ区切り）の両方に対応
                # まずJSON形式を試行
                is_json_format = False
                if ';' in content:
                    member_lines = content.split(';')
                    for line in member_lines:
                        line = line.strip()
                        if not line or line.startswith('#'):
                            continue
                        try:
                            member_data = json.loads(line)
                            user_id = member_data.get('id', '')
                            if user_id:
                                rde_members[user_id] = {
                                    'id': user_id,
                                    'userName': member_data.get('userName', member_data.get('name', 'Unknown')),
                                    'emailAddress': member_data.get('emailAddress', member_data.get('email', '')),
                                    'organizationName': member_data.get('organizationName', ''),
                                    'familyName': member_data.get('familyName', ''),
                                    'givenName': member_data.get('givenName', ''),
                                    'source': 'rde-member.txt',
                                    'source_format': 'json'
                                }
                                is_json_format = True
                        except json.JSONDecodeError:
                            continue
                
                # JSON形式でない場合はCSV形式（カンマ区切り）として読み込み
                if not is_json_format:
                    logger.debug("rde-member.txt: CSV形式として読み込みます")
                    
                    # CSV形式で読み込み: メールアドレス, role, canCreateDatasets, canEditMembers
                    lines = content.split('\n')
                    for line in lines:
                        line = line.strip()
                        if not line or line.startswith('#'):
                            continue
                        
                        # セミコロンで終わる場合は削除
                        if line.endswith(';'):
                            line = line[:-1]
                        
                        parts = [x.strip() for x in line.split(',')]
                        if len(parts) >= 4:
                            email = parts[0]
                            role_num = parts[1]  # 1=OWNER, 2=ASSISTANT
                            can_create = parts[2] == '1'
                            can_edit = parts[3] == '1'
                            
                            # メールアドレスからユーザー情報を取得（まずローカルマップから）
                            user_info = None
                            user_id = None
                            
                            if email in email_to_user_map:
                                user_info = email_to_user_map[email]
                                user_id = user_info['id']
                                logger.debug("CSV読み込み（ローカル）: %s -> %s (ID: %s...)", email, user_info.get('userName'), user_id[:20])
                            else:
                                # subGroup.jsonに存在しない場合、APIで検索
                                logger.debug("rde-member.txt: メールアドレス %s がローカルに存在しないため、APIで検索します", email)
                                
                                if bearer_token:
                                    import urllib.parse
                                    from net.http_helpers import proxy_get
                                    
                                    try:
                                        # APIでメールアドレス検索
                                        encoded_email = urllib.parse.quote_plus(email)
                                        filter_param = urllib.parse.quote('filter[emailAddress]', safe='')
                                        url = f"https://rde-user-api.nims.go.jp/users?{filter_param}={encoded_email}"
                                        
                                        headers = {
                                            'Authorization': f'Bearer {bearer_token}',
                                            'Accept': 'application/vnd.api+json',
                                            'Content-Type': 'application/json'
                                        }
                                        
                                        response = proxy_get(url, headers=headers)
                                        response.raise_for_status()
                                        
                                        data = response.json()
                                        users = data.get('data', [])
                                        
                                        if users:
                                            # 最初のユーザーを取得
                                            user = users[0]
                                            attr = user.get('attributes', {})
                                            user_id = user.get('id', '')
                                            user_info = {
                                                'id': user_id,
                                                'userName': attr.get('userName', 'Unknown'),
                                                'emailAddress': attr.get('emailAddress', email),
                                                'organizationName': attr.get('organizationName', ''),
                                                'familyName': attr.get('familyName', ''),
                                                'givenName': attr.get('givenName', '')
                                            }
                                            logger.debug("CSV読み込み（API）: %s -> %s (ID: %s...)", email, user_info.get('userName'), user_id[:20])
                                        else:
                                            logger.warning("rde-member.txt: メールアドレス %s のユーザーがAPIでも見つかりません", email)
                                    except Exception as api_error:
                                        logger.error("rde-member.txt API検索エラー (%s): %s", email, api_error)
                                else:
                                    logger.warning("rde-member.txt: Bearer token未提供のため、%s のAPI検索をスキップ", email)
                            
                            # ユーザー情報が取得できた場合のみ追加
                            if user_info and user_id:
                                rde_members[user_id] = {
                                    'id': user_id,
                                    'userName': user_info.get('userName', 'Unknown'),
                                    'emailAddress': email,
                                    'organizationName': user_info.get('organizationName', ''),
                                    'familyName': user_info.get('familyName', ''),
                                    'givenName': user_info.get('givenName', ''),
                                    'source': 'rde-member.txt',
                                    'source_format': 'csv',
                                    'role_from_file': 'OWNER' if role_num == '1' else 'ASSISTANT',
                                    'canCreateDatasets': can_create,
                                    'canEditMembers': can_edit
                                }
                
                logger.debug("CSV形式rde-member.txt読み込み完了: %s名", len(rde_members))
        except Exception as e:
            logger.warning("rde-member.txt読み込みエラー: %s", e)
            import traceback
            logger.debug("詳細エラー: %s", traceback.format_exc())
    else:
        logger.debug("修正タブモード: rde-member.txtの読み込みをスキップ")

    # 2. subGroup.jsonから既存ロール情報読み込み（既に読み込み済みのsubgroup_dataを使用）
    subgroup_roles = {}
    subgroup_all_users = {}  # 新規作成時：全ユーザー情報（実施機関メンバー）
    
    try:
        if subgroup_data:
            # 修正対象のサブグループを検索
            if subgroup_id:
                # まず included セクションから探す（新しい形式）
                included_items = subgroup_data.get("included", [])
                found_group = None
                
                for item in included_items:
                    if (item.get("type") == "group" and 
                        item.get("id") == subgroup_id and
                        item.get("attributes", {}).get("groupType") == "TEAM"):
                        found_group = item
                        break
                
                # 見つからない場合、従来のdata構造から探す
                if not found_group:
                    groups_to_check = []
                    
                    if isinstance(subgroup_data, dict):
                        if 'data' in subgroup_data:
                            data_section = subgroup_data['data']
                            if isinstance(data_section, list):
                                groups_to_check = data_section
                            elif isinstance(data_section, dict):
                                groups_to_check = [data_section]
                        else:
                            groups_to_check = [subgroup_data]
                    elif isinstance(subgroup_data, list):
                        groups_to_check = subgroup_data
                    
                    for group in groups_to_check:
                        if isinstance(group, dict) and group.get('id') == subgroup_id:
                            found_group = group
                            break
                
                # 見つかったグループからロール情報を抽出
                if found_group:
                    attributes = found_group.get('attributes', {})
                    if isinstance(attributes, dict):
                        roles = attributes.get('roles', [])
                        if isinstance(roles, list):
                            logger.debug("サブグループ %s のロール情報取得: %s件", subgroup_id, len(roles))
                            for role in roles:
                                if isinstance(role, dict):
                                    user_id = role.get('userId', '')
                                    if user_id:
                                        subgroup_roles[user_id] = {
                                            'role': role.get('role', 'MEMBER'),
                                            'canCreateDatasets': role.get('canCreateDatasets', False),
                                            'canEditMembers': role.get('canEditMembers', False)
                                        }
                else:
                    logger.warning("サブグループ %s が見つかりませんでした", subgroup_id)
            else:
                # 新規作成時：subGroup.jsonの全ユーザーを実施機関メンバーとして取得
                included_items = subgroup_data.get("included", [])
                for item in included_items:
                    if item.get("type") == "user":
                        user_id = item.get('id', '')
                        if user_id:
                            attr = item.get('attributes', {})
                            subgroup_all_users[user_id] = {
                                'id': user_id,
                                'userName': attr.get('userName', ''),
                                'emailAddress': attr.get('emailAddress', ''),
                                'organizationName': attr.get('organizationName', ''),
                                'familyName': attr.get('familyName', ''),
                                'givenName': attr.get('givenName', ''),
                                'source': 'subGroup.json',
                                'source_format': 'included'
                            }
                logger.debug("新規作成時：subGroup.jsonから全ユーザー取得 - %s名", len(subgroup_all_users))
    except Exception as e:
        logger.warning("subGroup.json読み込みエラー: %s", e)
        import traceback
        logger.debug("詳細エラー: %s", traceback.format_exc())

    # 3. 動的追加ユーザーを組み込み（一時ファイル + パラメータ）（修正タブモードではスキップ）
    dynamic_members = {}
    if subgroup_id is None:  # 新規作成時のみ動的ユーザーを読み込む
        # 一時ファイルから動的ユーザーを読み込み
        temp_dynamic_users = load_dynamic_users_from_temp()
        all_dynamic_users = temp_dynamic_users.copy()
        
        # パラメータで渡された動的ユーザーも追加
        if dynamic_users:
            all_dynamic_users.extend(dynamic_users)
        
        # 重複排除しながら動的メンバー辞書を作成
        for user in all_dynamic_users:
            user_id = user.get('id', '')
            if user_id and user_id not in dynamic_members:
                dynamic_members[user_id] = {
                    'id': user_id,
                    'userName': user.get('userName', user.get('name', 'Unknown')),
                    'emailAddress': user.get('emailAddress', user.get('email', '')),
                    'organizationName': user.get('organizationName', ''),
                    'familyName': user.get('familyName', ''),
                    'givenName': user.get('givenName', ''),
                    'source': 'dynamic'
                }
    else:
        logger.debug("修正タブモード: 動的追加ユーザーの読み込みをスキップ")

    # 4. 全てのユーザーIDを統合
    all_user_ids = set()
    all_user_ids.update(rde_members.keys())
    all_user_ids.update(subgroup_roles.keys())
    all_user_ids.update(subgroup_all_users.keys())  # 新規作成時の全ユーザー
    all_user_ids.update(dynamic_members.keys())
    if detail_attr_map:
        # サブグループ詳細ファイルのみ存在する場合でも全メンバーを候補に含める
        all_user_ids.update(detail_attr_map.keys())

    # 5. 統合ユーザーリスト作成
    unified_users = []
    member_info = {}
    
    for user_id in all_user_ids:
        detail_attr = detail_attr_map.get(user_id) if detail_attr_map else None
        # 優先順位: rde-member.txt > dynamic_users > subgroup_all_users（新規作成時） > subGroup.json（IDのみ）
        user_data = None
        
        if user_id in rde_members:
            user_data = rde_members[user_id]
        elif user_id in dynamic_members:
            user_data = dynamic_members[user_id]
        elif user_id in subgroup_all_users:
            # 新規作成時：subGroup.jsonの全ユーザー情報を使用
            user_data = subgroup_all_users[user_id]
        else:
            # subGroup.jsonにしか存在しない場合、API補完を試行
            # まずsubGroup.jsonのincludedセクションからユーザー情報を探す
            if detail_attr:
                user_data = {
                    'id': user_id,
                    'userName': detail_attr.get('userName', ''),
                    'emailAddress': detail_attr.get('emailAddress', ''),
                    'familyNameKanji': detail_attr.get('familyNameKanji', ''),
                    'givenNameKanji': detail_attr.get('givenNameKanji', ''),
                    'organizationName': detail_attr.get('organizationName', ''),
                    'familyName': detail_attr.get('familyName', ''),
                    'givenName': detail_attr.get('givenName', ''),
                    'source': 'subgroup_detail',
                    'source_format': 'included'
                }
            else:
                user_data = None
                
                # subGroup.jsonからユーザー詳細を探す
                try:
                    if subgroup_data:
                        included_items = subgroup_data.get("included", [])
                        for item in included_items:
                            if item.get("type") == "user" and item.get("id") == user_id:
                                attr = item.get("attributes", {})
                                user_data = {
                                    'id': user_id,
                                    'userName': attr.get('userName', ''),
                                    'emailAddress': attr.get('emailAddress', ''),
                                    'organizationName': attr.get('organizationName', ''),
                                    'familyName': attr.get('familyName', ''),
                                    'givenName': attr.get('givenName', ''),
                                    'source': 'subGroup.json',
                                    'source_format': 'included'
                                }
                                logger.debug("subGroup.jsonからユーザー詳細取得: %s (%s)", user_data.get('userName', 'Unknown'), user_id)
                                break
                except Exception as e:
                    logger.debug("subGroup.json ユーザー検索エラー: %s", e)
                
                # 見つからない場合は最小限のデータで初期化
                if not user_data:
                    user_data = {
                        'id': user_id,
                        'userName': 'Unknown User',
                        'emailAddress': '',
                        'organizationName': '',
                        'familyName': '',
                        'givenName': '',
                        'source': 'subGroup.json',
                        'source_format': 'roles_only'
                    }

        # サブグループ詳細に存在する属性で不足分を補う
        if detail_attr:
            if not user_data.get('emailAddress'):
                user_data['emailAddress'] = detail_attr.get('emailAddress', '')
            if not user_data.get('userName') or user_data.get('userName') in ['Unknown', 'Unknown User']:
                user_data['userName'] = detail_attr.get('userName', user_data.get('userName', ''))
            user_data.setdefault('organizationName', detail_attr.get('organizationName', ''))
            user_data.setdefault('familyName', detail_attr.get('familyName', ''))
            user_data.setdefault('givenName', detail_attr.get('givenName', ''))
            user_data.setdefault('familyNameKanji', detail_attr.get('familyNameKanji', ''))
            user_data.setdefault('givenNameKanji', detail_attr.get('givenNameKanji', ''))
        
        # 詳細データが不足している場合はAPI呼び出しで補完
        if (user_data.get('userName', '').strip() in ['', 'Unknown User', 'Unknown'] or 
            not user_data.get('emailAddress', '').strip()):
            
            logger.debug("ユーザー詳細データ不足: %s - API補完を試行", user_id)
            
            # 元のsourceを保存
            original_source = user_data.get('source', 'unknown')
            original_source_format = user_data.get('source_format', '')
            
            # API呼び出しでユーザー詳細を取得
            api_user_data = fetch_user_details_by_id(user_id, bearer_token)
            
            if api_user_data:
                # API取得データで既存データを更新（sourceとsource_formatは保持）
                user_data.update({
                    'userName': api_user_data.get('userName', user_data.get('userName', '')),
                    'emailAddress': api_user_data.get('emailAddress', user_data.get('emailAddress', '')),
                    'familyName': api_user_data.get('familyName', user_data.get('familyName', '')),
                    'givenName': api_user_data.get('givenName', user_data.get('givenName', '')),
                    'organizationName': api_user_data.get('organizationName', user_data.get('organizationName', '')),
                    'isDeleted': api_user_data.get('isDeleted', False),
                    'source': original_source,  # 元のsourceを保持
                    'source_format': original_source_format  # 元のsource_formatを保持
                })
                logger.debug("API補完成功: %s (source=%s, source_format=%s)", user_data.get('userName', 'Unknown'), original_source, original_source_format)
            else:
                if bearer_token:
                    logger.warning("API補完失敗: %s - 元データを使用", user_id)
                else:
                    logger.debug("bearer_token未提供のためAPI補完スキップ: %s", user_id)
        
        # 既存ロール情報があれば設定（優先順位: subgroup_roles > rde-member.txt）
        if user_id in subgroup_roles:
            user_data.update({
                'existingRole': subgroup_roles[user_id]['role'],
                'canCreateDatasets': subgroup_roles[user_id]['canCreateDatasets'],
                'canEditMembers': subgroup_roles[user_id]['canEditMembers']
            })
        # subgroup_rolesに存在しない場合、rde-member.txtからのロール情報を使用
        elif 'role_from_file' in user_data:
            user_data.update({
                'existingRole': user_data['role_from_file'],
                'canCreateDatasets': user_data.get('canCreateDatasets', False),
                'canEditMembers': user_data.get('canEditMembers', False)
            })
        
        unified_users.append(user_data)
        member_info[user_id] = user_data
    
    logger.debug("load_unified_member_list 完了:")
    logger.debug("  - rde-member.txt: %s名", len(rde_members))
    logger.debug("  - subgroup_roles: %s名", len(subgroup_roles))
    logger.debug("  - subgroup_all_users: %s名", len(subgroup_all_users))
    logger.debug("  - dynamic_members: %s名", len(dynamic_members)) 
    logger.debug("  - 統合結果: %s名", len(unified_users))
    
    return unified_users, member_info


def get_dynamic_users_temp_path():
    """動的ユーザー一時保存ファイルのパスを取得"""
    try:
        from config.common import get_dynamic_file_path
        temp_dir = get_dynamic_file_path('output/temp')
        os.makedirs(temp_dir, exist_ok=True)
        return os.path.join(temp_dir, 'dynamic_users.json')
    except Exception:
        # フォールバック: 相対パスで作成
        temp_dir = 'output/temp'
        os.makedirs(temp_dir, exist_ok=True)
        return os.path.join(temp_dir, 'dynamic_users.json')


def get_dynamic_users_backup_path():
    """動的ユーザーバックアップファイルのパスを取得"""
    try:
        from config.common import get_dynamic_file_path
        temp_dir = get_dynamic_file_path('output/temp')
        os.makedirs(temp_dir, exist_ok=True)
        return os.path.join(temp_dir, 'dynamic_users_backup.json')
    except Exception:
        # フォールバック: 相対パスで作成
        temp_dir = 'output/temp'
        os.makedirs(temp_dir, exist_ok=True)
        return os.path.join(temp_dir, 'dynamic_users_backup.json')


def load_dynamic_users_from_temp():
    """
    テンポラリファイルから動的ユーザーを読み込み
    
    Returns:
        list: 動的ユーザーリスト [{"id": "", "userName": "", ...}, ...]
    """
    temp_path = get_dynamic_users_temp_path()
    
    if not os.path.exists(temp_path):
        return []
    
    try:
        with open(temp_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get('dynamic_users', [])
    except Exception as e:
        logger.warning("動的ユーザー一時ファイル読み込みエラー: %s", e)
        return []


def save_dynamic_users_to_temp(dynamic_users):
    """
    動的ユーザーをテンポラリファイルに保存
    
    Args:
        dynamic_users (list): 動的ユーザーリスト
        
    Returns:
        bool: 保存成功可否
    """
    temp_path = get_dynamic_users_temp_path()
    
    try:
        data = {
            'dynamic_users': dynamic_users,
            'last_updated': json.dumps(os.path.getmtime(temp_path) if os.path.exists(temp_path) else 0)
        }
        
        with open(temp_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        logger.debug("動的ユーザーを一時ファイルに保存: %s件", len(dynamic_users))
        return True
        
    except Exception as e:
        logger.error("動的ユーザー一時保存エラー: %s", e)
        return False


def add_dynamic_user_to_member_list(user_data):
    """
    動的ユーザーを一時ファイルに追加保存（rde-member.txtは読取専用）
    
    Args:
        user_data: {"id": "", "userName": "", "emailAddress": "", ...}
    
    Returns:
        bool: 保存成功可否
    """
    try:
        # 既存の動的ユーザーを読み込み
        existing_dynamic_users = load_dynamic_users_from_temp()
        existing_ids = {user.get('id', '') for user in existing_dynamic_users}
        
        # 重複チェック
        user_id = user_data.get('id', '')
        if user_id in existing_ids:
            logger.debug("ユーザーID %s は既に動的ユーザーリストに存在します", user_id)
            return True
        
        # 新しいユーザーを追加
        existing_dynamic_users.append(user_data)
        
        # 一時ファイルに保存
        success = save_dynamic_users_to_temp(existing_dynamic_users)
        
        if success:
            logger.debug("動的ユーザーを一時ファイルに保存: %s", user_data.get('userName', 'Unknown'))
        
        return success
        
    except Exception as e:
        logger.error("動的ユーザー保存エラー: %s", e)
        return False


def load_dynamic_users_backup():
    """
    バックアップファイルから動的ユーザーを読み込み
    
    Returns:
        list: バックアップされた動的ユーザーリスト
    """
    backup_path = get_dynamic_users_backup_path()
    
    if not os.path.exists(backup_path):
        return []
    
    try:
        with open(backup_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data.get('backup_users', [])
    except Exception as e:
        logger.warning("動的ユーザーバックアップファイル読み込みエラー: %s", e)
        return []


def save_dynamic_users_backup(new_users):
    """
    動的ユーザーをバックアップファイルに保存（マージ処理付き）
    
    Args:
        new_users (list): 新たにバックアップするユーザーリスト
        
    Returns:
        bool: 保存成功可否
    """
    if not new_users:
        return True
    
    backup_path = get_dynamic_users_backup_path()
    
    try:
        # 既存のバックアップユーザーを読み込み
        existing_backup_users = load_dynamic_users_backup()
        existing_backup_ids = {user.get('id', '') for user in existing_backup_users}
        
        # 新規ユーザーのみをマージ
        merged_users = existing_backup_users.copy()
        new_count = 0
        
        for user in new_users:
            user_id = user.get('id', '')
            if user_id and user_id not in existing_backup_ids:
                # タイムスタンプを追加
                import datetime
                user_with_timestamp = user.copy()
                user_with_timestamp['backed_up_at'] = datetime.datetime.now().isoformat()
                
                merged_users.append(user_with_timestamp)
                existing_backup_ids.add(user_id)
                new_count += 1
        
        # バックアップファイルに保存
        backup_data = {
            'backup_users': merged_users,
            'last_backup': datetime.datetime.now().isoformat(),
            'total_users': len(merged_users)
        }
        
        with open(backup_path, 'w', encoding='utf-8') as f:
            json.dump(backup_data, f, ensure_ascii=False, indent=2)
        
        logger.debug("動的ユーザーバックアップ保存完了: 新規%s件、合計%s件", new_count, len(merged_users))
        return True
        
    except Exception as e:
        logger.error("動的ユーザーバックアップ保存エラー: %s", e)
        return False


def backup_and_clear_dynamic_users():
    """
    修正タブを開く際の動的ユーザー初期化処理
    現在の動的ユーザーをバックアップしてから一時ファイルをクリア
    
    Returns:
        bool: 処理成功可否
    """
    try:
        # 現在の動的ユーザーを読み込み
        current_dynamic_users = load_dynamic_users_from_temp()
        
        if current_dynamic_users:
            logger.debug("修正タブ開始: 動的ユーザー%s件をバックアップ中...", len(current_dynamic_users))
            
            # バックアップに保存
            backup_success = save_dynamic_users_backup(current_dynamic_users)
            
            if backup_success:
                # 一時ファイルをクリア
                clear_success = save_dynamic_users_to_temp([])
                
                if clear_success:
                    logger.debug("修正タブ開始: 動的ユーザー初期化完了 (バックアップ済み)")
                    return True
                else:
                    logger.warning("修正タブ開始: 動的ユーザークリア失敗")
                    return False
            else:
                logger.warning("修正タブ開始: 動的ユーザーバックアップ失敗")
                return False
        else:
            logger.debug("修正タブ開始: 動的ユーザーが存在しないため初期化処理スキップ")
            return True
            
    except Exception as e:
        logger.error("修正タブ開始時の動的ユーザー初期化エラー: %s", e)
        return False


def get_dynamic_users_backup_info():
    """
    動的ユーザーバックアップの情報を取得
    
    Returns:
        dict: バックアップ情報 {"total_users": int, "last_backup": str, "users": []}
    """
    backup_path = get_dynamic_users_backup_path()
    
    if not os.path.exists(backup_path):
        return {"total_users": 0, "last_backup": None, "users": []}
    
    try:
        with open(backup_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        return {
            "total_users": data.get('total_users', 0),
            "last_backup": data.get('last_backup', None),
            "users": data.get('backup_users', [])
        }
    except Exception as e:
        logger.error("バックアップ情報取得エラー: %s", e)
        return {"total_users": 0, "last_backup": None, "users": []}


def restore_dynamic_users_from_backup():
    """
    バックアップから動的ユーザーを復元
    注意: 現在の一時ファイルを上書きします
    
    Returns:
        bool: 復元成功可否
    """
    try:
        backup_users = load_dynamic_users_backup()
        
        if not backup_users:
            logger.info("バックアップが空のため復元処理をスキップ")
            return True
        
        # バックアップから一時ファイルに復元
        # backed_up_atタイムスタンプを除去して復元
        restored_users = []
        for user in backup_users:
            clean_user = {k: v for k, v in user.items() if k != 'backed_up_at'}
            restored_users.append(clean_user)
        
        success = save_dynamic_users_to_temp(restored_users)
        
        if success:
            logger.info("動的ユーザー復元完了: %s件", len(restored_users))
        
        return success
        
    except Exception as e:
        logger.error("動的ユーザー復元エラー: %s", e)
        return False


def build_subgroup_request(info, group_config, member_lines, idx, group, selected_user_ids=None, roles=None):
    """
    リクエストペイロード・ヘッダー等の組み立て
    
    Args:
        info: info.jsonの内容
        group_config: グループ設定
        member_lines: メンバー情報
        idx: グループインデックス
        group: 作成するグループ情報
        selected_user_ids: 選択されたユーザーID（旧方式互換）
        roles: ロール情報（新方式、こちらを優先）
    """
    raw_subjects = group.get("subjects", [])
    subjects = []
    for s in raw_subjects:
        if isinstance(s, dict):
            grant_number = s.get("grantNumber")
            title = s.get("title")
            if not grant_number and not title:
                continue
            if not grant_number:
                grant_number = title
            if not title:
                title = grant_number
            subjects.append({"grantNumber": grant_number, "title": title})
        else:
            grant_number = str(s)
            subjects.append({"grantNumber": grant_number, "title": grant_number})
    
    raw_funds = group.get("funds", [])
    funds = []
    for f in raw_funds:
        if isinstance(f, dict):
            fund_number = f.get("fundNumber")
            if fund_number:
                funds.append({"fundNumber": fund_number})
        else:
            fund_number = str(f)
            if fund_number:
                funds.append({"fundNumber": fund_number})
    
    # rolesが提供されていればそれを使用、なければ旧方式
    if roles:
        payload_roles = roles
    elif selected_user_ids:
        payload_roles = []
        for user_id in selected_user_ids:
            payload_roles.append({
                "userId": user_id,
                "role": "OWNER",  # デフォルトでOWNER。必要に応じてUIで選択できるよう拡張可
                "canCreateDatasets": True,
                "canEditMembers": True
            })
    else:
        payload_roles = []
    
    parent_id = info.get("project_group_id", "")
    payload = {
        "data": {
            "type": "group",
            "attributes": {
                "name": group.get("group_name", ""),
                "description": group.get("description", ""),
                "subjects": subjects,
                "funds": funds,
                "roles": payload_roles
            },
            "relationships": {
                "parent": {
                    "data": {
                        "type": "group",
                        "id": parent_id
                    }
                }
            }
        }
    }
    
    payload_str = json.dumps(payload, ensure_ascii=False, indent=2)
    api_url = "https://rde-api.nims.go.jp/groups"
    headers_dict = {
        "Accept": "application/vnd.api+json",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "ja,en-US;q=0.9,en;q=0.8",
        "Authorization": f"Bearer <YOUR_BEARER_TOKEN>",
        "Connection": "keep-alive",
        "Content-Type": "application/vnd.api+json",
        "Host": "rde-api.nims.go.jp",
        "Origin": "https://rde.nims.go.jp",
        "Referer": "https://rde.nims.go.jp/",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Safari/537.36",
        "sec-ch-ua": '"Not)A;Brand";v="8", "Chromium";v="138", "Google Chrome";v="138"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }
    
    header_order = [
        "Accept", "Accept-Encoding", "Accept-Language", "Authorization", "Connection",
        "Content-Type", "Host", "Origin", "Referer", "Sec-Fetch-Dest", "Sec-Fetch-Mode",
        "Sec-Fetch-Site", "User-Agent", "sec-ch-ua", "sec-ch-ua-mobile", "sec-ch-ua-platform"
    ]
    headers_str = '\n'.join([f'{k}: {headers_dict[k]}' for k in header_order if k in headers_dict])
    popup_text = (
        f"Request URL\n{api_url}\nRequest Method\nPOST\n\nPOST /groups HTTP/1.1\n"
        f"{headers_str}\n\n{payload_str}"
    )
    return popup_text, payload, api_url, headers_dict


def send_subgroup_request(widget, api_url, headers, payload, group_name, auto_refresh=True):
    """
    サブグループ作成APIリクエストの送信
    
    Args:
        widget: 親ウィジェット
        api_url: API URL
        headers: リクエストヘッダー
        payload: リクエストペイロード
        group_name: グループ名（メッセージ用）
        auto_refresh: 成功時の自動リフレッシュ有効/無効
    """
    from core.bearer_token_manager import BearerTokenManager
    bearer_token = BearerTokenManager.get_token_with_relogin_prompt(widget)
    if not bearer_token:
        QMessageBox.warning(widget, "認証エラー", "Bearerトークンが取得できません。ログイン状態を確認してください。")
        return False
    
    headers = headers.copy()
    headers["Authorization"] = f"Bearer {bearer_token}"
    
    try:
        from net.http_helpers import proxy_post
        resp = proxy_post(api_url, headers=headers, json=payload, timeout=15)
        if resp.status_code in (200, 201):
            QMessageBox.information(widget, "作成成功", f"サブグループ[{group_name}]の作成に成功しました。")
            
            # 成功時にsubGroup.jsonを自動再取得
            if auto_refresh:
                try:
                    def auto_refresh_func():
                        try:
                            from classes.basic.core.basic_info_logic import auto_refresh_subgroup_json
                            from classes.utils.progress_worker import SimpleProgressWorker
                            from classes.basic.ui.ui_basic_info import show_progress_dialog
                            
                            from core.bearer_token_manager import BearerTokenManager
                            bearer_token = BearerTokenManager.get_valid_token()
                            if bearer_token:
                                # プログレス表示付きで自動更新
                                worker = SimpleProgressWorker(
                                    task_func=auto_refresh_subgroup_json,
                                    task_kwargs={'bearer_token': bearer_token},
                                    task_name="サブグループ情報自動更新"
                                )
                                
                                # プログレス表示
                                progress_dialog = show_progress_dialog(widget, "サブグループ情報自動更新", worker)
                                
                                # サブグループ更新通知を送信
                                try:
                                    from classes.dataset.util.dataset_refresh_notifier import get_subgroup_refresh_notifier
                                    subgroup_notifier = get_subgroup_refresh_notifier()
                                    # 更新完了後に少し遅延して通知（ファイル更新完了を待つため）
                                    from qt_compat.core import QTimer
                                    def send_notification():
                                        try:
                                            subgroup_notifier.notify_refresh()
                                            logger.info("サブグループ更新通知を送信しました")
                                        except Exception as e:
                                            logger.warning("サブグループ更新通知送信に失敗: %s", e)
                                    QTimer.singleShot(2000, send_notification)  # 2秒後に通知
                                except Exception as e:
                                    logger.warning("サブグループ更新通知の設定に失敗: %s", e)
                                
                        except Exception as e:
                            logger.error("サブグループ情報自動更新でエラー: %s", e)
                    
                    # 少し遅延してから自動更新実行
                    QTimer.singleShot(1000, auto_refresh_func)
                    
                except Exception as e:
                    logger.warning("サブグループ情報自動更新の設定に失敗: %s", e)
            
            return True
        else:
            QMessageBox.warning(widget, "作成失敗", f"サブグループ[{group_name}]の作成に失敗しました。\n\nStatus: {resp.status_code}\n{resp.text}")
            return False
    except Exception as e:
        QMessageBox.warning(widget, "APIエラー", f"API送信中にエラーが発生しました: {e}")
        return False


def create_subgroup_payload(group_name, description, subjects, funds, roles, parent_id):
    """
    手動入力用のサブグループペイロード作成
    
    Args:
        group_name: グループ名
        description: 説明
        subjects: 課題リスト [{"grantNumber": "", "title": ""}, ...]
        funds: 研究資金リスト [{"fundNumber": ""}, ...]
        roles: ロールリスト [{"userId": "", "role": "", "canCreateDatasets": bool, "canEditMembers": bool}, ...]
        parent_id: 親グループID
    
    Returns:
        dict: APIペイロード
    """
    return {
        "data": {
            "type": "group",
            "attributes": {
                "name": group_name,
                "description": description,
                "subjects": subjects,
                "funds": funds,
                "roles": roles
            },
            "relationships": {
                "parent": {
                    "data": {
                        "type": "group",
                        "id": parent_id
                    }
                }
            }
        }
    }


def validate_user_roles(user_rows):
    """
    ユーザーロール選択のバリデーション
    
    Args:
        user_rows: ユーザー選択情報 [(user_id, owner_radio, assistant_cb, member_cb, agent_cb, viewer_cb), ...]

    Returns:
        tuple: (is_valid, error_message, roles, selected_user_ids)
    """
    owner_count = 0
    selected_user_ids = []
    roles = []
    owner_id = None

    for user_id, owner_radio, assistant_cb, member_cb, agent_cb, viewer_cb in user_rows or []:
        if owner_radio.isChecked():
            owner_count += 1
            owner_id = user_id
            roles.append({
                "userId": user_id,
                "role": "OWNER",
                "canCreateDatasets": True,
                "canEditMembers": True
            })
        elif assistant_cb.isChecked():
            selected_user_ids.append(user_id)
            roles.append({
                "userId": user_id,
                "role": "ASSISTANT",
                "canCreateDatasets": True,
                "canEditMembers": True
            })
        elif member_cb.isChecked():
            selected_user_ids.append(user_id)
            roles.append({
                "userId": user_id,
                "role": "MEMBER",
                "canCreateDatasets": False,
                "canEditMembers": False
            })
        elif agent_cb.isChecked():
            selected_user_ids.append(user_id)
            roles.append({
                "userId": user_id,
                "role": "AGENT",
                "canCreateDatasets": False,
                "canEditMembers": False
            })
        elif viewer_cb.isChecked():
            selected_user_ids.append(user_id)
            roles.append({
                "userId": user_id,
                "role": "VIEWER",
                "canCreateDatasets": False,
                "canEditMembers": False
            })

    # OWNER必須バリデーション
    if owner_count == 0:
        return False, "サブグループには必ずOWNERを1名選択してください。", [], []
    elif owner_count > 1:
        return False, f"OWNERは1名のみ選択してください。現在{owner_count}名が選択されています。", [], []
    
    if owner_id:
        selected_user_ids = [owner_id] + selected_user_ids
    
    if not selected_user_ids:
        return False, "サブグループに追加するユーザーを1人以上選択してください。", [], []
    
    return True, "", roles, selected_user_ids
