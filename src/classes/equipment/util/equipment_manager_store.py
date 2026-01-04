from __future__ import annotations

import json
import logging
from dataclasses import dataclass
from typing import Any, Dict, List, Tuple

from config.common import get_dynamic_file_path

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class EquipmentManager:
    name: str
    email: str
    note: str = ""


def _split_multi(text: str) -> List[str]:
    # ';' または改行区切りを許可
    raw = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    parts: List[str] = []
    for chunk in raw.split(";"):
        for line in chunk.split("\n"):
            v = (line or "").strip()
            if v:
                parts.append(v)
    return parts


def managers_to_placeholder_fields(managers: List[EquipmentManager]) -> Tuple[str, str, str]:
    """テンプレ向けの表示/プレースホルダ値を生成する。"""

    names = [m.name.strip() for m in (managers or []) if isinstance(m, EquipmentManager) and m.name.strip()]
    emails = [m.email.strip() for m in (managers or []) if isinstance(m, EquipmentManager) and m.email.strip()]
    notes = [m.note.strip() for m in (managers or []) if isinstance(m, EquipmentManager) and m.note.strip()]
    return ("; ".join(names), "; ".join(emails), "; ".join(notes))


def parse_managers_from_fields(names_text: str, emails_text: str, notes_text: str) -> List[EquipmentManager]:
    """UI入力（複数値）から EquipmentManager の配列を作る。"""

    names = _split_multi(names_text)
    emails = _split_multi(emails_text)
    notes = _split_multi(notes_text)
    n = max(len(names), len(emails), len(notes), 1)
    result: List[EquipmentManager] = []
    for i in range(n):
        name = names[i] if i < len(names) else ""
        email = emails[i] if i < len(emails) else ""
        note = notes[i] if i < len(notes) else ""
        m = EquipmentManager(name=name, email=email, note=note)
        if m.name or m.email or m.note:
            result.append(m)
    return result


def get_equipment_manager_store_path() -> str:
    """Return the user data path for the equipment manager store JSON."""

    return get_dynamic_file_path("input/equipment_managers.json")


def _normalize_manager_dict(item: Dict[str, Any]) -> EquipmentManager:
    name = str(item.get("name") or "").strip()
    email = str(item.get("email") or "").strip()
    note = str(item.get("note") or "").strip()
    return EquipmentManager(name=name, email=email, note=note)


def load_equipment_managers() -> Dict[str, List[EquipmentManager]]:
    """Load equipment managers mapping from JSON.

    JSON schema (minimal):
      {"AE-003": [{"name": "...", "email": "...", "note": "..."}, ...], ...}
    """

    path = get_equipment_manager_store_path()
    try:
        with open(path, "r", encoding="utf-8") as f:
            payload = json.load(f)
    except FileNotFoundError:
        return {}
    except Exception as exc:
        logger.debug("Failed to load equipment managers: %s", exc)
        return {}

    if not isinstance(payload, dict):
        return {}

    result: Dict[str, List[EquipmentManager]] = {}
    for k, v in payload.items():
        equip_id = str(k or "").strip()
        if not equip_id:
            continue
        if not isinstance(v, list):
            continue
        managers: List[EquipmentManager] = []
        for item in v:
            if not isinstance(item, dict):
                continue
            m = _normalize_manager_dict(item)
            if not (m.name or m.email or m.note):
                continue
            managers.append(m)
        if managers:
            result[equip_id] = managers
    return result


def save_equipment_managers(mapping: Dict[str, List[EquipmentManager]]) -> None:
    path = get_equipment_manager_store_path()

    payload: Dict[str, Any] = {}
    for equip_id, managers in (mapping or {}).items():
        eid = str(equip_id or "").strip()
        if not eid:
            continue
        arr: List[Dict[str, str]] = []
        for m in managers or []:
            if not isinstance(m, EquipmentManager):
                continue
            if not (m.name or m.email or m.note):
                continue
            arr.append({"name": m.name, "email": m.email, "note": m.note})
        if arr:
            payload[eid] = arr

    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def merge_equipment_manager_mappings(
    base: Dict[str, List[EquipmentManager]],
    incoming: Dict[str, List[EquipmentManager]],
) -> Dict[str, List[EquipmentManager]]:
    """Merge incoming mapping into base mapping (dedupe by name/email/note).

    - Keeps existing order
    - Appends new unique managers
    """

    result: Dict[str, List[EquipmentManager]] = {k: list(v or []) for k, v in (base or {}).items()}
    for equip_id, managers in (incoming or {}).items():
        eid = str(equip_id or "").strip()
        if not eid:
            continue
        existing = result.get(eid) or []
        seen = {(m.name, m.email, m.note) for m in existing if isinstance(m, EquipmentManager)}
        merged = list(existing)
        for m in managers or []:
            if not isinstance(m, EquipmentManager):
                continue
            if not (m.name or m.email or m.note):
                continue
            key = (m.name, m.email, m.note)
            if key in seen:
                continue
            seen.add(key)
            merged.append(m)
        if merged:
            result[eid] = merged
        else:
            result.pop(eid, None)
    return result


def export_equipment_managers_to_xlsx(
    *,
    path: str,
    mapping: Dict[str, List[EquipmentManager]],
    equipment_name_by_id: Dict[str, str] | None = None,
) -> None:
    """Export equipment manager mapping to XLSX.

    Sheet columns:
      設備ID, 装置名, 管理者名, 管理者メール, 備考

    Manager fields are stored as multi-value strings (newline-separated) compatible with UI.
    """

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "equipment_managers"

    headers = ["設備ID", "装置名", "管理者名", "管理者メール", "備考"]
    ws.append(headers)

    name_map = equipment_name_by_id or {}

    equip_ids = sorted({str(k or "").strip() for k in (mapping or {}).keys()} | {str(k or "").strip() for k in name_map.keys()})
    equip_ids = [eid for eid in equip_ids if eid]

    for eid in equip_ids:
        managers = (mapping or {}).get(eid, [])
        names, emails, notes = managers_to_placeholder_fields(list(managers or []))
        # XLSXでは編集しやすいよう改行区切りを優先
        names = (names or "").replace("; ", "\n")
        emails = (emails or "").replace("; ", "\n")
        notes = (notes or "").replace("; ", "\n")
        ws.append([eid, str(name_map.get(eid, "") or ""), names, emails, notes])

    wb.save(str(path))


def import_equipment_managers_from_xlsx(
    *,
    path: str,
) -> tuple[Dict[str, List[EquipmentManager]], Dict[str, str]]:
    """Import equipment manager mapping from XLSX.

    Returns: (mapping, equipment_name_by_id)
    """

    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return ({}, {})

    header_row = rows[0]
    headers = [str(h or "").strip() for h in header_row]
    col_index: Dict[str, int] = {h: i for i, h in enumerate(headers) if h}

    def _idx(*candidates: str) -> int:
        for c in candidates:
            if c in col_index:
                return int(col_index[c])
        return -1

    eid_i = _idx("設備ID", "equipment_id", "equipmentId")
    name_i = _idx("装置名", "device_name_ja", "deviceNameJa")
    names_i = _idx("管理者名", "manager_names", "managerNames")
    emails_i = _idx("管理者メール", "manager_emails", "managerEmails")
    notes_i = _idx("備考", "note", "notes")

    if eid_i < 0:
        raise ValueError("XLSXに '設備ID' 列が見つかりません")

    mapping: Dict[str, List[EquipmentManager]] = {}
    equipment_name_by_id: Dict[str, str] = {}

    for r in rows[1:]:
        if r is None:
            continue
        # all empty row
        if not any((str(v).strip() if v is not None else "") for v in r):
            continue

        eid_val = r[eid_i] if eid_i < len(r) else ""
        eid = str(eid_val or "").strip()
        if not eid:
            continue

        device_name_val = r[name_i] if (0 <= name_i < len(r)) else ""
        device_name = str(device_name_val or "").strip() if name_i >= 0 else ""
        if device_name:
            equipment_name_by_id[eid] = device_name

        names_val = r[names_i] if (0 <= names_i < len(r)) else ""
        emails_val = r[emails_i] if (0 <= emails_i < len(r)) else ""
        notes_val = r[notes_i] if (0 <= notes_i < len(r)) else ""

        names_text = str(names_val or "").strip() if names_i >= 0 else ""
        emails_text = str(emails_val or "").strip() if emails_i >= 0 else ""
        notes_text = str(notes_val or "").strip() if notes_i >= 0 else ""

        managers = parse_managers_from_fields(names_text, emails_text, notes_text)
        if managers:
            mapping[eid] = managers

    return (mapping, equipment_name_by_id)
