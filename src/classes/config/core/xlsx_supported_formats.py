from __future__ import annotations

import math
import re
from typing import List, Dict, Optional

from openpyxl import load_workbook

from .models import SupportedFileFormatEntry


# 版表記のゆれに対応（例: 【V3版】, 【V 3 版】, V3版, v3版, V3, 【V3】等）
VERSION_RE = re.compile(r"[【(\[]?\s*[Vv]\s*(\d+)(?:\s*版)?\s*[】)\]]?")


def _is_nonempty_cell_value(v: object) -> bool:
    """Excelセル値として「値が入っている」を判定する。

    - NaN/None/空白のみは空
    - それ以外（数値/文字列）は値あり
    """
    try:
        if v is None:
            return False
        if isinstance(v, float) and math.isnan(v):
            return False
    except Exception:
        # 数値変換できない型は後続で判定
        pass

    if isinstance(v, str):
        return bool(v.strip())
    return True


def _extract_versioned_template_columns(columns: List[str]) -> tuple[Dict[int, str], Dict[int, str]]:
    """見出しに【Vn版】を含むテンプレート列を収集する。

    Returns:
        (name_cols, id_cols): {version_int: column_name}
    """
    name_cols: Dict[int, str] = {}
    id_cols: Dict[int, str] = {}

    for col in columns:
        col_s = str(col)
        ver = _extract_version(col_s)
        if ver is None:
            continue

        # 日本語/英語両対応で緩く判定
        col_lower = col_s.lower()
        has_template = ("テンプレ" in col_s) or ("template" in col_lower)
        if not has_template:
            continue

        is_id = ("id" in col_lower) or ("id" in col_s) or ("テンプレートid" in col_s.lower()) or ("テンプレートid" in col_lower)
        # 「テンプレート名」「template_name」等
        is_name = ("名" in col_s) or ("名称" in col_s) or ("name" in col_lower)

        if is_id:
            id_cols[ver] = col
        elif is_name:
            name_cols[ver] = col

    return name_cols, id_cols


def _normalize_ext(raw: str) -> str:
    """拡張子を正規化（.txt → txt, .JPEG → jpg 等）"""
    s = raw.strip().lower()
    s = s.replace("．", ".")  # 全角ドット
    # 先頭のドット除去
    s = s.lstrip('.')
    # 空白や記号除去
    s = re.sub(r"[^a-z0-9]", "", s)
    # 名称マップ
    name_map = {
        "tiff": "tif", "tif": "tif",
        "jpeg": "jpg", "jpe": "jpg", "jpg": "jpg",
        "png": "png", "csv": "csv", "json": "json",
        "excel": "xlsx", "xls": "xls", "xlsx": "xlsx",
        "txt": "txt", "pdf": "pdf", "xml": "xml",
    }
    return name_map.get(s, s)


def _extract_exts_with_desc(text: str) -> Dict[str, str]:
    """複合記述から {拡張子: 説明} 辞書を抽出。
    
    対応パターン:
    - 基本: .rawファイル（スペクトル）
    - スラッシュ区切り: .dm3/4ファイル → dm3, dm4
    - カンマ区切り: .dm3, .dm4ファイル → dm3, dm4
    - 全角カンマ区切り: .dm3、.dm4ファイル → dm3, dm4
    - 改行区切り: .tif\n.tiff → tif, tiff
    - 「または」区切り: .txt または .csv → txt, csv
    - 「or」区切り: .txt or .dx → txt, dx
    - ドットなしスラッシュ: .jdf/.jdx → jdf, jdx
    
    例: ".rawファイル（スペクトル）\n.xlsxファイル（測定条件）" 
        → {'raw': 'スペクトル', 'xlsx': '測定条件'}
    例: ".dm3/4ファイル（スペクトル/測定条件）"
        → {'dm3': 'スペクトル/測定条件', 'dm4': 'スペクトル/測定条件'}
    """
    if not isinstance(text, str):
        return {}
    
    # === 前処理: 全角文字を半角に正規化 ===
    # 全角記号を半角に統一
    t = text
    t = t.replace('．', '.')  # 全角ドット
    t = t.replace('、', ',')  # 全角カンマ
    t = t.replace('，', ',')  # 全角カンマ（別形式）
    t = t.replace('（', '(')  # 全角括弧
    t = t.replace('）', ')')  
    t = t.replace('【', '[')
    t = t.replace('】', ']')
    t = t.replace('　', ' ')  # 全角スペース
    
    # 改行を空白に統一
    t = t.replace('\n', ' ').replace('\r', ' ')
    
    # 連続する空白を1つに圧縮
    t = re.sub(r"\s+", " ", t)
    
    result = {}
    
    # === パターン1: 複合拡張子（スラッシュ区切り）の展開 ===
    # 例: .dm3/4 → .dm3, .dm4
    # 例: .tif/tiff → .tif, .tiff
    # パターン: .拡張子1/拡張子2 または .拡張子1/数字
    def expand_slash_notation(m: re.Match) -> str:
        """スラッシュ記法を展開: .dm3/4 → .dm3, .dm4"""
        prefix = m.group(1)  # 先頭の拡張子部分
        suffix = m.group(2)  # スラッシュ後の部分
        
        # 数字のみの場合（例: dm3/4）
        if suffix.isdigit():
            # 先頭から数字以外を削除して基本部分を取得
            base = re.sub(r'\d+$', '', prefix)
            # 展開: dm3, dm4
            expanded = f".{prefix}, .{base}{suffix}"
            return expanded
        else:
            # 文字列の場合（例: tif/tiff）
            expanded = f".{prefix}, .{suffix}"
            return expanded
    
    # スラッシュ区切りを展開（後続の「ファイル」や括弧は保持）
    t = re.sub(r'\.([a-z0-9]+)/([a-z0-9]+)', expand_slash_notation, t, flags=re.IGNORECASE)
    
    # === パターン2: 「または」「or」を半角カンマに統一 ===
    t = re.sub(r'\s*または\s*', ', ', t)
    t = re.sub(r'\s+or\s+', ', ', t, flags=re.IGNORECASE)
    
    # === パターン3: 複数拡張子の抽出 ===
    # マッチパターン: .ext または ext のみ（カンマ・スペース区切り対応）
    # 全体を複数の拡張子候補として分割抽出
    
    # 括弧内の説明文を一時的に保護（先に抽出）
    desc_pattern = r'[([（]([^)\]）]+)[)\]）]'
    descriptions = []
    def save_desc(m: re.Match) -> str:
        idx = len(descriptions)
        descriptions.append(m.group(1).strip())
        return f"__DESC{idx}__"
    
    t_protected = re.sub(desc_pattern, save_desc, t)
    
    # 拡張子パターン抽出: .ext または ext（「ファイル」「file」等のキーワードも除外）
    # カンマやスペースで区切られた拡張子を全て抽出
    ext_pattern = r'\.?([a-z0-9]+)(?:ファイル|file)?'
    
    # 抽出候補をリスト化
    ext_candidates = []
    for token in re.split(r'[,\s]+', t_protected):
        token = token.strip()
        if not token or token.startswith('__DESC'):
            continue
        # ドットで始まるまたは英数字のみのトークン
        m = re.match(r'^\.?([a-z0-9]+)(?:ファイル|file)?$', token, re.IGNORECASE)
        if m:
            ext_raw = m.group(1)
            # 明らかに拡張子でない文字列をフィルタ（日本語・長すぎる等）
            if len(ext_raw) <= 10 and re.match(r'^[a-z0-9]+$', ext_raw, re.IGNORECASE):
                ext_candidates.append(ext_raw)
    
    # === パターン4: 正規表現による詳細マッチング（フォールバック） ===
    # より複雑なパターンに対応（「.ext(説明)」形式）
    detailed_pattern = r'\.([a-z0-9]+)(?:ファイル|file)?(?:[([（]([^)\]）]+)[)\]）])?'
    for m in re.finditer(detailed_pattern, t, flags=re.IGNORECASE):
        ext_raw = m.group(1)
        desc = m.group(2).strip() if m.group(2) else ""
        
        if len(ext_raw) <= 10 and re.match(r'^[a-z0-9]+$', ext_raw, re.IGNORECASE):
            ext_norm = _normalize_ext(ext_raw)
            if ext_norm:
                # 説明文の保護されたプレースホルダーを復元
                for idx, saved_desc in enumerate(descriptions):
                    desc = desc.replace(f"__DESC{idx}__", saved_desc)
                
                if ext_norm in result:
                    # 既存の説明がある場合は追記
                    if desc and desc not in result[ext_norm]:
                        result[ext_norm] = result[ext_norm] + "/" + desc
                else:
                    result[ext_norm] = desc
    
    # === パターン5: シンプルな拡張子候補の追加 ===
    # 上記パターンで拾えなかった候補を追加
    for ext_raw in ext_candidates:
        ext_norm = _normalize_ext(ext_raw)
        if ext_norm and ext_norm not in result:
            result[ext_norm] = ""
    
    return result


def _extract_version(s: str) -> Optional[int]:
    if not isinstance(s, str):
        return None
    m = VERSION_RE.search(s)
    if m:
        try:
            return int(m.group(1))
        except ValueError:
            return None
    return None


def _worksheet_to_rows(xlsx_path: str) -> Dict[str, List[List[object]]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheets: Dict[str, List[List[object]]] = {}
    for worksheet in workbook.worksheets:
        rows: List[List[object]] = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(list(row))
        sheets[worksheet.title] = rows
    return sheets


def _resolve_column(columns: List[str], candidates: List[str]) -> Optional[str]:
    for candidate in candidates:
        if candidate in columns:
            return candidate
    norm = lambda s: re.sub(r"[（）()【】]", "", s)
    normalized_candidates = [norm(c) for c in candidates]
    for col in columns:
        cnorm = norm(str(col))
        for cand in normalized_candidates:
            if cand and cand in cnorm:
                return col
    return None


def _extract_sheet_records(rows: List[List[object]], col_map_candidates: Dict[str, List[str]]) -> tuple[List[str], List[Dict[str, object]]]:
    if not rows:
        return ([], [])

    candidate_headers = set(sum(col_map_candidates.values(), []))
    best_idx = 0
    best_score = -1
    probe_count = min(10, len(rows))
    for idx in range(probe_count):
        row_vals = [str(v).strip() for v in rows[idx] if v is not None]
        score = sum(1 for value in row_vals if value in candidate_headers)
        if score > best_score:
            best_score = score
            best_idx = idx

    header_row = rows[best_idx]
    header_len = len(header_row)
    headers = [str(v).strip() if v is not None else "" for v in header_row]
    records: List[Dict[str, object]] = []
    for raw_row in rows[best_idx + 1:]:
        row = list(raw_row[:header_len])
        if len(row) < header_len:
            row.extend([None] * (header_len - len(row)))
        if not any(_is_nonempty_cell_value(value) for value in row):
            continue
        records.append({headers[idx]: row[idx] for idx in range(header_len) if headers[idx]})
    return (headers, records)


def parse_supported_formats(xlsx_path: str) -> List[SupportedFileFormatEntry]:
    """複数シートXLSXを解析し、対応ファイル形式の一覧を抽出する。

    期待列（ゆらぎあり）:
      - equipment_id: ["装置ID", "機器ID", "equipment_id"]
      - file_format: ["登録ファイル形式", "ファイル形式", "file_format"]
      - template_name: ["テンプレート名", "データセットテンプレート名", "template_name"]
      - template_version_label: ["テンプレート版", "テンプレート（【V?版】）", "template_version"]
      - dataset_id: ["データセットID", "dataset_id"] (任意)
    """

    entries: List[SupportedFileFormatEntry] = []

    col_map_candidates: Dict[str, List[str]] = {
        "equipment_id": ["装置ID", "機器ID", "equipment_id"],
        "file_format": ["登録ファイル形式", "ファイル形式", "file_format"],
        "template_name": ["テンプレート名", "データセットテンプレート名", "template_name"],
        # 全角/半角括弧・文字揺らぎの両対応
        "template_version_label": [
            "テンプレート版",
            "テンプレート（【V?版】）",  # 全角閉じ括弧
            "テンプレート（【V?版】)",   # 半角閉じ括弧
            "template_version",
        ],
        "dataset_id": ["データセットID", "dataset_id"],
    }

    for sheet_name, sheet_rows in _worksheet_to_rows(xlsx_path).items():
        columns, records = _extract_sheet_records(sheet_rows, col_map_candidates)
        if not columns or not records:
            continue

        # 列解決
        equipment_col = _resolve_column(columns, col_map_candidates["equipment_id"])
        file_col = _resolve_column(columns, col_map_candidates["file_format"])
        tmpl_col = _resolve_column(columns, col_map_candidates["template_name"])
        ver_col = _resolve_column(columns, col_map_candidates["template_version_label"])
        dataset_col = _resolve_column(columns, col_map_candidates["dataset_id"])

        # 【Vn版】列が複数あるケースに対応（行ごとに最新版を選ぶ）
        versioned_name_cols, versioned_id_cols = _extract_versioned_template_columns(columns)
        has_versioned_templates = bool(versioned_name_cols or versioned_id_cols)

        # 必須列がなければスキップ（テンプレ列は通常列 or 版付き列のいずれかで良い）
        if not (equipment_col and file_col and (tmpl_col or has_versioned_templates)):
            continue

        for row in records:
            equipment_value = row.get(equipment_col) if equipment_col else None
            file_value = row.get(file_col) if file_col else None
            equipment_id = str(equipment_value).strip() if _is_nonempty_cell_value(equipment_value) else ""
            file_fmt_raw = str(file_value).strip() if _is_nonempty_cell_value(file_value) else ""
            template_name = ""
            template_version: Optional[int] = None

            if has_versioned_templates:
                # 行ごとに「最新版（値が入っている最大Vn）」を選ぶ
                candidate_versions = sorted(set(versioned_name_cols.keys()) | set(versioned_id_cols.keys()), reverse=True)
                for v in candidate_versions:
                    id_col = versioned_id_cols.get(v)
                    name_col = versioned_name_cols.get(v)
                    id_val = row.get(id_col) if id_col is not None else None
                    name_val = row.get(name_col) if name_col is not None else None

                    if _is_nonempty_cell_value(id_val) or _is_nonempty_cell_value(name_val):
                        template_version = v
                        chosen = id_val if _is_nonempty_cell_value(id_val) else name_val
                        template_name = str(chosen).strip() if chosen is not None else ""
                        break

            # 版付き列で取れない場合は従来列を使用
            if not template_name and tmpl_col:
                template_value = row.get(tmpl_col)
                template_name = str(template_value).strip() if _is_nonempty_cell_value(template_value) else ""

            version_value = row.get(ver_col) if ver_col else None
            version_label = str(version_value).strip() if _is_nonempty_cell_value(version_value) else ""
            dataset_value = row.get(dataset_col) if dataset_col else None
            dataset_id = str(dataset_value).strip() if _is_nonempty_cell_value(dataset_value) else None

            if not equipment_id or not file_fmt_raw or not template_name:
                continue

            if template_version is None:
                template_version = _extract_version(version_label)
            ext_desc_map = _extract_exts_with_desc(file_fmt_raw)
            if not ext_desc_map:
                # フォールバック
                ext_desc_map = {_normalize_ext(file_fmt_raw): ""}
            
            exts_list = sorted(ext_desc_map.keys())
            
            entries.append(
                SupportedFileFormatEntry(
                    equipment_id=equipment_id,
                    file_exts=exts_list,
                    file_descs=ext_desc_map,
                    template_name=template_name,
                    template_version=template_version,
                    source_sheet=sheet_name,
                    original_format=file_fmt_raw,
                )
            )

    # 同一(equipment_id, template_name)で最新版に集約
    latest: Dict[tuple, SupportedFileFormatEntry] = {}
    for e in entries:
        key = (e.equipment_id, e.template_name)
        cur = latest.get(key)
        if cur is None:
            latest[key] = e
        else:
            # Noneより数値を優先、数値同士なら大きい方
            def ver(v: Optional[int]) -> int:
                return -1 if v is None else v
            if ver(e.template_version) >= ver(cur.template_version):
                latest[key] = e

    return list(latest.values())
