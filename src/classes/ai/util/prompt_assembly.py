from __future__ import annotations

import copy
import json
import logging
import math
import os
import re
import time
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from hashlib import sha1
from typing import Any, Dict, Iterable, List, Optional, Tuple

from config.common import get_dynamic_file_path
from classes.ai.util.prompt_dictionary_presets import build_prompt_dictionary_preset_config

logger = logging.getLogger("RDE_AI_PROMPT")

PROMPT_MODE_FULL_EMBED = "full_embed"
PROMPT_MODE_FILTERED_EMBED = "filtered_embed"
PROMPT_FALLBACK_FULL_EMBED = "full_embed"
PROMPT_FALLBACK_KEEP_FILTERED = "keep_filtered"

_LARGE_PLACEHOLDER_SOURCES = {
    "static_material_index": "material_index_tree",
    "material_index_data": "material_index_tree",
    "dataportal_material_index": "dataportal_flat",
    "dataportal_tag": "dataportal_flat",
    "dataportal_equipment": "dataportal_flat",
}

_SOURCE_METADATA = {
    "static_material_index": {
        "label": "STATIC_MI_TREE",
        "method": "データセット名・説明・ARIM情報・ファイルツリーから抽出した語を NFKC 正規化し、MI ツリーの親子パスに部分一致スコアを付けて上位候補だけ再構築します。",
        "example": "例: 'TEM nanowire dataset / TEM と STEM を用いたナノワイヤ観察' から 'ナノ・低次元材料・物質群 > ナノワイヤ' や 'TEM' を含む枝を優先します。",
    },
    "material_index_data": {
        "label": "material_index_data",
        "method": "STATIC_MI_TREE と同じく、材料インデックスの親子パスを対象にスコアリングし、上位候補のみを JSON 再構築します。",
        "example": "例: 'battery cathode / LiFePO4' を含む説明なら、電池材料に対応する枝を優先します。",
    },
    "dataportal_material_index": {
        "label": "データポータル材料マスタ",
        "method": "フラットなマスタの id とラベルを正規化し、データセット説明や ARIM 情報に含まれる語と部分一致した項目を上位から採用します。",
        "example": "例: 'GaN device / 薄膜' を含む説明なら、GaN や薄膜に一致する候補だけを残します。",
    },
    "dataportal_tag": {
        "label": "データポータルタグマスタ",
        "method": "タグ id と表示名を全文検索し、データセット説明・実験サマリー・構造化テキストに一致したタグだけを抽出します。",
        "example": "例: 'TEM image, STEM map, electron microscopy' から TEM / STEM / 電子顕微鏡系タグを優先します。",
    },
    "dataportal_equipment": {
        "label": "データポータル装置マスタ",
        "method": "装置分類マスタの各候補に対して、ファイルツリーや説明文に含まれる装置名・略称との部分一致スコアを計算して上位候補を残します。",
        "example": "例: 'SEM_observation.tif' や '走査電子顕微鏡' を含む場合、SEM 系の装置分類を優先します。",
    },
}

_PLACEHOLDER_PATTERN = re.compile(r"{([A-Za-z0-9_]+)}")
_ALNUM_TOKEN_PATTERN = re.compile(r"[a-z0-9][a-z0-9_+./:-]{1,}")
_JA_TOKEN_PATTERN = re.compile(r"[一-龯ぁ-んァ-ヶー]{2,}")
_MIXED_TOKEN_PATTERN = re.compile(r"[a-z0-9]+[一-龯ぁ-んァ-ヶー]+|[一-龯ぁ-んァ-ヶー]+[a-z0-9]+")

_DEFAULT_QUERY_SOURCE_WEIGHTS = {
    "title": 4.0,
    "description": 3.5,
    "summary": 3.0,
    "metadata": 2.0,
    "body": 1.5,
    "file_tree": 0.75,
    "system": 0.25,
}

_DEFAULT_QUERY_SOURCE_TOKEN_LIMITS = {
    "title": 20,
    "description": 40,
    "summary": 30,
    "metadata": 30,
    "body": 30,
    "file_tree": 16,
    "system": 8,
}

_DEFAULT_ALIAS_CONFIG = {
    "general_aliases": {
        "tem": ["透過電子顕微鏡", "透過電顕"],
        "stem": ["走査透過電子顕微鏡", "走査透過電顕", "走査透過"],
        "eds": ["edx", "元素分析", "エネルギー分散型x線分析", "エネルギー分散型x線分光"],
        "fib": ["集束イオンビーム"],
        "sem": ["走査電子顕微鏡", "走査電顕"],
        "xrd": ["x線回折", "x線回折測定"],
        "afm": ["原子間力顕微鏡"],
        "ebsd": ["電子後方散乱回折"],
    },
    "source_aliases": {},
    "stopwords": [
        "final", "new", "test", "temp", "img", "data", "sample", "file", "report",
        "output", "draft", "result", "results", "copy", "backup",
    ],
    "stopwords_ja": [
        "データ", "試料", "報告", "結果", "最終", "新規", "一時", "画像", "ファイル", "サンプル",
    ],
    "file_extensions": ["pdf", "tif", "tiff", "jpg", "jpeg", "png", "csv", "txt", "tmp", "json", "xml"],
    "allowlist": ["al", "fe", "cu", "ti", "w", "c", "n", "o", "si", "au", "ag"],
    "weak_stopwords": [],
    "canonical_terms": [],
    "aliases": {},
    "source_overrides": {},
    "allowlists": {
        "global": [],
        "sources": {},
    },
    "metadata": {
        "version": 2,
        "last_scanned_at": "",
        "last_applied_at": "",
        "scan_note": "",
    },
    "assist": {
        "web_enabled": False,
        "llm_enabled": False,
    },
    "generated": {
        "candidates": [],
        "scan_results": [],
        "term_inventory": [],
        "evaluation_report": {},
    },
}

_ALIAS_CONFIG_CACHE = {"mtime": None, "value": None}
_PROMPT_DICTIONARY_SUMMARY_CACHE = {"mtime": None, "source_mtime": None, "value": None}

_PROMPT_DICTIONARY_PII_FIELD_PATTERNS = (
    "user",
    "email",
    "contact",
    "familyname",
    "givenname",
    "owner_names",
    "applicant",
    "manager_name",
    "data_owner",
)

_PROMPT_DICTIONARY_SUMMARY_EXCLUDED_COLUMNS = {
    "dataset_manager_name",
    "dataset_applicant_name",
    "dataset_owner_names_str",
    "invoice.basic.data_owner",
    "contact",
}

_PROMPT_DICTIONARY_STOPWORD_SEEDS = {
    "strong": {
        "raw",
        "recipe",
        "apparatus",
        "dataset",
        "entry",
        "image",
        "images",
        "number",
        "project",
        "result",
        "results",
        "summary",
        "ファイル",
        "画像",
        "結果",
        "装置",
        "データ",
        "試料",
        "サンプル",
    },
    "weak": {
        "group",
        "mode",
        "date",
        "name",
        "sheet",
        "report",
        "利用日",
        "番号",
        "名称",
    },
}

_PROMPT_DICTIONARY_SCAN_SOURCES = {
    "summary_entries": "output/summary.xlsx",
    "summary_datasets": "output/summary.xlsx",
    "data_entry": "output/rde/data/dataEntry",
    "invoice": "output/rde/data/invoice",
    "search_results": "output/search_results",
    "datasets_tree": "output/datasets/.datatree.new.json",
}

_PROMPT_DICTIONARY_WEB_SCAN_SOURCE = "output/data_portal_debug"

_PROMPT_DICTIONARY_RECURSIVE_SCAN_TARGETS = {
    "output_rde_data_dataEntry_recursive": "output/rde/data/dataEntry",
    "output_rde_data_datasets_recursive": "output/rde/data/datasets",
    "output_rde_data_recursive": "output/rde/data",
    "output_data_portal_public_recursive": "output/data_portal_public",
    "output_arim_reports_recursive": "output/arim-site/reports",
    "output_arim_equipment_recursive": "output/arim-site/equipment",
}

_PROMPT_DICTIONARY_AUTO_APPROVE_THRESHOLDS = {
    "alias": 0.9,
    "allowlist": 0.75,
    "stopword": 0.96,
    "weak_stopword": 0.98,
}

_PROMPT_DICTIONARY_DOMAIN_HINTS = {
    "顕微鏡",
    "回折",
    "分光",
    "分析",
    "観察",
    "評価",
    "測定",
    "堆積",
    "成膜",
    "加工",
    "リソグラフィ",
    "エッチング",
    "ナノ",
    "薄膜",
    "電子",
    "x線",
    "laser",
    "spectroscopy",
    "microscopy",
    "diffraction",
    "analysis",
    "lithography",
    "etch",
    "deposition",
    "sem",
    "tem",
    "stem",
    "fib",
    "afm",
    "xrd",
    "ebsd",
    "edx",
    "raman",
    "ftir",
}

_PROMPT_DICTIONARY_GENERIC_TERM_SEEDS = {
    "data",
    "dataset",
    "datasets",
    "sample",
    "samples",
    "result",
    "results",
    "report",
    "reports",
    "entry",
    "entries",
    "image",
    "images",
    "name",
    "description",
    "metadata",
    "project",
    "group",
    "page",
    "cache",
    "json",
    "output",
    "file",
    "files",
    "text",
    "sheet",
    "value",
    "header",
    "項目",
    "名称",
    "番号",
    "情報",
    "一覧",
    "装置",
    "結果",
    "データ",
    "試料",
    "画像",
}

_PROMPT_DICTIONARY_PROGRESS_EVENT_FILE_GRANULARITY = 25
_PROMPT_DICTIONARY_TERM_INVENTORY_LIMIT = 20000
_PROMPT_DICTIONARY_TERM_VIEW_LIMIT = 500
_PROMPT_DICTIONARY_CANDIDATE_LIMIT = 12000
_PROMPT_DICTIONARY_PRESET_SEED_VERSION = 1


def default_prompt_assembly_config() -> Dict[str, Any]:
    return {
        "default_mode": PROMPT_MODE_FULL_EMBED,
        "fallback_behavior": PROMPT_FALLBACK_FULL_EMBED,
        "min_candidates": 3,
        "max_candidates": 20,
        "max_chars": 12000,
        "max_token_estimate": 3000,
        "low_score_threshold": 0,
        "query_source_weights": copy.deepcopy(_DEFAULT_QUERY_SOURCE_WEIGHTS),
        "query_source_token_limits": copy.deepcopy(_DEFAULT_QUERY_SOURCE_TOKEN_LIMITS),
        "comparison_report_enabled": True,
        "debug_save_enabled": False,
        "features": {},
    }


def normalize_prompt_assembly_config_inplace(config: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(config, dict):
        return config

    defaults = default_prompt_assembly_config()
    prompt_assembly = config.get("prompt_assembly")
    if not isinstance(prompt_assembly, dict):
        prompt_assembly = {}
        config["prompt_assembly"] = prompt_assembly

    for key, value in defaults.items():
        if key == "features":
            prompt_assembly.setdefault(key, {})
        else:
            prompt_assembly.setdefault(key, value)

    if prompt_assembly.get("default_mode") not in {PROMPT_MODE_FULL_EMBED, PROMPT_MODE_FILTERED_EMBED}:
        prompt_assembly["default_mode"] = PROMPT_MODE_FULL_EMBED

    if prompt_assembly.get("fallback_behavior") not in {
        PROMPT_FALLBACK_FULL_EMBED,
        PROMPT_FALLBACK_KEEP_FILTERED,
    }:
        prompt_assembly["fallback_behavior"] = PROMPT_FALLBACK_FULL_EMBED

    try:
        prompt_assembly["min_candidates"] = max(1, int(prompt_assembly.get("min_candidates", defaults["min_candidates"])))
    except Exception:
        prompt_assembly["min_candidates"] = defaults["min_candidates"]
    try:
        prompt_assembly["max_candidates"] = max(1, int(prompt_assembly.get("max_candidates", defaults["max_candidates"])))
    except Exception:
        prompt_assembly["max_candidates"] = defaults["max_candidates"]
    try:
        prompt_assembly["max_chars"] = max(500, int(prompt_assembly.get("max_chars", defaults["max_chars"])))
    except Exception:
        prompt_assembly["max_chars"] = defaults["max_chars"]
    try:
        prompt_assembly["max_token_estimate"] = max(
            100,
            int(prompt_assembly.get("max_token_estimate", defaults["max_token_estimate"])),
        )
    except Exception:
        prompt_assembly["max_token_estimate"] = defaults["max_token_estimate"]
    try:
        prompt_assembly["low_score_threshold"] = max(0, int(prompt_assembly.get("low_score_threshold", 0)))
    except Exception:
        prompt_assembly["low_score_threshold"] = 0
    prompt_assembly["comparison_report_enabled"] = bool(prompt_assembly.get("comparison_report_enabled", True))
    prompt_assembly["debug_save_enabled"] = bool(prompt_assembly.get("debug_save_enabled", False))

    prompt_assembly["query_source_weights"] = _normalize_weight_map(
        prompt_assembly.get("query_source_weights"),
        _DEFAULT_QUERY_SOURCE_WEIGHTS,
        numeric_type=float,
    )
    prompt_assembly["query_source_token_limits"] = _normalize_weight_map(
        prompt_assembly.get("query_source_token_limits"),
        _DEFAULT_QUERY_SOURCE_TOKEN_LIMITS,
        numeric_type=int,
        minimum=1,
    )

    features = prompt_assembly.get("features")
    if not isinstance(features, dict):
        features = {}
        prompt_assembly["features"] = features

    for feature_id, feature_config in list(features.items()):
        if not isinstance(feature_config, dict):
            features[feature_id] = {}
            feature_config = features[feature_id]
        _normalize_feature_config_inplace(feature_config, prompt_assembly)

    return config


def estimate_token_count(text: str) -> int:
    return max(0, math.ceil(len(text or "") / 4))


def get_prompt_assembly_source_catalog() -> Dict[str, Dict[str, str]]:
    catalog: Dict[str, Dict[str, str]] = {}
    for placeholder, source_type in _LARGE_PLACEHOLDER_SOURCES.items():
        metadata = dict(_SOURCE_METADATA.get(placeholder) or {})
        metadata.setdefault("label", placeholder)
        metadata.setdefault("method", "データセット文脈との部分一致で候補を絞り込みます。")
        metadata.setdefault("example", "例は未定義です。")
        metadata["source_type"] = source_type
        catalog[placeholder] = metadata
    return catalog


def detect_prompt_assembly_sources(prompt_template: str) -> List[str]:
    placeholders = _extract_placeholders(prompt_template or "")
    return [placeholder for placeholder in placeholders if placeholder in _LARGE_PLACEHOLDER_SOURCES]


@dataclass
class PromptAssemblyResult:
    prompt: str
    diagnostics: Dict[str, Any]
    rendered_context: Dict[str, Any]


def build_prompt(
    prompt_template: str,
    context_data: Dict[str, Any],
    *,
    ai_config: Optional[Dict[str, Any]] = None,
    feature_id: str = "",
    template_name: str = "",
    template_path: str = "",
    alias_config_override: Optional[Dict[str, Any]] = None,
) -> PromptAssemblyResult:
    started = time.perf_counter()
    safe_template = prompt_template or ""
    safe_context = dict(context_data or {})
    placeholders = _extract_placeholders(safe_template)
    assembly_config = _resolve_prompt_assembly_config(ai_config, feature_id, template_name)
    request_id = _build_request_id(feature_id or template_name or template_path or "prompt")

    source_diagnostics: List[Dict[str, Any]] = []
    filtered_context = dict(safe_context)

    for placeholder in placeholders:
        source_type = _LARGE_PLACEHOLDER_SOURCES.get(placeholder)
        if not source_type:
            continue
        raw_value = _safe_str(filtered_context.get(placeholder))
        if not raw_value.strip():
            continue
        filtered_value, source_diag = _resolve_placeholder_value(
            placeholder=placeholder,
            source_type=source_type,
            raw_value=raw_value,
            context_data=filtered_context,
            assembly_config=assembly_config,
            alias_config_override=alias_config_override,
        )
        filtered_context[placeholder] = filtered_value
        source_diagnostics.append(source_diag)

    rendered_prompt = safe_template
    for key, value in filtered_context.items():
        placeholder = f"{{{key}}}"
        if placeholder in rendered_prompt:
            rendered_prompt = rendered_prompt.replace(placeholder, _safe_str(value))

    unresolved_placeholders = _extract_placeholders(rendered_prompt)
    elapsed = round(time.perf_counter() - started, 3)
    estimated_full_prompt_chars = len(rendered_prompt) + sum(
        max(0, int(item.get("original_chars", 0)) - int(item.get("filtered_chars", 0)))
        for item in source_diagnostics
        if item.get("mode") == PROMPT_MODE_FILTERED_EMBED
    )
    diagnostics = {
        "request_id": request_id,
        "feature_id": feature_id or "unknown",
        "template_name": template_name or feature_id or "unknown",
        "template_path": template_path or "",
        "mode": assembly_config.get("mode", PROMPT_MODE_FULL_EMBED),
        "fallback_behavior": assembly_config.get("fallback_behavior", PROMPT_FALLBACK_FULL_EMBED),
        "source_diagnostics": source_diagnostics,
        "fallback_used": any(item.get("fallback_used") for item in source_diagnostics),
        "prompt_chars": len(rendered_prompt),
        "prompt_token_estimate": estimate_token_count(rendered_prompt),
        "estimated_full_prompt_chars": estimated_full_prompt_chars,
        "estimated_full_prompt_tokens": estimate_token_count("x" * estimated_full_prompt_chars),
        "prompt_char_delta_vs_full": estimated_full_prompt_chars - len(rendered_prompt),
        "template_chars": len(safe_template),
        "unresolved_placeholders": unresolved_placeholders,
        "build_elapsed_seconds": elapsed,
        "comparison_report_enabled": bool(assembly_config.get("comparison_report_enabled", True)),
        "debug_save_enabled": bool(assembly_config.get("debug_save_enabled", False)),
        "source_summary": [item.get("summary") for item in source_diagnostics if item.get("summary")],
    }

    _log_build_summary(diagnostics)
    _append_event({"event": "build", **diagnostics})
    _save_debug_artifacts_if_needed(
        diagnostics=diagnostics,
        template_text=safe_template,
        prompt_text=rendered_prompt,
        source_diagnostics=source_diagnostics,
    )

    return PromptAssemblyResult(
        prompt=rendered_prompt,
        diagnostics=diagnostics,
        rendered_context=filtered_context,
    )


def log_prompt_request_completion(
    prompt_diagnostics: Optional[Dict[str, Any]],
    *,
    result: Optional[Dict[str, Any]] = None,
    error: Optional[str] = None,
) -> None:
    if not isinstance(prompt_diagnostics, dict) or not prompt_diagnostics:
        return

    success = bool(isinstance(result, dict) and result.get("success"))
    response_text = ""
    if isinstance(result, dict):
        response_text = _safe_str(result.get("response") or result.get("content"))

    record = {
        "event": "result",
        "request_id": prompt_diagnostics.get("request_id"),
        "feature_id": prompt_diagnostics.get("feature_id"),
        "template_name": prompt_diagnostics.get("template_name"),
        "template_path": prompt_diagnostics.get("template_path"),
        "mode": prompt_diagnostics.get("mode"),
        "success": success,
        "provider": (result or {}).get("provider") if isinstance(result, dict) else None,
        "model": (result or {}).get("model") if isinstance(result, dict) else None,
        "elapsed_seconds": (result or {}).get("elapsed_seconds") if isinstance(result, dict) else None,
        "started_at": (result or {}).get("started_at") if isinstance(result, dict) else None,
        "finished_at": (result or {}).get("finished_at") if isinstance(result, dict) else None,
        "error": error or ((result or {}).get("error") if isinstance(result, dict) else None),
        "response_chars": len(response_text),
        "response_token_estimate": estimate_token_count(response_text),
    }

    logger.info(
        "prompt request completed feature=%s template=%s mode=%s success=%s prompt_chars=%s response_chars=%s elapsed=%s",
        record.get("feature_id"),
        record.get("template_name"),
        record.get("mode"),
        record.get("success"),
        prompt_diagnostics.get("prompt_chars"),
        record.get("response_chars"),
        record.get("elapsed_seconds"),
    )
    _append_event(record)
    _append_comparison_report(prompt_diagnostics, record)
    _save_response_artifacts_if_needed(prompt_diagnostics, result=result, error=error)


def _normalize_feature_config_inplace(feature_config: Dict[str, Any], base_config: Dict[str, Any]) -> None:
    if feature_config.get("mode") not in {PROMPT_MODE_FULL_EMBED, PROMPT_MODE_FILTERED_EMBED, None, ""}:
        feature_config["mode"] = base_config.get("default_mode", PROMPT_MODE_FULL_EMBED)
    if not feature_config.get("mode"):
        feature_config["mode"] = base_config.get("default_mode", PROMPT_MODE_FULL_EMBED)
    if feature_config.get("fallback_behavior") not in {
        PROMPT_FALLBACK_FULL_EMBED,
        PROMPT_FALLBACK_KEEP_FILTERED,
        None,
        "",
    }:
        feature_config["fallback_behavior"] = base_config.get("fallback_behavior", PROMPT_FALLBACK_FULL_EMBED)
    if not feature_config.get("fallback_behavior"):
        feature_config["fallback_behavior"] = base_config.get("fallback_behavior", PROMPT_FALLBACK_FULL_EMBED)

    for key in ("min_candidates", "max_candidates", "max_chars", "max_token_estimate", "low_score_threshold"):
        if key in feature_config:
            try:
                feature_config[key] = int(feature_config[key])
            except Exception:
                feature_config.pop(key, None)

    if "query_source_weights" in feature_config:
        feature_config["query_source_weights"] = _normalize_weight_map(
            feature_config.get("query_source_weights"),
            base_config.get("query_source_weights") or _DEFAULT_QUERY_SOURCE_WEIGHTS,
            numeric_type=float,
        )
    if "query_source_token_limits" in feature_config:
        feature_config["query_source_token_limits"] = _normalize_weight_map(
            feature_config.get("query_source_token_limits"),
            base_config.get("query_source_token_limits") or _DEFAULT_QUERY_SOURCE_TOKEN_LIMITS,
            numeric_type=int,
            minimum=1,
        )

    feature_config["debug_save_enabled"] = bool(
        feature_config.get("debug_save_enabled", base_config.get("debug_save_enabled", False))
    )
    if "comparison_report_enabled" in feature_config:
        feature_config["comparison_report_enabled"] = bool(feature_config.get("comparison_report_enabled"))
    sources = feature_config.get("sources")
    if not isinstance(sources, dict):
        feature_config["sources"] = {}
    else:
        normalized_sources = {}
        for placeholder, source_override in sources.items():
            if not isinstance(source_override, dict):
                continue
            cloned_override = dict(source_override)
            for key in ("min_candidates", "max_candidates", "max_chars", "max_token_estimate", "low_score_threshold"):
                if key in cloned_override:
                    try:
                        cloned_override[key] = int(cloned_override[key])
                    except Exception:
                        cloned_override.pop(key, None)
            normalized_sources[placeholder] = cloned_override
        feature_config["sources"] = normalized_sources


def _normalize_weight_map(
    value: Any,
    defaults: Dict[str, Any],
    *,
    numeric_type,
    minimum: Optional[float] = None,
) -> Dict[str, Any]:
    base = dict(defaults or {})
    if not isinstance(value, dict):
        return base
    for key, default_value in defaults.items():
        raw_value = value.get(key, default_value)
        try:
            converted = numeric_type(raw_value)
            if minimum is not None:
                converted = max(minimum, converted)
            base[key] = converted
        except Exception:
            base[key] = default_value
    return base


def _resolve_prompt_assembly_config(
    ai_config: Optional[Dict[str, Any]],
    feature_id: str,
    template_name: str,
) -> Dict[str, Any]:
    config = dict(ai_config or {})
    normalize_prompt_assembly_config_inplace(config)
    prompt_assembly = dict(config.get("prompt_assembly") or {})
    features = prompt_assembly.get("features") or {}

    resolved = {
        "mode": prompt_assembly.get("default_mode", PROMPT_MODE_FULL_EMBED),
        "fallback_behavior": prompt_assembly.get("fallback_behavior", PROMPT_FALLBACK_FULL_EMBED),
        "min_candidates": prompt_assembly.get("min_candidates", 3),
        "max_candidates": prompt_assembly.get("max_candidates", 20),
        "max_chars": prompt_assembly.get("max_chars", 12000),
        "max_token_estimate": prompt_assembly.get("max_token_estimate", 3000),
        "low_score_threshold": prompt_assembly.get("low_score_threshold", 0),
        "query_source_weights": dict(prompt_assembly.get("query_source_weights") or _DEFAULT_QUERY_SOURCE_WEIGHTS),
        "query_source_token_limits": dict(prompt_assembly.get("query_source_token_limits") or _DEFAULT_QUERY_SOURCE_TOKEN_LIMITS),
        "comparison_report_enabled": bool(prompt_assembly.get("comparison_report_enabled", True)),
        "debug_save_enabled": bool(prompt_assembly.get("debug_save_enabled", False)),
        "sources": {},
    }

    for key in [feature_id, template_name]:
        if not key:
            continue
        feature_override = features.get(key)
        if not isinstance(feature_override, dict):
            continue
        _normalize_feature_config_inplace(feature_override, prompt_assembly)
        for field in (
            "mode",
            "fallback_behavior",
            "min_candidates",
            "max_candidates",
            "max_chars",
            "max_token_estimate",
            "low_score_threshold",
            "comparison_report_enabled",
            "debug_save_enabled",
        ):
            if field in feature_override:
                resolved[field] = feature_override[field]
        if "query_source_weights" in feature_override:
            resolved["query_source_weights"] = _normalize_weight_map(
                feature_override.get("query_source_weights"),
                resolved.get("query_source_weights") or _DEFAULT_QUERY_SOURCE_WEIGHTS,
                numeric_type=float,
            )
        if "query_source_token_limits" in feature_override:
            resolved["query_source_token_limits"] = _normalize_weight_map(
                feature_override.get("query_source_token_limits"),
                resolved.get("query_source_token_limits") or _DEFAULT_QUERY_SOURCE_TOKEN_LIMITS,
                numeric_type=int,
                minimum=1,
            )
        source_overrides = feature_override.get("sources")
        if isinstance(source_overrides, dict):
            merged_sources = dict(resolved.get("sources") or {})
            for placeholder, source_override in source_overrides.items():
                if not isinstance(source_override, dict):
                    continue
                merged_sources[placeholder] = {**merged_sources.get(placeholder, {}), **source_override}
            resolved["sources"] = merged_sources

    return resolved


def _resolve_placeholder_value(
    *,
    placeholder: str,
    source_type: str,
    raw_value: str,
    context_data: Dict[str, Any],
    assembly_config: Dict[str, Any],
    alias_config_override: Optional[Dict[str, Any]] = None,
) -> Tuple[str, Dict[str, Any]]:
    started = time.perf_counter()
    source_config = dict(assembly_config)
    source_overrides = (assembly_config.get("sources") or {}).get(placeholder)
    if isinstance(source_overrides, dict):
        source_config.update(source_overrides)

    original_chars = len(raw_value)
    original_token_estimate = estimate_token_count(raw_value)
    mode = source_config.get("mode", PROMPT_MODE_FULL_EMBED)
    if mode == PROMPT_MODE_FULL_EMBED:
        summary = f"{placeholder}: full {original_chars} chars"
        return raw_value, {
            "placeholder": placeholder,
            "mode": PROMPT_MODE_FULL_EMBED,
            "original_chars": original_chars,
            "original_token_estimate": original_token_estimate,
            "candidate_pool_size": None,
            "selected_candidates": None,
            "filtered_chars": original_chars,
            "filtered_token_estimate": original_token_estimate,
            "fallback_used": False,
            "trimmed": False,
            "query_preview": "",
            "query_terms": [],
            "top_matches": [],
            "summary": summary,
            "elapsed_seconds": round(time.perf_counter() - started, 3),
        }

    entries, formatter_name = _parse_entries_for_placeholder(placeholder, source_type, raw_value)
    alias_index = _get_alias_index(alias_config_override)
    query_package = _build_weighted_query_package(
        context_data,
        assembly_config,
        placeholder=placeholder,
        source_type=source_type,
        alias_index=alias_index,
    )

    scored_entries = []
    for entry in entries:
        score, breakdown, matched_tokens = _score_entry(
            entry,
            query_package,
            placeholder=placeholder,
            source_type=source_type,
            alias_index=alias_index,
        )
        if score > 0:
            scored_entries.append(
                {
                    "score": score,
                    "entry": entry,
                    "breakdown": breakdown,
                    "matched_tokens": matched_tokens,
                }
            )

    scored_entries.sort(
        key=lambda item: (
            item["score"],
            len(item["entry"].get("path") or []),
            item["entry"].get("label") or "",
        ),
        reverse=True,
    )

    max_candidates = max(1, int(source_config.get("max_candidates", 20)))
    min_candidates = max(1, int(source_config.get("min_candidates", 3)))
    fallback_behavior = source_config.get("fallback_behavior", PROMPT_FALLBACK_FULL_EMBED)
    low_score_threshold = max(0, int(source_config.get("low_score_threshold", 0)))
    selected_scored_entries = scored_entries[:max_candidates]
    selected_entries = [item["entry"] for item in selected_scored_entries]
    fallback_used = False
    fallback_reason = ""

    top_score = selected_scored_entries[0]["score"] if selected_scored_entries else 0
    if not selected_entries:
        fallback_reason = "no_candidates"
    elif top_score < low_score_threshold:
        fallback_reason = "low_score_confidence"
    elif len(selected_entries) < min_candidates:
        fallback_reason = "too_few_candidates"

    if fallback_reason:
        if fallback_behavior == PROMPT_FALLBACK_FULL_EMBED:
            fallback_used = True
            summary = f"{placeholder}: filtered fallback to full ({fallback_reason})"
            return raw_value, {
                "placeholder": placeholder,
                "mode": PROMPT_MODE_FILTERED_EMBED,
                "original_chars": original_chars,
                "original_token_estimate": original_token_estimate,
                "candidate_pool_size": len(entries),
                "selected_candidates": len(selected_entries),
                "filtered_chars": original_chars,
                "filtered_token_estimate": original_token_estimate,
                "fallback_used": True,
                "fallback_reason": fallback_reason,
                "trimmed": False,
                "query_preview": _truncate_diagnostic_text(query_package.get("query_text", ""), 500),
                "query_terms": _summarize_weighted_tokens(query_package, limit=20),
                "alias_expanded_tokens": query_package.get("alias_expansions", [])[:20],
                "ignored_noise_tokens": query_package.get("ignored_tokens", [])[:20],
                "query_context_by_source": query_package.get("sources", []),
                "top_matches": _build_match_preview(scored_entries, max_items=8),
                "selected_summary": _build_selected_summary(selected_entries),
                "summary": summary,
                "elapsed_seconds": round(time.perf_counter() - started, 3),
            }

    filtered_text, trimmed, trimmed_count = _format_filtered_entries(
        formatter_name=formatter_name,
        original_text=raw_value,
        selected_entries=selected_entries,
        max_chars=max(500, int(source_config.get("max_chars", 12000))),
        max_token_estimate=max(100, int(source_config.get("max_token_estimate", 3000))),
    )

    summary = (
        f"{placeholder}: filtered {len(selected_entries)}/{len(entries)} candidates, "
        f"{len(filtered_text)} chars"
    )

    if not filtered_text.strip() and selected_entries and fallback_behavior == PROMPT_FALLBACK_FULL_EMBED:
        fallback_used = True
        fallback_reason = "truncation_failure"
        filtered_text = raw_value

    return filtered_text, {
        "placeholder": placeholder,
        "mode": PROMPT_MODE_FILTERED_EMBED,
        "original_chars": original_chars,
        "original_token_estimate": original_token_estimate,
        "candidate_pool_size": len(entries),
        "selected_candidates": len(selected_entries),
        "filtered_chars": len(filtered_text),
        "filtered_token_estimate": estimate_token_count(filtered_text),
        "fallback_used": fallback_used,
        "fallback_reason": fallback_reason or None,
        "trimmed": trimmed,
        "trimmed_candidates": trimmed_count,
        "query_preview": _truncate_diagnostic_text(query_package.get("query_text", ""), 500),
        "query_terms": _summarize_weighted_tokens(query_package, limit=20),
        "alias_expanded_tokens": query_package.get("alias_expansions", [])[:20],
        "ignored_noise_tokens": query_package.get("ignored_tokens", [])[:20],
        "query_context_by_source": query_package.get("sources", []),
        "top_matches": _build_match_preview(scored_entries, max_items=8),
        "selected_summary": _build_selected_summary(selected_entries),
        "candidate_pool_summary": {
            "raw_candidate_count": len(entries),
            "scored_candidate_count": len(scored_entries),
            "selected_candidate_count": len(selected_entries),
            "top_score": top_score,
            "low_score_threshold": low_score_threshold,
        },
        "injected_block_preview": _truncate_diagnostic_text(filtered_text, 600),
        "summary": summary,
        "elapsed_seconds": round(time.perf_counter() - started, 3),
    }


def _parse_entries_for_placeholder(placeholder: str, source_type: str, raw_value: str) -> Tuple[List[Dict[str, Any]], str]:
    if source_type == "dataportal_flat":
        return _parse_flat_master_entries(raw_value), "flat"
    return _parse_material_index_entries(raw_value), "material_index"


def _parse_flat_master_entries(raw_value: str) -> List[Dict[str, Any]]:
    try:
        payload = json.loads(raw_value)
    except Exception:
        return []

    data = payload.get("data") if isinstance(payload, dict) else payload
    if not isinstance(data, dict):
        return []

    entries = []
    for key, value in data.items():
        label = _safe_str(value).strip()
        entry_id = _safe_str(key).strip()
        if not label:
            continue
        entries.append(
            {
                "id": entry_id,
                "label": label,
                "path": [label],
                "search_text": _normalize_text(f"{entry_id} {label}"),
                "original_value": value,
            }
        )
    return entries


def _parse_material_index_entries(raw_value: str) -> List[Dict[str, Any]]:
    try:
        payload = json.loads(raw_value)
    except Exception:
        return []

    entries: List[Dict[str, Any]] = []

    def visit(node: Any, path: List[str]) -> None:
        if isinstance(node, dict):
            for key, value in node.items():
                visit(value, path + [_safe_str(key)])
            return
        if isinstance(node, list):
            for item in node:
                if isinstance(item, (dict, list)):
                    visit(item, path)
                else:
                    label = _safe_str(item).strip()
                    if not label:
                        continue
                    full_path = path + [label]
                    entries.append(
                        {
                            "id": " > ".join(full_path),
                            "label": label,
                            "path": full_path,
                            "search_text": _normalize_text(" ".join(full_path)),
                            "original_value": item,
                        }
                    )
            return
        label = _safe_str(node).strip()
        if not label:
            return
        full_path = path + [label]
        entries.append(
            {
                "id": " > ".join(full_path),
                "label": label,
                "path": full_path,
                "search_text": _normalize_text(" ".join(full_path)),
                "original_value": node,
            }
        )

    visit(payload, [])
    return entries


def _format_filtered_entries(
    *,
    formatter_name: str,
    original_text: str,
    selected_entries: List[Dict[str, Any]],
    max_chars: int,
    max_token_estimate: int,
) -> Tuple[str, bool, int]:
    if formatter_name == "flat":
        formatter = _build_flat_subset
    else:
        formatter = _build_material_index_subset

    working_entries = list(selected_entries)
    trimmed = False
    trimmed_count = 0
    if not working_entries:
        return formatter(original_text, []), False, 0

    while working_entries:
        rendered = formatter(original_text, working_entries)
        if len(rendered) <= max_chars and estimate_token_count(rendered) <= max_token_estimate:
            return rendered, trimmed, trimmed_count
        working_entries.pop()
        trimmed = True
        trimmed_count += 1

    return formatter(original_text, []), trimmed, trimmed_count


def _build_flat_subset(original_text: str, selected_entries: List[Dict[str, Any]]) -> str:
    try:
        payload = json.loads(original_text)
    except Exception:
        return "{}"

    if not isinstance(payload, dict):
        return json.dumps({}, ensure_ascii=False, indent=2)

    selected_map = {}
    for entry in selected_entries:
        entry_id = entry.get("id")
        if entry_id is None:
            continue
        selected_map[_safe_str(entry_id)] = entry.get("original_value")
    new_payload = dict(payload)
    new_payload["data"] = selected_map
    new_payload["count"] = len(selected_map)
    return json.dumps(new_payload, ensure_ascii=False, indent=2)


def _build_material_index_subset(original_text: str, selected_entries: List[Dict[str, Any]]) -> str:
    subset: Dict[str, Any] = {}
    for entry in selected_entries:
        path = list(entry.get("path") or [])
        if not path:
            continue
        cursor = subset
        for index, component in enumerate(path):
            if index == len(path) - 2:
                key = component
                leaf = path[index + 1]
                cursor.setdefault(key, [])
                if isinstance(cursor[key], list) and leaf not in cursor[key]:
                    cursor[key].append(leaf)
                break
            if index >= len(path) - 1:
                break
            cursor = cursor.setdefault(component, {})
            if not isinstance(cursor, dict):
                break
    return json.dumps(subset, ensure_ascii=False, indent=2)


def _build_query_text(context_data: Dict[str, Any]) -> str:
    return "\n".join(source.get("text", "") for source in _collect_query_sources(context_data, default_prompt_assembly_config()))


def _score_entry(
    entry: Dict[str, Any],
    query_package: Dict[str, Any],
    *,
    placeholder: str,
    source_type: str,
    alias_index: Dict[str, Any],
) -> Tuple[int, Dict[str, int], List[Dict[str, Any]]]:
    base_terms, alias_terms = _build_entry_term_sets(entry, placeholder, source_type, alias_index)
    search_text = entry.get("search_text") or ""
    path_components = [_normalize_text(component) for component in (entry.get("path") or []) if _normalize_text(component)]
    weighted_tokens = query_package.get("weighted_tokens") or {}
    query_norm = query_package.get("query_norm") or ""

    breakdown = {
        "exact_match": 0,
        "normalized_exact_match": 0,
        "alias_match": 0,
        "substring_match": 0,
        "token_overlap": 0,
        "weighted_source_overlap": 0,
        "path_bonus": 0,
        "parent_child_bonus": 0,
        "description_hit": 0,
        "repeated_token_damping": 0,
        "noise_penalty": 0,
    }
    matched_tokens: List[Dict[str, Any]] = []
    matched_path_indices = []
    strong_source_hit = False
    file_tree_only = True

    if search_text and search_text in query_norm:
        breakdown["normalized_exact_match"] += 120

    for index, component in enumerate(path_components):
        if component and component in query_norm:
            breakdown["path_bonus"] += 18 + min(len(component), 20)
            matched_path_indices.append(index)

    for token, token_info in weighted_tokens.items():
        token_weight = float(token_info.get("weight") or 0.0)
        if token_weight <= 0:
            continue
        source_categories = set(token_info.get("source_categories") or [])
        if source_categories - {"file_tree"}:
            file_tree_only = False
        if source_categories & {"title", "description", "summary"}:
            strong_source_hit = True

        if token in base_terms:
            breakdown["token_overlap"] += min(60, 8 + int(token_weight * 10))
            breakdown["weighted_source_overlap"] += int(token_weight * 6)
            if source_categories & {"title", "description", "summary"}:
                breakdown["description_hit"] += 6
            matched_tokens.append(
                {
                    "token": token,
                    "weight": round(token_weight, 3),
                    "kind": "direct",
                    "source_categories": sorted(source_categories),
                }
            )
        elif token in alias_terms:
            breakdown["alias_match"] += min(52, 6 + int(token_weight * 9))
            breakdown["weighted_source_overlap"] += int(token_weight * 5)
            matched_tokens.append(
                {
                    "token": token,
                    "weight": round(token_weight, 3),
                    "kind": "alias",
                    "source_categories": sorted(source_categories),
                }
            )
        elif len(token) >= 3 and any(token in term or term in token for term in list(base_terms)[:48]):
            breakdown["substring_match"] += min(18, 3 + int(token_weight * 4))

        damping_penalty = float(token_info.get("damping_penalty") or 0.0)
        if damping_penalty > 0 and (token in base_terms or token in alias_terms):
            breakdown["repeated_token_damping"] -= min(12, int(math.ceil(damping_penalty * 4)))

    if len(matched_path_indices) >= 2:
        consecutive_matches = 0
        for prev_idx, next_idx in zip(matched_path_indices, matched_path_indices[1:]):
            if next_idx == prev_idx + 1:
                consecutive_matches += 1
        if consecutive_matches:
            breakdown["parent_child_bonus"] += 16 * consecutive_matches

    if matched_tokens and file_tree_only and not strong_source_hit:
        breakdown["noise_penalty"] -= 12

    total_score = sum(int(value) for value in breakdown.values())
    return total_score, breakdown, matched_tokens[:16]


def _collect_query_sources(context_data: Dict[str, Any], assembly_config: Dict[str, Any]) -> List[Dict[str, Any]]:
    skip_keys = set(_LARGE_PLACEHOLDER_SOURCES) | {
        "llm_provider",
        "llm_model",
        "llm_model_name",
        "dataset_id",
    }
    weights = assembly_config.get("query_source_weights") or _DEFAULT_QUERY_SOURCE_WEIGHTS
    token_limits = assembly_config.get("query_source_token_limits") or _DEFAULT_QUERY_SOURCE_TOKEN_LIMITS
    sources = []
    for key, value in (context_data or {}).items():
        if key in skip_keys:
            continue
        text = _safe_str(value).strip()
        if not text:
            continue
        if text.startswith("[") and text.endswith("未設定]"):
            continue
        category = _classify_context_source(key)
        weight = float(weights.get(category, 1.0))
        if weight <= 0:
            continue
        sources.append(
            {
                "source_key": key,
                "category": category,
                "weight": weight,
                "token_limit": int(token_limits.get(category, 20)),
                "text": text[:4000],
            }
        )
    return sources


def _classify_context_source(key: str) -> str:
    normalized_key = (key or "").strip().lower()
    if normalized_key in {"name", "title", "dataset_name", "item_name"}:
        return "title"
    if normalized_key in {"existing_description", "description", "abstract"}:
        return "description"
    if normalized_key in {"experiment_summary", "dataset_existing_info", "summary"}:
        return "summary"
    if normalized_key in {"file_tree"}:
        return "file_tree"
    if normalized_key in {"text_from_structured_files", "json_from_structured_files", "arim_extension_data", "arim_experiment_data", "arim_detailed_experiment"}:
        return "metadata"
    if normalized_key.startswith("report_") or normalized_key.startswith("arim_report_"):
        return "body"
    return "body"


def _build_weighted_query_package(
    context_data: Dict[str, Any],
    assembly_config: Dict[str, Any],
    *,
    placeholder: str,
    source_type: str,
    alias_index: Dict[str, Any],
) -> Dict[str, Any]:
    sources = _collect_query_sources(context_data, assembly_config)
    combined_text = "\n".join(source.get("text", "") for source in sources)
    weighted_tokens: Dict[str, Dict[str, Any]] = {}
    alias_expansions: List[Dict[str, Any]] = []
    ignored_tokens: List[Dict[str, Any]] = []
    source_summaries: List[Dict[str, Any]] = []

    for source in sources:
        normalized_text = _normalize_text(source.get("text", ""))
        raw_tokens = _extract_query_tokens(normalized_text)
        limited_tokens = raw_tokens[: max(1, int(source.get("token_limit") or 1))]
        source_kept_tokens = []
        for token in limited_tokens:
            ignore_reason = _classify_ignored_token_for_source(token, alias_index, placeholder, source_type)
            if ignore_reason:
                ignored_tokens.append({
                    "token": token,
                    "reason": ignore_reason,
                    "source_key": source.get("source_key"),
                    "category": source.get("category"),
                })
                continue
            source_kept_tokens.append(token)
            _register_weighted_token(weighted_tokens, token, source, alias_of="")
            expanded_aliases = sorted(_resolve_alias_terms(token, alias_index, placeholder, source_type) - {token})[:8]
            if expanded_aliases:
                alias_expansions.append(
                    {
                        "token": token,
                        "expanded": expanded_aliases,
                        "source_key": source.get("source_key"),
                        "category": source.get("category"),
                    }
                )
            for alias_token in expanded_aliases:
                _register_weighted_token(weighted_tokens, alias_token, source, alias_of=token, weight_scale=0.85)
        source_summaries.append(
            {
                "source_key": source.get("source_key"),
                "category": source.get("category"),
                "weight": source.get("weight"),
                "text_preview": _truncate_diagnostic_text(source.get("text", ""), 220),
                "tokens": source_kept_tokens[:16],
            }
        )

    return {
        "query_text": combined_text,
        "query_norm": _normalize_text(combined_text),
        "weighted_tokens": weighted_tokens,
        "alias_expansions": alias_expansions,
        "ignored_tokens": ignored_tokens,
        "sources": source_summaries,
    }


def _register_weighted_token(
    weighted_tokens: Dict[str, Dict[str, Any]],
    token: str,
    source: Dict[str, Any],
    *,
    alias_of: str,
    weight_scale: float = 1.0,
) -> None:
    normalized_token = _normalize_text(token)
    if not normalized_token:
        return
    current = weighted_tokens.setdefault(
        normalized_token,
        {
            "weight": 0.0,
            "occurrences": 0,
            "source_categories": set(),
            "source_keys": set(),
            "alias_of": set(),
            "damping_penalty": 0.0,
        },
    )
    base_weight = float(source.get("weight") or 1.0) * weight_scale
    damping_factor = 0.6 ** current["occurrences"]
    added_weight = base_weight * damping_factor
    current["weight"] += added_weight
    current["damping_penalty"] += max(0.0, base_weight - added_weight)
    current["occurrences"] += 1
    current["source_categories"].add(source.get("category"))
    current["source_keys"].add(source.get("source_key"))
    if alias_of:
        current["alias_of"].add(alias_of)


def _build_entry_term_sets(
    entry: Dict[str, Any],
    placeholder: str,
    source_type: str,
    alias_index: Dict[str, Any],
) -> Tuple[set, set]:
    base_terms = set()
    for component in list(entry.get("path") or []) + [entry.get("label") or "", entry.get("search_text") or ""]:
        normalized_component = _normalize_text(component)
        if not normalized_component:
            continue
        base_terms.add(normalized_component)
        for token in _extract_query_tokens(normalized_component):
            base_terms.add(token)

    alias_terms = set()
    for term in list(base_terms):
        alias_terms.update(_resolve_alias_terms(term, alias_index, placeholder, source_type))
    alias_terms -= base_terms
    return base_terms, alias_terms


def _summarize_weighted_tokens(query_package: Dict[str, Any], limit: int = 20) -> List[str]:
    weighted_tokens = query_package.get("weighted_tokens") or {}
    ranked = sorted(weighted_tokens.items(), key=lambda item: item[1].get("weight", 0.0), reverse=True)
    return [token for token, _info in ranked[:limit]]


def _extract_placeholders(text: str) -> List[str]:
    return list(dict.fromkeys(_PLACEHOLDER_PATTERN.findall(text or "")))


def _get_alias_index(alias_config_override: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    alias_config = _normalize_prompt_dictionary_config(alias_config_override) if alias_config_override is not None else _load_prompt_alias_config()
    global_index = _build_bidirectional_alias_index(alias_config.get("general_aliases") or {})
    source_indexes = {}
    for source_key, aliases in (alias_config.get("source_aliases") or {}).items():
        source_indexes[source_key] = _build_bidirectional_alias_index(aliases or {})
    source_stopwords = {}
    source_weak_stopwords = {}
    source_allowlist = {}
    for source_key, override in (alias_config.get("source_overrides") or {}).items():
        if not isinstance(override, dict):
            continue
        source_stopwords[source_key] = {
            _normalize_text(token)
            for token in override.get("stopwords") or []
            if _normalize_text(token)
        }
        source_weak_stopwords[source_key] = {
            _normalize_text(token)
            for token in override.get("weak_stopwords") or []
            if _normalize_text(token)
        }
        source_allowlist[source_key] = {
            _normalize_text(token)
            for token in override.get("allowlist") or []
            if _normalize_text(token)
        }
    for source_key, values in ((alias_config.get("allowlists") or {}).get("sources") or {}).items():
        normalized_values = {
            _normalize_text(token)
            for token in values or []
            if _normalize_text(token)
        }
        if normalized_values:
            source_allowlist.setdefault(source_key, set()).update(normalized_values)
    return {
        "global": global_index,
        "source": source_indexes,
        "stopwords": {_normalize_text(token) for token in alias_config.get("stopwords") or [] if _normalize_text(token)},
        "stopwords_ja": {_normalize_text(token) for token in alias_config.get("stopwords_ja") or [] if _normalize_text(token)},
        "weak_stopwords": {_normalize_text(token) for token in alias_config.get("weak_stopwords") or [] if _normalize_text(token)},
        "file_extensions": {_normalize_text(token) for token in alias_config.get("file_extensions") or [] if _normalize_text(token)},
        "allowlist": {
            _normalize_text(token)
            for token in list(alias_config.get("allowlist") or []) + list(((alias_config.get("allowlists") or {}).get("global") or []))
            if _normalize_text(token)
        },
        "source_stopwords": source_stopwords,
        "source_weak_stopwords": source_weak_stopwords,
        "source_allowlist": source_allowlist,
    }


def _load_prompt_alias_config() -> Dict[str, Any]:
    path = get_prompt_dictionary_config_path()
    mtime = os.path.getmtime(path) if os.path.exists(path) else None
    cached = _ALIAS_CONFIG_CACHE.get("value")
    if cached is not None and _ALIAS_CONFIG_CACHE.get("mtime") == mtime:
        return copy.deepcopy(cached)

    merged = {}
    loaded = None
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as handle:
                loaded = json.load(handle)
        except Exception:
            logger.warning("prompt alias config load failed: %s", path, exc_info=True)

    loaded_metadata = (loaded.get("metadata") or {}) if isinstance(loaded, dict) else {}
    current_seed_version = int(loaded_metadata.get("preset_seed_version") or 0)
    should_seed_presets = not isinstance(loaded, dict) or current_seed_version < _PROMPT_DICTIONARY_PRESET_SEED_VERSION
    if should_seed_presets:
        merged = _deep_merge_prompt_alias_config(merged, _DEFAULT_ALIAS_CONFIG)
        merged = _deep_merge_prompt_alias_config(merged, build_prompt_dictionary_preset_config())
    if isinstance(loaded, dict):
        merged = _deep_merge_prompt_alias_config(merged, loaded)
    if should_seed_presets:
        metadata = merged.setdefault("metadata", {})
        metadata["preset_seed_version"] = _PROMPT_DICTIONARY_PRESET_SEED_VERSION
        metadata.setdefault("preset_seed_name", "materials_device_aliases")

    merged = _normalize_prompt_dictionary_config(merged)

    _ALIAS_CONFIG_CACHE["mtime"] = mtime
    _ALIAS_CONFIG_CACHE["value"] = copy.deepcopy(merged)
    return merged


def _deep_merge_prompt_alias_config(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    merged = copy.deepcopy(base)
    for key, value in (override or {}).items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = _deep_merge_prompt_alias_config(merged[key], value)
        elif isinstance(value, list):
            merged[key] = list(value)
        else:
            merged[key] = value
    return merged


def _build_bidirectional_alias_index(aliases: Dict[str, Any]) -> Dict[str, set]:
    index: Dict[str, set] = {}
    for canonical, values in (aliases or {}).items():
        terms = {_normalize_text(canonical)}
        if isinstance(values, list):
            terms.update(_normalize_text(item) for item in values if _normalize_text(item))
        terms = {term for term in terms if term}
        for term in terms:
            index.setdefault(term, set()).update(terms)
    return index


def _resolve_alias_terms(token: str, alias_index: Dict[str, Any], placeholder: str, source_type: str) -> set:
    normalized_token = _normalize_text(token)
    if not normalized_token:
        return set()
    resolved = {normalized_token}
    resolved.update(alias_index.get("global", {}).get(normalized_token, set()))
    for source_key in (placeholder, source_type):
        if not source_key:
            continue
        resolved.update((alias_index.get("source", {}).get(source_key, {}) or {}).get(normalized_token, set()))
    return {term for term in resolved if term}


def _classify_ignored_token(token: str, alias_index: Dict[str, Any]) -> str:
    return _classify_ignored_token_for_source(token, alias_index, "", "")


def _classify_ignored_token_for_source(token: str, alias_index: Dict[str, Any], placeholder: str, source_type: str) -> str:
    normalized_token = _normalize_text(token)
    if not normalized_token:
        return "empty"
    contextual_allowlist = set(alias_index.get("allowlist", set()))
    contextual_stopwords = set(alias_index.get("stopwords", set())) | set(alias_index.get("stopwords_ja", set()))
    contextual_weak_stopwords = set(alias_index.get("weak_stopwords", set()))
    for source_key in (placeholder, source_type):
        if not source_key:
            continue
        contextual_allowlist.update(alias_index.get("source_allowlist", {}).get(source_key, set()))
        contextual_stopwords.update(alias_index.get("source_stopwords", {}).get(source_key, set()))
        contextual_weak_stopwords.update(alias_index.get("source_weak_stopwords", {}).get(source_key, set()))

    if normalized_token in contextual_allowlist:
        return ""
    if normalized_token in contextual_stopwords:
        return "stopword"
    if normalized_token in contextual_weak_stopwords:
        return "weak_stopword"
    if normalized_token in alias_index.get("file_extensions", set()):
        return "file_extension"
    if len(normalized_token) <= 1:
        return "too_short"
    if len(normalized_token) <= 2 and not any(ch.isdigit() for ch in normalized_token) and normalized_token not in contextual_allowlist:
        return "too_short"
    return ""


def get_prompt_dictionary_config_path() -> str:
    return get_dynamic_file_path("input/ai/prompt_assembly_aliases.json")


def get_prompt_dictionary_summary_path() -> str:
    config_path = get_prompt_dictionary_config_path()
    root, _ext = os.path.splitext(config_path)
    return f"{root}.summary.json"


def load_prompt_dictionary_config() -> Dict[str, Any]:
    return _load_prompt_alias_config()


def save_prompt_dictionary_config(config: Dict[str, Any]) -> bool:
    try:
        normalized = _normalize_prompt_dictionary_config(config)
        path = get_prompt_dictionary_config_path()
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as handle:
            json.dump(normalized, handle, ensure_ascii=False, indent=2)
        _save_prompt_dictionary_summary(normalized)
        _ALIAS_CONFIG_CACHE["mtime"] = None
        _ALIAS_CONFIG_CACHE["value"] = None
        return True
    except Exception:
        logger.warning("prompt dictionary config save failed", exc_info=True)
        return False


def clear_prompt_dictionary_caches(*, remove_summary_file: bool = False) -> None:
    """Clear in-memory prompt dictionary caches and optional summary cache file."""

    _ALIAS_CONFIG_CACHE["mtime"] = None
    _ALIAS_CONFIG_CACHE["value"] = None
    _PROMPT_DICTIONARY_SUMMARY_CACHE["mtime"] = None
    _PROMPT_DICTIONARY_SUMMARY_CACHE["source_mtime"] = None
    _PROMPT_DICTIONARY_SUMMARY_CACHE["value"] = None
    if remove_summary_file:
        try:
            summary_path = get_prompt_dictionary_summary_path()
            if os.path.exists(summary_path):
                os.remove(summary_path)
        except Exception:
            logger.debug("prompt dictionary summary cache removal failed", exc_info=True)


def get_prompt_dictionary_cache_metadata() -> Dict[str, Any]:
    """Return metadata for prompt dictionary cache summary and in-memory snapshots."""

    summary_path = get_prompt_dictionary_summary_path()
    summary = _load_prompt_dictionary_summary() or {}
    size_bytes = 0
    updated_at = None
    try:
        if os.path.exists(summary_path):
            stat = os.stat(summary_path)
            size_bytes = int(stat.st_size)
            updated_at = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)
    except Exception:
        pass

    in_memory_active = bool(
        _ALIAS_CONFIG_CACHE.get("value") is not None
        or _PROMPT_DICTIONARY_SUMMARY_CACHE.get("value") is not None
    )
    item_count = 0
    try:
        item_count = int(summary.get("candidate_count") or 0) + int(summary.get("general_alias_count") or 0)
    except Exception:
        item_count = 0
    return {
        "path": summary_path,
        "item_count": item_count,
        "size_bytes": size_bytes,
        "updated_at": updated_at,
        "active": bool(in_memory_active or os.path.exists(summary_path)),
    }


def get_prompt_dictionary_summary(config: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    if config is None:
        summary = _load_prompt_dictionary_summary()
        if isinstance(summary, dict):
            return summary
    payload = _normalize_prompt_dictionary_config(config or _load_prompt_alias_config())
    summary = _build_prompt_dictionary_summary(payload)
    if config is None:
        _save_prompt_dictionary_summary(payload)
    return summary


def _build_prompt_dictionary_summary(payload: Dict[str, Any]) -> Dict[str, Any]:
    candidates = list(((payload.get("generated") or {}).get("candidates") or []))
    term_inventory = list(((payload.get("generated") or {}).get("term_inventory") or []))
    counts = Counter(str(candidate.get("status") or "pending") for candidate in candidates)
    kind_counts = Counter(str(candidate.get("kind") or "") for candidate in candidates)
    return {
        "general_alias_count": len(payload.get("general_aliases") or {}),
        "source_alias_count": sum(len(values or {}) for values in (payload.get("source_aliases") or {}).values()),
        "stopword_count": len(payload.get("stopwords") or []),
        "weak_stopword_count": len(payload.get("weak_stopwords") or []),
        "allowlist_count": len(payload.get("allowlist") or []) + len(((payload.get("allowlists") or {}).get("global") or [])),
        "candidate_count": len(candidates),
        "extracted_term_count": len(term_inventory),
        "status_counts": dict(counts),
        "kind_counts": dict(kind_counts),
        "scan_result_count": len(((payload.get("generated") or {}).get("scan_results") or [])),
        "last_scanned_at": ((payload.get("metadata") or {}).get("last_scanned_at") or ""),
        "last_applied_at": ((payload.get("metadata") or {}).get("last_applied_at") or ""),
        "web_enabled": bool((payload.get("assist") or {}).get("web_enabled", False)),
        "llm_enabled": bool((payload.get("assist") or {}).get("llm_enabled", False)),
        "auto_approved_candidate_count": sum(1 for candidate in candidates if candidate.get("status") in {"approved", "applied"} and candidate.get("reason", "").find("auto_approved") >= 0),
    }


def _load_prompt_dictionary_summary() -> Optional[Dict[str, Any]]:
    summary_path = get_prompt_dictionary_summary_path()
    config_path = get_prompt_dictionary_config_path()
    if not os.path.exists(summary_path):
        return None
    try:
        summary_mtime = os.path.getmtime(summary_path)
        config_mtime = os.path.getmtime(config_path) if os.path.exists(config_path) else None
        cached = _PROMPT_DICTIONARY_SUMMARY_CACHE.get("value")
        if (
            cached is not None
            and _PROMPT_DICTIONARY_SUMMARY_CACHE.get("mtime") == summary_mtime
            and _PROMPT_DICTIONARY_SUMMARY_CACHE.get("source_mtime") == config_mtime
        ):
            return dict(cached)
        if config_mtime is not None and summary_mtime < config_mtime:
            return None
        with open(summary_path, "r", encoding="utf-8") as handle:
            loaded = json.load(handle)
        if not isinstance(loaded, dict):
            return None
        _PROMPT_DICTIONARY_SUMMARY_CACHE["mtime"] = summary_mtime
        _PROMPT_DICTIONARY_SUMMARY_CACHE["source_mtime"] = config_mtime
        _PROMPT_DICTIONARY_SUMMARY_CACHE["value"] = dict(loaded)
        return dict(loaded)
    except Exception:
        logger.warning("prompt dictionary summary load failed", exc_info=True)
        return None


def _save_prompt_dictionary_summary(payload: Dict[str, Any]) -> None:
    summary = _build_prompt_dictionary_summary(payload)
    summary_path = get_prompt_dictionary_summary_path()
    try:
        os.makedirs(os.path.dirname(summary_path), exist_ok=True)
        with open(summary_path, "w", encoding="utf-8") as handle:
            json.dump(summary, handle, ensure_ascii=False, indent=2)
        config_path = get_prompt_dictionary_config_path()
        _PROMPT_DICTIONARY_SUMMARY_CACHE["mtime"] = os.path.getmtime(summary_path)
        _PROMPT_DICTIONARY_SUMMARY_CACHE["source_mtime"] = os.path.getmtime(config_path) if os.path.exists(config_path) else None
        _PROMPT_DICTIONARY_SUMMARY_CACHE["value"] = dict(summary)
    except Exception:
        logger.warning("prompt dictionary summary save failed", exc_info=True)


def build_prompt_dictionary_merge_preview(config: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    payload = _normalize_prompt_dictionary_config(config or _load_prompt_alias_config())
    preview = {
        "general_aliases": {},
        "source_overrides": {},
        "stopwords": [],
        "weak_stopwords": [],
        "allowlist": [],
    }
    for candidate in ((payload.get("generated") or {}).get("candidates") or []):
        if candidate.get("status") != "approved":
            continue
        _merge_candidate_into_preview(preview, candidate)
    return preview


def apply_prompt_dictionary_candidates(config: Optional[Dict[str, Any]] = None, candidate_ids: Optional[List[str]] = None) -> Dict[str, Any]:
    payload = _normalize_prompt_dictionary_config(config or _load_prompt_alias_config())
    candidate_id_set = set(candidate_ids or [])
    for candidate in ((payload.get("generated") or {}).get("candidates") or []):
        if candidate.get("status") != "approved":
            continue
        if candidate_id_set and candidate.get("id") not in candidate_id_set:
            continue
        _apply_candidate_to_config(payload, candidate)
        candidate["status"] = "applied"
        candidate["updated_at"] = _utc_now_iso()
    payload.setdefault("metadata", {})["last_applied_at"] = _utc_now_iso()
    payload.setdefault("generated", {})["evaluation_report"] = _build_prompt_dictionary_evaluation_report(payload)
    return payload


def scan_prompt_dictionary_outputs(
    config: Optional[Dict[str, Any]] = None,
    max_files_per_source: int = 40,
    *,
    web_fetcher=None,
    llm_expander=None,
    progress_callback=None,
) -> Dict[str, Any]:
    payload = _normalize_prompt_dictionary_config(config or _load_prompt_alias_config())
    scan_results: List[Dict[str, Any]] = []
    candidates: Dict[str, Dict[str, Any]] = {
        candidate.get("id"): copy.deepcopy(candidate)
        for candidate in ((payload.get("generated") or {}).get("candidates") or [])
        if isinstance(candidate, dict) and candidate.get("id")
    }
    observed_generic_tokens: Counter = Counter()
    term_inventory_map: Dict[str, Dict[str, Any]] = {}

    _emit_prompt_dictionary_progress(
        progress_callback,
        phase="build_start",
        message="辞書作成を開始しました",
        source_count=len(_PROMPT_DICTIONARY_SCAN_SOURCES) + len(_PROMPT_DICTIONARY_RECURSIVE_SCAN_TARGETS),
    )

    for source_name, source_path in _PROMPT_DICTIONARY_SCAN_SOURCES.items():
        abs_path = get_dynamic_file_path(source_path)
        _emit_prompt_dictionary_progress(
            progress_callback,
            phase="source_start",
            message=f"{source_name} を走査中",
            source=source_name,
            path=abs_path,
        )
        if source_name.startswith("summary_"):
            result = _scan_summary_workbook(abs_path, source_name)
        elif source_name == "data_entry":
            result = _scan_json_directory(abs_path, source_name, max_files_per_source, _extract_data_entry_texts, progress_callback=progress_callback)
        elif source_name == "invoice":
            result = _scan_json_directory(abs_path, source_name, max_files_per_source, _extract_invoice_texts, progress_callback=progress_callback)
        elif source_name == "search_results":
            result = _scan_json_directory(abs_path, source_name, min(max_files_per_source, 20), _extract_generic_json_texts, progress_callback=progress_callback)
        else:
            result = _scan_single_json(abs_path, source_name, _extract_generic_json_texts)
        for term_entry in result.pop("_term_inventory_raw", []):
            _merge_prompt_dictionary_term_inventory_entry(term_inventory_map, term_entry)
        scan_results.append(result)
        observed_generic_tokens.update(result.get("generic_tokens", {}))
        for candidate in result.get("candidates", []):
            existing = candidates.get(candidate.get("id"))
            if existing:
                candidate["status"] = existing.get("status") or candidate.get("status")
                candidate["created_at"] = existing.get("created_at") or candidate.get("created_at")
            candidates[candidate["id"]] = _merge_prompt_dictionary_candidate(existing, candidate)
        _emit_prompt_dictionary_progress(
            progress_callback,
            phase="source_complete",
            message=f"{source_name} の走査が完了しました",
            source=source_name,
            scanned_items=result.get("scanned_items", 0),
            text_items=result.get("text_items", 0),
            candidate_count=len(candidates),
            extracted_term_count=len(term_inventory_map),
        )

    if (payload.get("assist") or {}).get("web_enabled"):
        _emit_prompt_dictionary_progress(progress_callback, phase="web_start", message="WEB補助ソースを走査中")
        web_result = _scan_web_debug_directory(get_dynamic_file_path(_PROMPT_DICTIONARY_WEB_SCAN_SOURCE), max_files_per_source)
        for term_entry in web_result.pop("_term_inventory_raw", []):
            _merge_prompt_dictionary_term_inventory_entry(term_inventory_map, term_entry)
        scan_results.append(web_result)
        observed_generic_tokens.update(web_result.get("generic_tokens", {}))
        for candidate in web_result.get("candidates", []):
            existing = candidates.get(candidate.get("id"))
            candidates[candidate["id"]] = _merge_prompt_dictionary_candidate(existing, candidate)
        if callable(web_fetcher):
            for candidate in _normalize_external_candidates(web_fetcher(payload)):
                existing = candidates.get(candidate.get("id"))
                candidates[candidate["id"]] = _merge_prompt_dictionary_candidate(existing, candidate)

    for source_name, source_path in _PROMPT_DICTIONARY_RECURSIVE_SCAN_TARGETS.items():
        _emit_prompt_dictionary_progress(
            progress_callback,
            phase="source_start",
            message=f"{source_name} を再帰走査中",
            source=source_name,
            path=get_dynamic_file_path(source_path),
        )
        recursive_result = _scan_recursive_directory(
            get_dynamic_file_path(source_path),
            source_name,
            max_files=max_files_per_source,
            progress_callback=progress_callback,
        )
        for term_entry in recursive_result.pop("_term_inventory_raw", []):
            _merge_prompt_dictionary_term_inventory_entry(term_inventory_map, term_entry)
        scan_results.append(recursive_result)
        observed_generic_tokens.update(recursive_result.get("generic_tokens", {}))
        for candidate in recursive_result.get("candidates", []):
            existing = candidates.get(candidate.get("id"))
            candidates[candidate["id"]] = _merge_prompt_dictionary_candidate(existing, candidate)
        _emit_prompt_dictionary_progress(
            progress_callback,
            phase="source_complete",
            message=f"{source_name} の再帰走査が完了しました",
            source=source_name,
            scanned_items=recursive_result.get("scanned_items", 0),
            text_items=recursive_result.get("text_items", 0),
            candidate_count=len(candidates),
            extracted_term_count=len(term_inventory_map),
        )

    term_inventory = _materialize_prompt_dictionary_term_inventory(term_inventory_map)

    for candidate in _build_allowlist_candidates_from_term_inventory(term_inventory):
        existing = candidates.get(candidate.get("id"))
        if existing:
            candidate["status"] = existing.get("status") or candidate.get("status")
            candidate["created_at"] = existing.get("created_at") or candidate.get("created_at")
        candidates[candidate["id"]] = _merge_prompt_dictionary_candidate(existing, candidate)

    if (payload.get("assist") or {}).get("llm_enabled"):
        _emit_prompt_dictionary_progress(progress_callback, phase="llm_start", message="LLMで抽出語を整理中")
        llm_candidates = []
        if callable(llm_expander):
            llm_candidates = _normalize_external_candidates(llm_expander(payload, scan_results))
        else:
            llm_candidates = _run_llm_prompt_dictionary_review(payload, scan_results, term_inventory)
        for candidate in llm_candidates:
            existing = candidates.get(candidate.get("id"))
            if existing:
                candidate["status"] = existing.get("status") or candidate.get("status")
                candidate["created_at"] = existing.get("created_at") or candidate.get("created_at")
            candidates[candidate["id"]] = _merge_prompt_dictionary_candidate(existing, candidate)
        _emit_prompt_dictionary_progress(progress_callback, phase="llm_complete", message="LLM整理が完了しました", candidate_count=len(candidates))

    for token, occurrences in observed_generic_tokens.items():
        seed_kind = _classify_generic_stopword_seed(token)
        if not seed_kind:
            continue
        score = min(0.98, 0.35 + (occurrences * 0.05))
        candidate = _make_prompt_dictionary_candidate(
            kind=seed_kind,
            canonical="",
            value=token,
            source="observed_generic_tokens",
            score=score,
            reason=f"scan_frequency={occurrences}",
            examples=[token],
            occurrences=occurrences,
        )
        existing = candidates.get(candidate.get("id"))
        if existing:
            candidate["status"] = existing.get("status") or candidate.get("status")
            candidate["created_at"] = existing.get("created_at") or candidate.get("created_at")
        candidates[candidate["id"]] = _merge_prompt_dictionary_candidate(existing, candidate)

    payload.setdefault("generated", {})["term_inventory"] = term_inventory[:_PROMPT_DICTIONARY_TERM_INVENTORY_LIMIT]

    payload.setdefault("generated", {})["candidates"] = _trim_prompt_dictionary_candidates(
        sorted(
            candidates.values(),
            key=lambda item: (
                {"pending": 0, "approved": 1, "applied": 2, "rejected": 3}.get(str(item.get("status") or "pending"), 9),
                -float(item.get("score") or 0.0),
                str(item.get("kind") or ""),
                str(item.get("canonical") or item.get("value") or ""),
            ),
        ),
        limit=_PROMPT_DICTIONARY_CANDIDATE_LIMIT,
    )
    _apply_prompt_dictionary_candidate_auto_approval(payload)
    payload["generated"]["scan_results"] = scan_results
    payload.setdefault("metadata", {})["last_scanned_at"] = _utc_now_iso()
    payload["generated"]["evaluation_report"] = _build_prompt_dictionary_evaluation_report(payload)
    _emit_prompt_dictionary_progress(
        progress_callback,
        phase="build_complete",
        message="辞書作成が完了しました",
        candidate_count=len(payload["generated"].get("candidates") or []),
        extracted_term_count=len(payload["generated"].get("term_inventory") or []),
    )
    return payload


def build_prompt_dictionary_from_output(
    config: Optional[Dict[str, Any]] = None,
    max_files_per_source: int = 200,
    *,
    web_fetcher=None,
    llm_expander=None,
    progress_callback=None,
    persist: bool = True,
) -> Dict[str, Any]:
    payload = scan_prompt_dictionary_outputs(
        config=config,
        max_files_per_source=max_files_per_source,
        web_fetcher=web_fetcher,
        llm_expander=llm_expander,
        progress_callback=progress_callback,
    )
    payload.setdefault("metadata", {})["scan_note"] = "output_recursive_scan"
    if persist:
        save_prompt_dictionary_config(payload)
    return payload


def _normalize_prompt_dictionary_config(config: Dict[str, Any]) -> Dict[str, Any]:
    payload = copy.deepcopy(config or {})
    if not isinstance(payload, dict):
        payload = {}

    payload.setdefault("general_aliases", {})
    payload.setdefault("source_aliases", {})
    payload.setdefault("stopwords", [])
    payload.setdefault("stopwords_ja", [])
    payload.setdefault("file_extensions", [])
    payload.setdefault("allowlist", [])
    payload.setdefault("weak_stopwords", [])
    payload.setdefault("canonical_terms", [])
    payload.setdefault("aliases", {})
    payload.setdefault("source_overrides", {})
    payload.setdefault("allowlists", {"global": [], "sources": {}})
    payload.setdefault("metadata", {})
    payload.setdefault("assist", {})
    payload.setdefault("generated", {})

    aliases = payload.get("aliases") or {}
    if isinstance(aliases, dict):
        payload["general_aliases"] = _merge_alias_maps(payload.get("general_aliases") or {}, aliases)

    payload["general_aliases"] = _normalize_alias_map(payload.get("general_aliases") or {})
    payload["source_aliases"] = {
        source_key: _normalize_alias_map(values or {})
        for source_key, values in (payload.get("source_aliases") or {}).items()
        if isinstance(values, dict)
    }

    allowlists = payload.get("allowlists")
    if not isinstance(allowlists, dict):
        allowlists = {"global": [], "sources": {}}
    allowlists["global"] = _unique_normalized_tokens(allowlists.get("global") or [])
    normalized_source_allowlists = {}
    for source_key, values in (allowlists.get("sources") or {}).items():
        normalized_values = _unique_normalized_tokens(values or [])
        if normalized_values:
            normalized_source_allowlists[source_key] = normalized_values
    allowlists["sources"] = normalized_source_allowlists
    payload["allowlists"] = allowlists

    for key in ("stopwords", "stopwords_ja", "weak_stopwords", "file_extensions", "allowlist", "canonical_terms"):
        payload[key] = _unique_normalized_tokens(payload.get(key) or [], preserve_case=False)

    source_overrides = payload.get("source_overrides")
    if not isinstance(source_overrides, dict):
        source_overrides = {}
    normalized_overrides = {}
    for source_key, override in source_overrides.items():
        if not isinstance(override, dict):
            continue
        normalized_overrides[source_key] = {
            "aliases": _normalize_alias_map(override.get("aliases") or {}),
            "stopwords": _unique_normalized_tokens(override.get("stopwords") or [], preserve_case=False),
            "weak_stopwords": _unique_normalized_tokens(override.get("weak_stopwords") or [], preserve_case=False),
            "allowlist": _unique_normalized_tokens(override.get("allowlist") or [], preserve_case=False),
        }
    payload["source_overrides"] = normalized_overrides

    generated = payload.get("generated")
    if not isinstance(generated, dict):
        generated = {}
    generated["candidates"] = [
        _normalize_prompt_dictionary_candidate(candidate)
        for candidate in (generated.get("candidates") or [])
        if isinstance(candidate, dict)
    ]
    generated["scan_results"] = [
        scan_result
        for scan_result in (generated.get("scan_results") or [])
        if isinstance(scan_result, dict)
    ]
    generated["term_inventory"] = [
        term
        for term in (generated.get("term_inventory") or [])
        if isinstance(term, dict) and _normalize_text(term.get("term") or "")
    ]
    if not isinstance(generated.get("evaluation_report"), dict):
        generated["evaluation_report"] = {}
    payload["generated"] = generated

    metadata = payload.get("metadata")
    if not isinstance(metadata, dict):
        metadata = {}
    metadata.setdefault("version", 2)
    metadata.setdefault("last_scanned_at", "")
    metadata.setdefault("last_applied_at", "")
    metadata.setdefault("scan_note", "")
    payload["metadata"] = metadata
    assist = payload.get("assist")
    if not isinstance(assist, dict):
        assist = {}
    assist["web_enabled"] = bool(assist.get("web_enabled", False))
    assist["llm_enabled"] = bool(assist.get("llm_enabled", False))
    payload["assist"] = assist
    return payload


def evaluate_prompt_dictionary_benchmarks(
    config: Optional[Dict[str, Any]] = None,
    benchmark_cases: Optional[List[Dict[str, Any]]] = None,
) -> Dict[str, Any]:
    payload = _normalize_prompt_dictionary_config(config or _load_prompt_alias_config())
    cases = benchmark_cases or build_prompt_dictionary_benchmark_cases(payload)
    before_config = _strip_applied_candidate_effects(payload)
    after_config = apply_prompt_dictionary_candidates(copy.deepcopy(payload))
    report = {
        "total_cases": len(cases),
        "generated_at": _utc_now_iso(),
        "before": _evaluate_benchmark_side(cases, before_config),
        "after": _evaluate_benchmark_side(cases, after_config),
        "cases": [],
    }
    report["delta"] = _diff_evaluation_metrics(report["before"], report["after"])
    for case in cases:
        before_result = _run_benchmark_case(case, before_config)
        after_result = _run_benchmark_case(case, after_config)
        report["cases"].append(
            {
                "name": case.get("name") or "case",
                "expected_terms": case.get("expected_terms") or [],
                "before": before_result,
                "after": after_result,
            }
        )
    payload.setdefault("generated", {})["evaluation_report"] = report
    return payload


def build_prompt_dictionary_benchmark_cases(config: Optional[Dict[str, Any]] = None, max_cases: int = 20) -> List[Dict[str, Any]]:
    payload = _normalize_prompt_dictionary_config(config or _load_prompt_alias_config())
    candidates = list(((payload.get("generated") or {}).get("candidates") or []))
    cases: List[Dict[str, Any]] = []
    for candidate in candidates:
        if candidate.get("kind") != "alias":
            continue
        canonical = _normalize_text(candidate.get("canonical") or "")
        value = _normalize_text(candidate.get("value") or "")
        if not canonical or not value:
            continue
        label_value = candidate.get("value") or value
        template = "TAGS\n{dataportal_tag}\nEND"
        cases.append(
            {
                "name": f"alias:{canonical}->{value}",
                "template": template,
                "context_data": {
                    "name": canonical,
                    "existing_description": f"{canonical} observation dataset",
                    "dataportal_tag": json.dumps({"data": {"1": label_value, "2": "control"}}, ensure_ascii=False, indent=2),
                },
                "ai_config": {"prompt_assembly": {**default_prompt_assembly_config(), "default_mode": PROMPT_MODE_FILTERED_EMBED, "max_candidates": 2, "min_candidates": 1}},
                "feature_id": "json_suggest_tag",
                "template_name": "json_suggest_tag",
                "expected_terms": [label_value, value],
            }
        )
        if len(cases) >= max_cases:
            break
    return cases


def _strip_applied_candidate_effects(config: Dict[str, Any]) -> Dict[str, Any]:
    payload = _normalize_prompt_dictionary_config(copy.deepcopy(config))
    applied_candidates = [
        candidate for candidate in ((payload.get("generated") or {}).get("candidates") or [])
        if candidate.get("status") in {"approved", "applied"}
    ]
    for candidate in applied_candidates:
        kind = candidate.get("kind")
        canonical = _normalize_text(candidate.get("canonical") or "")
        value = _normalize_text(candidate.get("value") or "")
        source = _safe_str(candidate.get("source") or "").strip()
        if kind == "alias" and canonical and value:
            if source and source not in {"global", "observed_generic_tokens"}:
                override = (payload.get("source_overrides") or {}).get(source) or {}
                alias_map = override.get("aliases") or {}
                alias_map[canonical] = [item for item in alias_map.get(canonical, []) if _normalize_text(item) != value]
                if not alias_map[canonical]:
                    alias_map.pop(canonical, None)
            else:
                payload["general_aliases"][canonical] = [item for item in payload["general_aliases"].get(canonical, []) if _normalize_text(item) != value]
                if not payload["general_aliases"][canonical]:
                    payload["general_aliases"].pop(canonical, None)
        elif kind == "stopword" and value:
            payload["stopwords"] = [item for item in payload.get("stopwords", []) if _normalize_text(item) != value]
        elif kind == "weak_stopword" and value:
            payload["weak_stopwords"] = [item for item in payload.get("weak_stopwords", []) if _normalize_text(item) != value]
        elif kind == "allowlist" and value:
            payload["allowlist"] = [item for item in payload.get("allowlist", []) if _normalize_text(item) != value]
    return payload


def _evaluate_benchmark_side(cases: List[Dict[str, Any]], alias_config: Dict[str, Any]) -> Dict[str, Any]:
    metrics = {
        "zero_candidate_rate": 0.0,
        "over_candidate_rate": 0.0,
        "average_candidates": 0.0,
        "top_candidate_validity_rate": 0.0,
        "consistency_vs_full_embed_rate": 0.0,
        "case_count": len(cases),
    }
    if not cases:
        return metrics
    results = [_run_benchmark_case(case, alias_config) for case in cases]
    metrics["zero_candidate_rate"] = round(sum(1 for item in results if item.get("selected_candidates", 0) == 0) / len(results), 4)
    metrics["over_candidate_rate"] = round(sum(1 for item in results if item.get("selected_candidates", 0) > 3) / len(results), 4)
    metrics["average_candidates"] = round(sum(float(item.get("selected_candidates", 0)) for item in results) / len(results), 3)
    metrics["top_candidate_validity_rate"] = round(sum(1 for item in results if item.get("top_candidate_valid")) / len(results), 4)
    metrics["consistency_vs_full_embed_rate"] = round(sum(1 for item in results if item.get("consistent_with_full_embed")) / len(results), 4)
    return metrics


def _run_benchmark_case(case: Dict[str, Any], alias_config: Dict[str, Any]) -> Dict[str, Any]:
    result = build_prompt(
        case.get("template") or "",
        case.get("context_data") or {},
        ai_config=case.get("ai_config") or {"prompt_assembly": default_prompt_assembly_config()},
        feature_id=case.get("feature_id") or "",
        template_name=case.get("template_name") or "",
        alias_config_override=alias_config,
    )
    full_config = copy.deepcopy(case.get("ai_config") or {"prompt_assembly": default_prompt_assembly_config()})
    normalize_prompt_assembly_config_inplace(full_config)
    full_config["prompt_assembly"]["default_mode"] = PROMPT_MODE_FULL_EMBED
    full_result = build_prompt(
        case.get("template") or "",
        case.get("context_data") or {},
        ai_config=full_config,
        feature_id=case.get("feature_id") or "",
        template_name=case.get("template_name") or "",
        alias_config_override=alias_config,
    )
    source_diag = ((result.diagnostics.get("source_diagnostics") or [{}]) or [{}])[0]
    top_match = ((source_diag.get("top_matches") or [{}]) or [{}])[0]
    expected_terms = [_normalize_text(term) for term in (case.get("expected_terms") or []) if _normalize_text(term)]
    top_path_text = _normalize_text(top_match.get("path_text") or top_match.get("label") or "")
    prompt_text = _normalize_text(result.prompt)
    full_prompt_text = _normalize_text(full_result.prompt)
    return {
        "selected_candidates": int(source_diag.get("selected_candidates") or 0),
        "fallback_used": bool(source_diag.get("fallback_used")),
        "top_candidate_valid": any(term in top_path_text for term in expected_terms) if expected_terms else False,
        "consistent_with_full_embed": all(term in full_prompt_text for term in expected_terms) and any(term in prompt_text for term in expected_terms) if expected_terms else False,
        "top_match": top_match,
    }


def _diff_evaluation_metrics(before: Dict[str, Any], after: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "zero_candidate_rate_delta": round(float(after.get("zero_candidate_rate", 0.0)) - float(before.get("zero_candidate_rate", 0.0)), 4),
        "over_candidate_rate_delta": round(float(after.get("over_candidate_rate", 0.0)) - float(before.get("over_candidate_rate", 0.0)), 4),
        "average_candidates_delta": round(float(after.get("average_candidates", 0.0)) - float(before.get("average_candidates", 0.0)), 4),
        "top_candidate_validity_rate_delta": round(float(after.get("top_candidate_validity_rate", 0.0)) - float(before.get("top_candidate_validity_rate", 0.0)), 4),
        "consistency_vs_full_embed_rate_delta": round(float(after.get("consistency_vs_full_embed_rate", 0.0)) - float(before.get("consistency_vs_full_embed_rate", 0.0)), 4),
    }


def _scan_web_debug_directory(path: str, max_files: int) -> Dict[str, Any]:
    texts: List[Tuple[str, str]] = []
    pii_filtered_items = 0
    scanned_items = 0
    if not os.path.isdir(path):
        return _empty_scan_result("web_debug_html", path)
    for index, file_name in enumerate(sorted(os.listdir(path))):
        if index >= max_files:
            break
        if not file_name.lower().endswith((".html", ".htm")):
            continue
        file_path = os.path.join(path, file_name)
        try:
            with open(file_path, "r", encoding="utf-8", errors="ignore") as handle:
                raw_html = handle.read()
            scanned_items += 1
            text = _html_to_text(raw_html)
            if _looks_like_pii_value(text):
                pii_filtered_items += 1
                continue
            if text:
                texts.append((file_name, text))
        except Exception:
            logger.warning("web debug scan failed: %s", file_path, exc_info=True)
    return _build_scan_result("web_debug_html", path, texts, scanned_items, pii_filtered_items)


def _scan_recursive_directory(path: str, source_name: str, max_files: int, *, progress_callback=None) -> Dict[str, Any]:
    texts: List[Tuple[str, str]] = []
    pii_filtered_items = 0
    scanned_items = 0
    try:
        from classes.dataset.util.file_text_extractor import get_file_text_extractor

        extractor = get_file_text_extractor()
    except Exception:
        logger.warning("recursive scan extractor unavailable", exc_info=True)
        return _empty_scan_result(source_name, path)

    if not os.path.isdir(path):
        return _empty_scan_result(source_name, path)

    discovered_files: List[Tuple[str, str, str]] = []
    for root, _dirs, files in os.walk(path):
        for file_name in sorted(files):
            file_path = os.path.join(root, file_name)
            if not extractor.is_extractable(file_path):
                continue
            rel_path = os.path.relpath(file_path, path)
            ext = os.path.splitext(file_name)[1].lower()
            discovered_files.append((file_path, rel_path, ext))
            if len(discovered_files) >= max_files:
                break
        if len(discovered_files) >= max_files:
            break

    total_files = len(discovered_files)
    for index, (file_path, rel_path, ext) in enumerate(discovered_files, start=1):
        scanned_items += 1
        extracted_entries = _extract_entries_from_file(file_path, rel_path, ext, extractor)
        for key, value in extracted_entries:
            if _looks_like_pii_field(key) or _looks_like_pii_value(value):
                pii_filtered_items += 1
                continue
            text = _safe_str(value).strip()
            if text:
                texts.append((key, text))
        if index == 1 or index % _PROMPT_DICTIONARY_PROGRESS_EVENT_FILE_GRANULARITY == 0 or index == total_files:
            _emit_prompt_dictionary_progress(
                progress_callback,
                phase="file_progress",
                message=f"{source_name}: {index}/{total_files} files",
                source=source_name,
                current=index,
                total=total_files,
                scanned_items=scanned_items,
            )
    return _build_scan_result(source_name, path, texts, scanned_items, pii_filtered_items)


def _extract_entries_from_file(file_path: str, rel_path: str, ext: str, extractor) -> List[Tuple[str, str]]:
    if ext == ".json":
        try:
            with open(file_path, "r", encoding="utf-8") as handle:
                payload = json.load(handle)
            return [(f"{rel_path}:{key}", value) for key, value in _extract_generic_json_texts(payload)]
        except Exception:
            logger.debug("generic json parse failed during recursive scan: %s", file_path, exc_info=True)
    text = extractor.extract_text(file_path, os.path.basename(file_path))
    if not text:
        return []
    if ext in {".html", ".htm"}:
        text = _html_to_text(text)
    return [(rel_path, text)]


def _html_to_text(raw_html: str) -> str:
    text = re.sub(r"<script[\s\S]*?</script>", " ", raw_html or "", flags=re.IGNORECASE)
    text = re.sub(r"<style[\s\S]*?</style>", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _normalize_external_candidates(items: Any) -> List[Dict[str, Any]]:
    normalized = []
    for item in items or []:
        if not isinstance(item, dict):
            continue
        normalized.append(_normalize_prompt_dictionary_candidate(item))
    return normalized


def _normalize_alias_map(aliases: Dict[str, Any]) -> Dict[str, List[str]]:
    normalized = {}
    for canonical, values in (aliases or {}).items():
        normalized_key = _normalize_text(canonical)
        if not normalized_key:
            continue
        collected = []
        for value in values or []:
            normalized_value = _normalize_text(value)
            if normalized_value and normalized_value != normalized_key:
                collected.append(normalized_value)
        if collected:
            normalized[normalized_key] = list(dict.fromkeys(collected))
        elif normalized_key not in normalized:
            normalized[normalized_key] = []
    return normalized


def _merge_alias_maps(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, List[str]]:
    merged = _normalize_alias_map(base or {})
    for canonical, values in _normalize_alias_map(override or {}).items():
        merged.setdefault(canonical, [])
        merged[canonical] = list(dict.fromkeys(list(merged[canonical]) + list(values)))
    return merged


def _unique_normalized_tokens(values: Iterable[Any], preserve_case: bool = False) -> List[str]:
    tokens = []
    for value in values or []:
        token = _safe_str(value).strip()
        if not token:
            continue
        tokens.append(token if preserve_case else _normalize_text(token))
    return list(dict.fromkeys(token for token in tokens if token))


def _normalize_prompt_dictionary_candidate(candidate: Dict[str, Any]) -> Dict[str, Any]:
    normalized = {
        "id": candidate.get("id") or "",
        "kind": _safe_str(candidate.get("kind") or "alias").strip() or "alias",
        "canonical": _normalize_text(candidate.get("canonical") or ""),
        "value": _normalize_text(candidate.get("value") or ""),
        "source": _safe_str(candidate.get("source") or "").strip(),
        "status": _safe_str(candidate.get("status") or "pending").strip() or "pending",
        "score": float(candidate.get("score") or 0.0),
        "reason": _safe_str(candidate.get("reason") or "").strip(),
        "examples": list(dict.fromkeys(_safe_str(example).strip() for example in (candidate.get("examples") or []) if _safe_str(example).strip()))[:5],
        "occurrences": max(1, int(candidate.get("occurrences") or 1)),
        "created_at": _safe_str(candidate.get("created_at") or _utc_now_iso()).strip(),
        "updated_at": _safe_str(candidate.get("updated_at") or _utc_now_iso()).strip(),
    }
    if not normalized["id"]:
        normalized["id"] = _prompt_dictionary_candidate_id(
            normalized["kind"],
            normalized["canonical"],
            normalized["value"],
            normalized["source"],
        )
    return normalized


def _merge_prompt_dictionary_candidate(existing: Optional[Dict[str, Any]], incoming: Dict[str, Any]) -> Dict[str, Any]:
    normalized = _normalize_prompt_dictionary_candidate(incoming)
    if not existing:
        return normalized
    merged = _normalize_prompt_dictionary_candidate(existing)
    if merged.get("status") not in {"rejected", "applied"}:
        if normalized.get("status") in {"applied", "approved"}:
            merged["status"] = normalized.get("status")
        elif merged.get("status") not in {"approved"}:
            merged["status"] = normalized.get("status") or merged.get("status")
    merged["score"] = max(float(merged.get("score") or 0.0), float(normalized.get("score") or 0.0))
    merged["occurrences"] = int(merged.get("occurrences") or 0) + int(normalized.get("occurrences") or 0)
    merged["examples"] = list(dict.fromkeys(list(merged.get("examples") or []) + list(normalized.get("examples") or [])))[:5]
    merged["reason"] = merged.get("reason") or normalized.get("reason")
    merged["updated_at"] = _utc_now_iso()
    return merged


def _merge_candidate_into_preview(preview: Dict[str, Any], candidate: Dict[str, Any]) -> None:
    kind = candidate.get("kind")
    canonical = _normalize_text(candidate.get("canonical") or "")
    value = _normalize_text(candidate.get("value") or "")
    source = _safe_str(candidate.get("source") or "").strip()
    if kind == "alias" and canonical and value:
        if source and source not in {"observed_generic_tokens", "global"}:
            source_override = preview.setdefault("source_overrides", {}).setdefault(source, {"aliases": {}, "stopwords": [], "weak_stopwords": [], "allowlist": []})
            source_override.setdefault("aliases", {}).setdefault(canonical, [])
            if value not in source_override["aliases"][canonical]:
                source_override["aliases"][canonical].append(value)
        else:
            preview.setdefault("general_aliases", {}).setdefault(canonical, [])
            if value not in preview["general_aliases"][canonical]:
                preview["general_aliases"][canonical].append(value)
    elif kind == "stopword" and value:
        preview.setdefault("stopwords", [])
        if value not in preview["stopwords"]:
            preview["stopwords"].append(value)
    elif kind == "weak_stopword" and value:
        preview.setdefault("weak_stopwords", [])
        if value not in preview["weak_stopwords"]:
            preview["weak_stopwords"].append(value)
    elif kind == "allowlist" and value:
        preview.setdefault("allowlist", [])
        if value not in preview["allowlist"]:
            preview["allowlist"].append(value)


def _apply_candidate_to_config(config: Dict[str, Any], candidate: Dict[str, Any]) -> None:
    kind = candidate.get("kind")
    canonical = _normalize_text(candidate.get("canonical") or "")
    value = _normalize_text(candidate.get("value") or "")
    source = _safe_str(candidate.get("source") or "").strip()
    if kind == "alias" and canonical and value:
        if source and source not in {"observed_generic_tokens", "global"}:
            source_override = config.setdefault("source_overrides", {}).setdefault(source, {"aliases": {}, "stopwords": [], "weak_stopwords": [], "allowlist": []})
            source_override.setdefault("aliases", {}).setdefault(canonical, [])
            if value not in source_override["aliases"][canonical]:
                source_override["aliases"][canonical].append(value)
        else:
            config.setdefault("general_aliases", {}).setdefault(canonical, [])
            if value not in config["general_aliases"][canonical]:
                config["general_aliases"][canonical].append(value)
    elif kind == "stopword" and value:
        if value not in config.setdefault("stopwords", []):
            config["stopwords"].append(value)
    elif kind == "weak_stopword" and value:
        if value not in config.setdefault("weak_stopwords", []):
            config["weak_stopwords"].append(value)
    elif kind == "allowlist" and value:
        if value not in config.setdefault("allowlist", []):
            config["allowlist"].append(value)


def _build_prompt_dictionary_evaluation_report(config: Dict[str, Any]) -> Dict[str, Any]:
    candidates = list(((config.get("generated") or {}).get("candidates") or []))
    scan_results = list(((config.get("generated") or {}).get("scan_results") or []))
    term_inventory = list(((config.get("generated") or {}).get("term_inventory") or []))
    return {
        "pending_candidates": sum(1 for candidate in candidates if candidate.get("status") == "pending"),
        "approved_candidates": sum(1 for candidate in candidates if candidate.get("status") == "approved"),
        "applied_candidates": sum(1 for candidate in candidates if candidate.get("status") == "applied"),
        "rejected_candidates": sum(1 for candidate in candidates if candidate.get("status") == "rejected"),
        "scan_sources": len(scan_results),
        "scanned_items": sum(int(item.get("scanned_items") or 0) for item in scan_results),
        "pii_filtered_items": sum(int(item.get("pii_filtered_items") or 0) for item in scan_results),
        "generated_alias_candidates": sum(1 for candidate in candidates if candidate.get("kind") == "alias"),
        "generated_stopword_candidates": sum(1 for candidate in candidates if candidate.get("kind") in {"stopword", "weak_stopword"}),
        "generated_allowlist_candidates": sum(1 for candidate in candidates if candidate.get("kind") == "allowlist"),
        "extracted_term_count": len(term_inventory),
        "auto_approved_candidates": sum(1 for candidate in candidates if candidate.get("status") in {"approved", "applied"} and "auto_approved" in _safe_str(candidate.get("reason"))),
        "recursive_scan_sources": sum(1 for item in scan_results if str(item.get("source") or "").endswith("_recursive")),
        "updated_at": _utc_now_iso(),
    }


def _emit_prompt_dictionary_progress(progress_callback, **payload: Any) -> None:
    if not callable(progress_callback):
        return
    try:
        progress_callback(payload)
    except Exception:
        logger.debug("prompt dictionary progress callback failed", exc_info=True)


def _merge_prompt_dictionary_term_inventory_entry(term_inventory_map: Dict[str, Dict[str, Any]], incoming: Dict[str, Any]) -> None:
    term = _normalize_text(incoming.get("term") or "")
    if not term:
        return
    existing = term_inventory_map.get(term)
    if not existing:
        term_inventory_map[term] = {
            "term": term,
            "occurrences": max(1, int(incoming.get("occurrences") or 1)),
            "document_count": max(1, int(incoming.get("document_count") or 1)),
            "source_hits": list(dict.fromkeys(_safe_str(item).strip() for item in (incoming.get("source_hits") or []) if _safe_str(item).strip())),
            "examples": list(dict.fromkeys(_safe_str(item).strip() for item in (incoming.get("examples") or []) if _safe_str(item).strip()))[:5],
            "pattern_hits": list(dict.fromkeys(_safe_str(item).strip() for item in (incoming.get("pattern_hits") or []) if _safe_str(item).strip())),
        }
        return
    existing["occurrences"] = int(existing.get("occurrences") or 0) + max(1, int(incoming.get("occurrences") or 1))
    existing["document_count"] = int(existing.get("document_count") or 0) + max(1, int(incoming.get("document_count") or 1))
    existing["source_hits"] = list(dict.fromkeys(list(existing.get("source_hits") or []) + list(incoming.get("source_hits") or [])))
    existing["examples"] = list(dict.fromkeys(list(existing.get("examples") or []) + list(incoming.get("examples") or [])))[:5]
    existing["pattern_hits"] = list(dict.fromkeys(list(existing.get("pattern_hits") or []) + list(incoming.get("pattern_hits") or [])))


def _materialize_prompt_dictionary_term_inventory(term_inventory_map: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
    inventory: List[Dict[str, Any]] = []
    for term, raw in term_inventory_map.items():
        entry = {
            "term": term,
            "occurrences": max(1, int(raw.get("occurrences") or 1)),
            "document_count": max(1, int(raw.get("document_count") or 1)),
            "source_hits": list(dict.fromkeys(raw.get("source_hits") or [])),
            "source_count": len(set(raw.get("source_hits") or [])),
            "examples": list(dict.fromkeys(raw.get("examples") or []))[:5],
            "pattern_hits": list(dict.fromkeys(raw.get("pattern_hits") or [])),
        }
        entry["quality_score"] = round(_score_prompt_dictionary_term_entry(entry), 4)
        inventory.append(entry)
    inventory.sort(
        key=lambda item: (
            -float(item.get("quality_score") or 0.0),
            -int(item.get("source_count") or 0),
            -int(item.get("occurrences") or 0),
            item.get("term") or "",
        )
    )
    return inventory


def _score_prompt_dictionary_term_entry(entry: Dict[str, Any]) -> float:
    term = _normalize_text(entry.get("term") or "")
    if not term or not _is_prompt_dictionary_term_candidate(term):
        return 0.0
    occurrences = max(1, int(entry.get("occurrences") or 1))
    document_count = max(1, int(entry.get("document_count") or 1))
    source_count = max(1, int(entry.get("source_count") or len(entry.get("source_hits") or [])))
    pattern_hits = set(entry.get("pattern_hits") or [])

    score = 0.12
    score += min(0.22, math.log1p(occurrences) * 0.06)
    score += min(0.18, math.log1p(document_count) * 0.05)
    score += min(0.2, source_count * 0.06)
    if pattern_hits & {"slash_pair", "paren_pair", "reverse_paren_pair", "acronym", "technical_phrase"}:
        score += 0.16
    if any(hint in term for hint in _PROMPT_DICTIONARY_DOMAIN_HINTS):
        score += 0.18
    if _script_kind(term) == "latin" and re.fullmatch(r"[a-z]{2,8}[0-9]{0,2}", term):
        score += 0.08
    if re.fullmatch(r"[a-z]{2,12}(?:\s+[a-z]{2,16}){1,3}", term):
        score += 0.05
    if term in _PROMPT_DICTIONARY_GENERIC_TERM_SEEDS:
        score -= 0.4
    if len(term) <= 2 and term not in (_DEFAULT_ALIAS_CONFIG.get("allowlist") or []):
        score -= 0.3
    if len(term) > 40:
        score -= 0.25
    if any(ch.isdigit() for ch in term) and not re.fullmatch(r"[a-z]{1,8}[0-9]{1,4}", term):
        score -= 0.1
    return max(0.0, min(0.99, score))


def _build_allowlist_candidates_from_term_inventory(term_inventory: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    candidates: List[Dict[str, Any]] = []
    for entry in term_inventory[:_PROMPT_DICTIONARY_TERM_INVENTORY_LIMIT]:
        term = _normalize_text(entry.get("term") or "")
        quality_score = float(entry.get("quality_score") or 0.0)
        if not term or quality_score < 0.58:
            continue
        if term in _PROMPT_DICTIONARY_GENERIC_TERM_SEEDS or term in (_DEFAULT_ALIAS_CONFIG.get("stopwords") or []):
            continue
        candidates.append(
            _make_prompt_dictionary_candidate(
                kind="allowlist",
                canonical="",
                value=term,
                source="term_inventory",
                score=quality_score,
                reason=f"term_inventory quality={quality_score:.3f} occ={int(entry.get('occurrences') or 0)} sources={int(entry.get('source_count') or 0)}",
                examples=list(entry.get("examples") or [])[:5],
                occurrences=int(entry.get("occurrences") or 1),
            )
        )
    return candidates


def _apply_prompt_dictionary_candidate_auto_approval(payload: Dict[str, Any]) -> None:
    for candidate in ((payload.get("generated") or {}).get("candidates") or []):
        if candidate.get("status") != "pending":
            continue
        threshold = _PROMPT_DICTIONARY_AUTO_APPROVE_THRESHOLDS.get(candidate.get("kind"))
        if threshold is None:
            continue
        score = float(candidate.get("score") or 0.0)
        occurrences = int(candidate.get("occurrences") or 1)
        reason = _safe_str(candidate.get("reason") or "")
        if score < threshold:
            continue
        if candidate.get("kind") == "alias" and occurrences < 2 and "llm_review" not in _safe_str(candidate.get("source")) and not any(tag in reason for tag in ("paren_pair", "slash_pair", "reverse_paren_pair")):
            continue
        if candidate.get("kind") == "allowlist" and occurrences < 2 and "llm_review" not in _safe_str(candidate.get("source")):
            continue
        candidate["status"] = "approved"
        candidate["reason"] = f"auto_approved; {reason}".strip("; ")
        candidate["updated_at"] = _utc_now_iso()


def _trim_prompt_dictionary_candidates(candidates: List[Dict[str, Any]], limit: int) -> List[Dict[str, Any]]:
    if limit <= 0 or len(candidates) <= limit:
        return candidates
    alias_candidates = [candidate for candidate in candidates if candidate.get("kind") == "alias"]
    stopword_candidates = [candidate for candidate in candidates if candidate.get("kind") in {"stopword", "weak_stopword"}]
    remainder = [candidate for candidate in candidates if candidate.get("kind") not in {"alias", "stopword", "weak_stopword"}]
    preserved = alias_candidates + stopword_candidates
    if len(preserved) >= limit:
        return preserved[:limit]
    remaining_slots = limit - len(preserved)
    return preserved + remainder[:remaining_slots]


def _run_llm_prompt_dictionary_review(payload: Dict[str, Any], scan_results: List[Dict[str, Any]], term_inventory: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    try:
        from classes.ai.core.ai_manager import AIManager
    except Exception:
        logger.debug("AIManager unavailable for prompt dictionary review", exc_info=True)
        return []

    ai_manager = AIManager()
    provider = ai_manager.get_default_provider()
    model = ai_manager.get_default_model(provider)
    if not provider or not model:
        return []

    top_terms = [
        {
            "term": item.get("term"),
            "quality_score": item.get("quality_score"),
            "occurrences": item.get("occurrences"),
            "source_count": item.get("source_count"),
            "examples": list(item.get("examples") or [])[:2],
        }
        for item in term_inventory[:120]
    ]
    top_aliases = [
        {
            "canonical": candidate.get("canonical"),
            "value": candidate.get("value"),
            "score": candidate.get("score"),
            "reason": candidate.get("reason"),
        }
        for candidate in ((payload.get("generated") or {}).get("candidates") or [])[:80]
        if candidate.get("kind") == "alias"
    ]
    if not top_terms and not top_aliases:
        return []

    prompt = (
        "ARIM/RDE の材料・装置・分析用語辞書を整理してください。"
        "以下の extracted_terms と alias_candidates を見て、"
        "高品質な alias / allowlist / stopword を JSON で返してください。"
        "\n\n"
        "返却形式は厳密な JSON のみで、"
        "aliases, allowlist, stopwords の3キーを持ちます。"
        "\n"
        "aliases は canonical, value, confidence, reason。"
        "allowlist / stopwords は term, confidence, reason。"
        "\n"
        "人物名、メール、組織名、ID、汎用語は除外し、材料・装置・手法・測定語だけを残してください。"
        "\n\n"
        f"extracted_terms={json.dumps(top_terms, ensure_ascii=False)}\n"
        f"alias_candidates={json.dumps(top_aliases, ensure_ascii=False)}"
    )
    result = ai_manager.send_prompt(prompt, provider, model)
    if not isinstance(result, dict) or not result.get("success"):
        logger.info("prompt dictionary llm review skipped: %s", (result or {}).get("error"))
        return []
    response_text = _safe_str(result.get("response") or result.get("content") or result.get("raw_response"))
    llm_payload = _extract_json_object_from_text(response_text)
    if not isinstance(llm_payload, dict):
        return []
    return _normalize_llm_review_payload(llm_payload)


def _extract_json_object_from_text(text: str) -> Any:
    raw = _safe_str(text).strip()
    if not raw:
        return None
    try:
        return json.loads(raw)
    except Exception:
        pass
    match = re.search(r"\{[\s\S]*\}", raw)
    if not match:
        return None
    try:
        return json.loads(match.group(0))
    except Exception:
        logger.debug("llm review json parse failed", exc_info=True)
        return None


def _normalize_llm_review_payload(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    candidates: List[Dict[str, Any]] = []
    for item in payload.get("aliases") or []:
        canonical = _normalize_text(item.get("canonical") or "")
        value = _normalize_text(item.get("value") or "")
        if not canonical or not value or canonical == value:
            continue
        confidence = max(0.0, min(0.99, float(item.get("confidence") or 0.0)))
        candidates.append(
            _make_prompt_dictionary_candidate(
                kind="alias",
                canonical=canonical,
                value=value,
                source="llm_review",
                score=max(0.7, confidence),
                reason=f"llm_review confidence={confidence:.3f}; {_safe_str(item.get('reason') or '').strip()}",
                examples=[f"{canonical} <-> {value}"],
                occurrences=2,
                status="approved" if confidence >= _PROMPT_DICTIONARY_AUTO_APPROVE_THRESHOLDS["alias"] else "pending",
            )
        )
    for group_name, candidate_kind in (("allowlist", "allowlist"), ("stopwords", "stopword")):
        for item in payload.get(group_name) or []:
            term = _normalize_text(item.get("term") or "")
            if not term:
                continue
            confidence = max(0.0, min(0.99, float(item.get("confidence") or 0.0)))
            status = "approved" if confidence >= _PROMPT_DICTIONARY_AUTO_APPROVE_THRESHOLDS[candidate_kind] else "pending"
            candidates.append(
                _make_prompt_dictionary_candidate(
                    kind=candidate_kind,
                    canonical="",
                    value=term,
                    source="llm_review",
                    score=max(0.68, confidence),
                    reason=f"llm_review confidence={confidence:.3f}; {_safe_str(item.get('reason') or '').strip()}",
                    examples=[term],
                    occurrences=2,
                    status=status,
                )
            )
    return candidates


def _is_prompt_dictionary_term_candidate(term: str) -> bool:
    normalized = _normalize_text(term)
    if not normalized:
        return False
    if _looks_like_pii_value(normalized):
        return False
    if normalized in _PROMPT_DICTIONARY_GENERIC_TERM_SEEDS:
        return False
    if normalized.startswith("http") or "/" in normalized and len(normalized) > 30:
        return False
    if re.fullmatch(r"[0-9\-_.]+", normalized):
        return False
    return 2 <= len(normalized) <= 40


def _extract_prompt_dictionary_term_inventory(text: str, source_name: str, document_key: str) -> Dict[str, Dict[str, Any]]:
    raw_text = _safe_str(text).strip()
    if not raw_text:
        return {}
    terms: Dict[str, Dict[str, Any]] = {}
    segments = [segment.strip() for segment in re.split(r"[\n\r\t|;,]+", raw_text) if segment.strip()]
    patterns = []
    for segment in segments[:120]:
        patterns.extend(_extract_term_patterns_from_segment(segment))
    if not patterns:
        patterns = [(raw_text[:120], "raw_text")]
    per_document_seen = set()
    for raw_term, pattern in patterns:
        normalized = _normalize_text(raw_term)
        if not _is_prompt_dictionary_term_candidate(normalized):
            continue
        if normalized in per_document_seen:
            continue
        per_document_seen.add(normalized)
        terms[normalized] = {
            "term": normalized,
            "occurrences": 1,
            "document_count": 1,
            "source_hits": [source_name],
            "examples": [f"{document_key}: {raw_term[:80]}"],
            "pattern_hits": [pattern],
        }
    return terms


def _extract_term_patterns_from_segment(segment: str) -> List[Tuple[str, str]]:
    results: List[Tuple[str, str]] = []
    cleaned = _safe_str(segment).strip()
    if not cleaned:
        return results
    if len(cleaned) <= 80:
        results.append((cleaned, "segment"))
    for match in re.finditer(r"([A-Za-z][A-Za-z0-9+._-]{1,12})\s*\(([^()]{3,80})\)", cleaned):
        results.append((match.group(1), "acronym"))
        results.append((match.group(2), "paren_pair"))
    for match in re.finditer(r"([^/]{3,80})\s*/\s*([^/]{3,80})", cleaned):
        results.append((match.group(1), "slash_pair"))
        results.append((match.group(2), "slash_pair"))
    for match in re.finditer(r"[A-Z]{2,12}(?:[-/][A-Z0-9]{1,8})?", cleaned):
        results.append((match.group(0), "acronym"))
    for match in re.finditer(r"[一-龯ぁ-んァ-ヶー]{2,24}", cleaned):
        results.append((match.group(0), "technical_phrase"))
    for match in re.finditer(r"[A-Za-z][A-Za-z0-9+._-]{1,24}(?:\s+[A-Za-z][A-Za-z0-9+._-]{1,24}){0,3}", cleaned):
        results.append((match.group(0), "technical_phrase"))
    return results[:40]


def _scan_summary_workbook(path: str, source_name: str) -> Dict[str, Any]:
    texts: List[Tuple[str, str]] = []
    pii_filtered_items = 0
    scanned_items = 0
    try:
        if not os.path.exists(path):
            return _empty_scan_result(source_name, path)
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet_name = "entries" if source_name == "summary_entries" else "datasets"
        if sheet_name not in workbook.sheetnames:
            return _empty_scan_result(source_name, path)
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return _empty_scan_result(source_name, path)
        headers = [_safe_str(value).strip() for value in rows[0]]
        for row in rows[1:]:
            record = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
            if all(_normalize_text(record.get(header, "")) == _normalize_text(header) for header in headers if header):
                continue
            for key, value in record.items():
                scanned_items += 1
                if not key or key in _PROMPT_DICTIONARY_SUMMARY_EXCLUDED_COLUMNS or _looks_like_pii_field(key) or _looks_like_pii_value(value):
                    pii_filtered_items += 1
                    continue
                text = _safe_str(value).strip()
                if text:
                    texts.append((f"{sheet_name}:{key}", text))
    except Exception:
        logger.warning("summary workbook scan failed: %s", path, exc_info=True)
        return _empty_scan_result(source_name, path)
    return _build_scan_result(source_name, path, texts, scanned_items, pii_filtered_items)


def _scan_json_directory(path: str, source_name: str, max_files: int, extractor, *, progress_callback=None) -> Dict[str, Any]:
    texts: List[Tuple[str, str]] = []
    pii_filtered_items = 0
    scanned_items = 0
    if not os.path.isdir(path):
        return _empty_scan_result(source_name, path)
    try:
        file_names = [file_name for file_name in sorted(os.listdir(path)) if file_name.lower().endswith(".json")][:max_files]
        total_files = len(file_names)
        for index, file_name in enumerate(file_names, start=1):
            file_path = os.path.join(path, file_name)
            file_texts, file_scanned, file_pii = _scan_single_json(file_path, source_name, extractor, return_tuple=True)
            texts.extend(file_texts)
            scanned_items += file_scanned
            pii_filtered_items += file_pii
            if index == 1 or index % _PROMPT_DICTIONARY_PROGRESS_EVENT_FILE_GRANULARITY == 0 or index == total_files:
                _emit_prompt_dictionary_progress(
                    progress_callback,
                    phase="file_progress",
                    message=f"{source_name}: {index}/{total_files} files",
                    source=source_name,
                    current=index,
                    total=total_files,
                    scanned_items=scanned_items,
                )
    except Exception:
        logger.warning("json directory scan failed: %s", path, exc_info=True)
    return _build_scan_result(source_name, path, texts, scanned_items, pii_filtered_items)


def _scan_single_json(path: str, source_name: str, extractor, return_tuple: bool = False):
    texts: List[Tuple[str, str]] = []
    pii_filtered_items = 0
    scanned_items = 0
    try:
        if not os.path.exists(path):
            result = _empty_scan_result(source_name, path)
            return ([], 0, 0) if return_tuple else result
        with open(path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
        extracted = extractor(payload)
        for key, value in extracted:
            scanned_items += 1
            if _looks_like_pii_field(key) or _looks_like_pii_value(value):
                pii_filtered_items += 1
                continue
            text = _safe_str(value).strip()
            if text:
                texts.append((key, text))
    except Exception:
        logger.warning("json scan failed: %s", path, exc_info=True)
    if return_tuple:
        return texts, scanned_items, pii_filtered_items
    return _build_scan_result(source_name, path, texts, scanned_items, pii_filtered_items)


def _build_scan_result(source_name: str, source_path: str, texts: List[Tuple[str, str]], scanned_items: int, pii_filtered_items: int) -> Dict[str, Any]:
    generic_tokens = Counter()
    alias_candidates: List[Dict[str, Any]] = []
    examples: List[str] = []
    term_inventory_map: Dict[str, Dict[str, Any]] = {}
    for key, text in texts:
        if len(examples) < 5:
            examples.append(f"{key}: {text[:80]}")
        generic_tokens.update(_extract_generic_candidate_tokens(text))
        alias_candidates.extend(_extract_alias_candidates_from_text(text, source_name))
        for term_entry in _extract_prompt_dictionary_term_inventory(text, source_name, key).values():
            _merge_prompt_dictionary_term_inventory_entry(term_inventory_map, term_entry)
    merged_candidates: Dict[str, Dict[str, Any]] = {}
    for candidate in alias_candidates:
        merged_candidates[candidate["id"]] = _merge_prompt_dictionary_candidate(merged_candidates.get(candidate["id"]), candidate)
    term_inventory = _materialize_prompt_dictionary_term_inventory(term_inventory_map)
    return {
        "source": source_name,
        "path": source_path,
        "scanned_items": scanned_items,
        "text_items": len(texts),
        "pii_filtered_items": pii_filtered_items,
        "term_count": len(term_inventory),
        "top_terms": [item.get("term") for item in term_inventory[:8]],
        "generic_tokens": dict(generic_tokens),
        "candidates": list(merged_candidates.values()),
        "_term_inventory_raw": term_inventory,
        "examples": examples,
    }


def _empty_scan_result(source_name: str, source_path: str) -> Dict[str, Any]:
    return {
        "source": source_name,
        "path": source_path,
        "scanned_items": 0,
        "text_items": 0,
        "pii_filtered_items": 0,
        "term_count": 0,
        "top_terms": [],
        "generic_tokens": {},
        "candidates": [],
        "_term_inventory_raw": [],
        "examples": [],
    }


def _extract_data_entry_texts(payload: Any) -> List[Tuple[str, str]]:
    results: List[Tuple[str, str]] = []
    for item in payload.get("data", []) or []:
        attributes = item.get("attributes") or {}
        if attributes.get("name"):
            results.append(("data.name", attributes.get("name")))
        metadata = attributes.get("metadata") or {}
        for key, value in metadata.items():
            results.append((f"metadata.key:{key}", key))
            if isinstance(value, dict):
                if value.get("value"):
                    results.append((f"metadata.value:{key}", value.get("value")))
            elif value:
                results.append((f"metadata.value:{key}", value))
    return results


def _extract_invoice_texts(payload: Any) -> List[Tuple[str, str]]:
    results: List[Tuple[str, str]] = []
    data = payload.get("data") or {}
    if isinstance(data, dict):
        basic = ((data.get("attributes") or {}).get("basic") or {})
        sample = ((data.get("attributes") or {}).get("sample") or {})
        for key in ("dataName", "description"):
            if basic.get(key):
                results.append((f"invoice.basic.{key}", basic.get(key)))
        for value in sample.get("names") or []:
            results.append(("invoice.sample.names", value))
        if sample.get("primaryName"):
            results.append(("invoice.sample.primaryName", sample.get("primaryName")))
    for included in payload.get("included", []) or []:
        item_type = included.get("type")
        attributes = included.get("attributes") or {}
        if item_type == "instrument":
            for key in ("nameJa", "nameEn", "organizationNameJa", "organizationNameEn"):
                if attributes.get(key):
                    results.append((f"instrument.{key}", attributes.get(key)))
            for program in attributes.get("programs") or []:
                if isinstance(program, dict) and program.get("programName"):
                    results.append(("instrument.programName", program.get("programName")))
    return results


def _extract_generic_json_texts(payload: Any) -> List[Tuple[str, str]]:
    results: List[Tuple[str, str]] = []

    def walk(node: Any, path: List[str]) -> None:
        if isinstance(node, dict):
            for key, value in node.items():
                walk(value, path + [str(key)])
            return
        if isinstance(node, list):
            for value in node[:50]:
                walk(value, path)
            return
        if isinstance(node, str):
            joined = ".".join(path[-4:])
            results.append((joined, node))

    walk(payload, [])
    return results


def _extract_alias_candidates_from_text(text: str, source_name: str) -> List[Dict[str, Any]]:
    candidates: List[Dict[str, Any]] = []
    raw_text = _safe_str(text).strip()
    if not raw_text:
        return candidates

    slash_match = re.match(r"^\s*([^/]{2,64})\s*/\s*([^/]{2,64})\s*$", raw_text)
    if slash_match:
        left = slash_match.group(1).strip()
        right = slash_match.group(2).strip()
        candidate = _make_alias_candidate_from_pair(left, right, source_name, reason="slash_pair")
        if candidate:
            candidates.append(candidate)

    paren_match = re.match(r"^\s*([^()]{3,80})\s*\(([A-Za-z0-9+._-]{2,16})\)\s*$", raw_text)
    if paren_match:
        long_form = paren_match.group(1).strip()
        short_form = paren_match.group(2).strip()
        candidate = _make_alias_candidate_from_pair(short_form, long_form, source_name, reason="paren_pair")
        if candidate:
            candidates.append(candidate)

    reverse_paren_match = re.match(r"^\s*([A-Za-z0-9+._-]{2,16})\s*\(([^()]{3,80})\)\s*$", raw_text)
    if reverse_paren_match:
        short_form = reverse_paren_match.group(1).strip()
        long_form = reverse_paren_match.group(2).strip()
        candidate = _make_alias_candidate_from_pair(short_form, long_form, source_name, reason="reverse_paren_pair")
        if candidate:
            candidates.append(candidate)

    bracket_match = re.match(r"^\s*\[[^\]]+\]\s*(.{3,80})$", raw_text)
    if bracket_match:
        cleaned = bracket_match.group(1).strip()
        candidate = _make_alias_candidate_from_pair(cleaned, raw_text, source_name, reason="bracket_cleanup")
        if candidate:
            candidates.append(candidate)

    return candidates


def _make_alias_candidate_from_pair(left: str, right: str, source_name: str, reason: str) -> Optional[Dict[str, Any]]:
    left_norm = _normalize_text(left)
    right_norm = _normalize_text(right)
    if not left_norm or not right_norm or left_norm == right_norm:
        return None
    if _looks_like_pii_value(left) or _looks_like_pii_value(right):
        return None
    if _script_kind(left_norm) == _script_kind(right_norm) and max(len(left_norm), len(right_norm)) > 20:
        return None
    canonical = left_norm if _prefer_as_canonical(left_norm, right_norm) else right_norm
    alias_value = right_norm if canonical == left_norm else left_norm
    score = 0.72 if reason in {"slash_pair", "paren_pair", "reverse_paren_pair"} else 0.58
    return _make_prompt_dictionary_candidate(
        kind="alias",
        canonical=canonical,
        value=alias_value,
        source=source_name,
        score=score,
        reason=reason,
        examples=[f"{left} <-> {right}"],
        occurrences=1,
    )


def _make_prompt_dictionary_candidate(kind: str, canonical: str, value: str, source: str, score: float, reason: str, examples: List[str], occurrences: int, status: str = "pending") -> Dict[str, Any]:
    canonical_norm = _normalize_text(canonical)
    value_norm = _normalize_text(value)
    candidate = {
        "id": _prompt_dictionary_candidate_id(kind, canonical_norm, value_norm, source),
        "kind": kind,
        "canonical": canonical_norm,
        "value": value_norm,
        "source": source,
        "status": status,
        "score": round(float(score), 3),
        "reason": reason,
        "examples": examples[:5],
        "occurrences": max(1, int(occurrences)),
        "created_at": _utc_now_iso(),
        "updated_at": _utc_now_iso(),
    }
    return candidate


def _prompt_dictionary_candidate_id(kind: str, canonical: str, value: str, source: str) -> str:
    return sha1(f"{kind}|{canonical}|{value}|{source}".encode("utf-8")).hexdigest()[:16]


def _extract_generic_candidate_tokens(text: str) -> List[str]:
    normalized = _normalize_text(text)
    tokens = []
    for token in _extract_query_tokens(normalized):
        if _looks_like_pii_value(token):
            continue
        if len(token) < 2 or len(token) > 24:
            continue
        tokens.append(token)
    return tokens


def _classify_generic_stopword_seed(token: str) -> str:
    if token in _PROMPT_DICTIONARY_STOPWORD_SEEDS["strong"]:
        return "stopword"
    if token in _PROMPT_DICTIONARY_STOPWORD_SEEDS["weak"]:
        return "weak_stopword"
    return ""


def _looks_like_pii_field(field_name: Any) -> bool:
    normalized = _normalize_text(field_name)
    return any(pattern in normalized for pattern in _PROMPT_DICTIONARY_PII_FIELD_PATTERNS)


def _looks_like_pii_value(value: Any) -> bool:
    text = _safe_str(value).strip()
    if not text:
        return False
    if "@" in text:
        return True
    if re.search(r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b", text, flags=re.IGNORECASE):
        return True
    return False


def _prefer_as_canonical(left: str, right: str) -> bool:
    if len(left) <= 12 and any(ch.isascii() and ch.isalpha() for ch in left) and not any('\u3040' <= ch <= '\u30ff' or '\u4e00' <= ch <= '\u9fff' for ch in left):
        return True
    if len(right) <= 12 and any(ch.isascii() and ch.isalpha() for ch in right) and not any('\u3040' <= ch <= '\u30ff' or '\u4e00' <= ch <= '\u9fff' for ch in right):
        return False
    return len(left) <= len(right)


def _script_kind(text: str) -> str:
    if any('\u3040' <= ch <= '\u30ff' or '\u4e00' <= ch <= '\u9fff' for ch in text):
        return "ja"
    if any(ch.isascii() and ch.isalpha() for ch in text):
        return "latin"
    return "other"


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _normalize_text(text: str) -> str:
    value = unicodedata.normalize("NFKC", _safe_str(text)).lower()
    value = re.sub(r"[-_/\\|]+", " ", value)
    value = re.sub(r"[()\[\]{}:,;]+", " ", value)
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def _extract_query_tokens(query_norm: str) -> List[str]:
    tokens: List[str] = []
    for token in _ALNUM_TOKEN_PATTERN.findall(query_norm):
        if len(token) >= 2:
            tokens.append(token)
        slash_parts = [part for part in re.split(r"[+./:_-]", token) if len(part) >= 2]
        tokens.extend(slash_parts)
    for token in _JA_TOKEN_PATTERN.findall(query_norm):
        if len(token) >= 2:
            tokens.append(token)
            if len(token) > 6:
                for index in range(0, min(len(token) - 1, 8)):
                    fragment = token[index:index + 4]
                    if len(fragment) >= 2:
                        tokens.append(fragment)
    for token in _MIXED_TOKEN_PATTERN.findall(query_norm):
        if len(token) >= 2:
            tokens.append(token)
    return list(dict.fromkeys(tokens))[:120]


def _build_selected_summary(entries: Iterable[Dict[str, Any]]) -> List[str]:
    summary = []
    for entry in list(entries)[:5]:
        path = entry.get("path") or []
        if path:
            summary.append(" > ".join(path))
        else:
            summary.append(_safe_str(entry.get("label")))
    return summary


def _build_match_preview(scored_entries: List[Tuple[int, Dict[str, Any]]], max_items: int = 8) -> List[Dict[str, Any]]:
    preview: List[Dict[str, Any]] = []
    for item in scored_entries[:max_items]:
        if isinstance(item, dict):
            score = item.get("score", 0)
            entry = item.get("entry") or {}
            breakdown = item.get("breakdown") or {}
            matched_tokens = item.get("matched_tokens") or []
        else:
            score, entry = item
            breakdown = {}
            matched_tokens = []
        path = list(entry.get("path") or [])
        preview.append(
            {
                "score": score,
                "label": entry.get("label") or "",
                "path": path,
                "path_text": " > ".join(path) if path else _safe_str(entry.get("label")),
                "score_breakdown": breakdown,
                "matched_tokens": matched_tokens[:8],
            }
        )
    return preview


def _truncate_diagnostic_text(text: str, max_chars: int) -> str:
    value = _safe_str(text).strip()
    if max_chars <= 0 or len(value) <= max_chars:
        return value
    return value[: max_chars - 1] + "…"


def _build_request_id(seed: str) -> str:
    now = datetime.now(timezone.utc).astimezone().strftime("%Y%m%d_%H%M%S_%f")
    digest = sha1(_safe_str(seed).encode("utf-8")).hexdigest()[:8]
    return f"{now}_{digest}"


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    try:
        return str(value)
    except Exception:
        return ""


def _log_build_summary(diagnostics: Dict[str, Any]) -> None:
    if diagnostics.get("fallback_used"):
        logger.warning(
            "prompt build fallback feature=%s template=%s mode=%s reasons=%s",
            diagnostics.get("feature_id"),
            diagnostics.get("template_name"),
            diagnostics.get("mode"),
            [item.get("fallback_reason") for item in diagnostics.get("source_diagnostics") or [] if item.get("fallback_used")],
        )
    logger.info(
        "prompt build feature=%s template=%s mode=%s prompt_chars=%s prompt_tokens=%s fallback=%s sources=%s elapsed=%s",
        diagnostics.get("feature_id"),
        diagnostics.get("template_name"),
        diagnostics.get("mode"),
        diagnostics.get("prompt_chars"),
        diagnostics.get("prompt_token_estimate"),
        diagnostics.get("fallback_used"),
        diagnostics.get("source_summary"),
        diagnostics.get("build_elapsed_seconds"),
    )


def _append_event(record: Dict[str, Any]) -> None:
    if os.environ.get("PYTEST_CURRENT_TEST"):
        return
    try:
        path = get_dynamic_file_path("output/ai_prompt_diagnostics/events.jsonl")
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "a", encoding="utf-8") as handle:
            handle.write(json.dumps({**record, "timestamp": _now_iso()}, ensure_ascii=False) + "\n")
    except Exception:
        logger.debug("prompt diagnostics event write failed", exc_info=True)


def _save_debug_artifacts_if_needed(
    *,
    diagnostics: Dict[str, Any],
    template_text: str,
    prompt_text: str,
    source_diagnostics: List[Dict[str, Any]],
) -> None:
    if not diagnostics.get("debug_save_enabled") or os.environ.get("PYTEST_CURRENT_TEST"):
        return
    try:
        debug_dir = _ensure_debug_dir(diagnostics)
        with open(os.path.join(debug_dir, "template.txt"), "w", encoding="utf-8") as handle:
            handle.write(template_text)
        with open(os.path.join(debug_dir, "prompt.txt"), "w", encoding="utf-8") as handle:
            handle.write(prompt_text)
        with open(os.path.join(debug_dir, "build.json"), "w", encoding="utf-8") as handle:
            json.dump({**diagnostics, "source_diagnostics": source_diagnostics}, handle, ensure_ascii=False, indent=2)
        diagnostics["debug_dir"] = debug_dir
    except Exception:
        logger.debug("prompt debug artifact write failed", exc_info=True)


def _save_response_artifacts_if_needed(
    prompt_diagnostics: Dict[str, Any],
    *,
    result: Optional[Dict[str, Any]] = None,
    error: Optional[str] = None,
) -> None:
    if not prompt_diagnostics.get("debug_save_enabled") or os.environ.get("PYTEST_CURRENT_TEST"):
        return
    try:
        debug_dir = prompt_diagnostics.get("debug_dir") or _ensure_debug_dir(prompt_diagnostics)
        payload = {
            "success": bool(isinstance(result, dict) and result.get("success")),
            "provider": (result or {}).get("provider") if isinstance(result, dict) else None,
            "model": (result or {}).get("model") if isinstance(result, dict) else None,
            "started_at": (result or {}).get("started_at") if isinstance(result, dict) else None,
            "finished_at": (result or {}).get("finished_at") if isinstance(result, dict) else None,
            "elapsed_seconds": (result or {}).get("elapsed_seconds") if isinstance(result, dict) else None,
            "error": error or ((result or {}).get("error") if isinstance(result, dict) else None),
            "request_params": (result or {}).get("request_params") if isinstance(result, dict) else None,
            "response_params": (result or {}).get("response_params") if isinstance(result, dict) else None,
        }
        with open(os.path.join(debug_dir, "result.json"), "w", encoding="utf-8") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
        response_text = _safe_str((result or {}).get("response") or (result or {}).get("content"))
        with open(os.path.join(debug_dir, "response.txt"), "w", encoding="utf-8") as handle:
            handle.write(response_text)
    except Exception:
        logger.debug("prompt response artifact write failed", exc_info=True)


def _append_comparison_report(prompt_diagnostics: Dict[str, Any], result_record: Dict[str, Any]) -> None:
    if os.environ.get("PYTEST_CURRENT_TEST"):
        return
    if not prompt_diagnostics.get("comparison_report_enabled", True):
        return
    try:
        rows = _build_comparison_rows(prompt_diagnostics, result_record)
        if not rows:
            return
        jsonl_path = get_dynamic_file_path("output/ai_prompt_diagnostics/comparisons.jsonl")
        os.makedirs(os.path.dirname(jsonl_path), exist_ok=True)
        with open(jsonl_path, "a", encoding="utf-8") as handle:
            for row in rows:
                handle.write(json.dumps({**row, "timestamp": _now_iso()}, ensure_ascii=False) + "\n")

        markdown_path = get_dynamic_file_path("output/ai_prompt_diagnostics/comparisons.md")
        markdown_exists = os.path.exists(markdown_path)
        with open(markdown_path, "a", encoding="utf-8") as handle:
            if not markdown_exists:
                handle.write("| feature name | template name | source placeholder | prompt mode | raw candidate count | filtered candidate count | final prompt chars | estimated full prompt chars | fallback | request duration | success |\n")
                handle.write("| --- | --- | --- | --- | ---: | ---: | ---: | ---: | --- | ---: | --- |\n")
            for row in rows:
                handle.write(
                    f"| {row.get('feature_name','')} | {row.get('template_name','')} | {row.get('source_placeholder_name','')} | {row.get('prompt_mode','')} | {row.get('raw_candidate_count',0)} | {row.get('filtered_candidate_count',0)} | {row.get('final_prompt_char_count',0)} | {row.get('estimated_full_prompt_char_count',0)} | {row.get('fallback_occurred',False)} | {row.get('request_duration',0) or 0} | {row.get('response_success',False)} |\n"
                )
    except Exception:
        logger.debug("comparison report write failed", exc_info=True)


def _build_comparison_rows(prompt_diagnostics: Dict[str, Any], result_record: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows = []
    for source_diag in prompt_diagnostics.get("source_diagnostics") or []:
        rows.append(
            {
                "feature_name": prompt_diagnostics.get("feature_id", ""),
                "template_name": prompt_diagnostics.get("template_name", ""),
                "source_placeholder_name": source_diag.get("placeholder", ""),
                "prompt_mode": source_diag.get("mode", prompt_diagnostics.get("mode", "")),
                "raw_candidate_count": (source_diag.get("candidate_pool_summary") or {}).get("raw_candidate_count", source_diag.get("candidate_pool_size", 0)),
                "filtered_candidate_count": (source_diag.get("candidate_pool_summary") or {}).get("selected_candidate_count", source_diag.get("selected_candidates", 0)),
                "final_prompt_char_count": prompt_diagnostics.get("prompt_chars", 0),
                "estimated_full_prompt_char_count": prompt_diagnostics.get("estimated_full_prompt_chars", 0),
                "estimated_tokens": prompt_diagnostics.get("prompt_token_estimate", 0),
                "fallback_occurred": bool(source_diag.get("fallback_used")),
                "fallback_reason": source_diag.get("fallback_reason"),
                "request_duration": result_record.get("elapsed_seconds"),
                "response_success": result_record.get("success", False),
                "response_error": result_record.get("error"),
                "selected_summary": source_diag.get("selected_summary") or [],
            }
        )
    return rows


def _ensure_debug_dir(diagnostics: Dict[str, Any]) -> str:
    feature_id = _safe_filename(_safe_str(diagnostics.get("feature_id") or diagnostics.get("template_name") or "prompt"))
    request_id = _safe_filename(_safe_str(diagnostics.get("request_id") or "request"))
    path = get_dynamic_file_path(f"output/ai_prompt_diagnostics/debug/{feature_id}/{request_id}")
    os.makedirs(path, exist_ok=True)
    return path


def _safe_filename(text: str) -> str:
    value = _safe_str(text).strip() or "unknown"
    for char in ['<', '>', ':', '"', '/', '\\', '|', '?', '*']:
        value = value.replace(char, '_')
    value = value.strip('._ ')
    return value[:180] if len(value) > 180 else value


def _now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
