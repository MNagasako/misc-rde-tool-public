"""
XLSX書き出し専用ロジック
"""
import os
import time
import logging
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Set

import openpyxl
from dateutil.parser import parse as parse_datetime

from classes.basic.conf.summary_export_options import SummaryExportMode, SummaryExportOptions
from config.common import (
    get_dynamic_file_path,
    INPUT_DIR,
    OUTPUT_DIR,
    OUTPUT_RDE_DIR,
    DATAFILES_DIR,
    SUBGROUP_JSON_PATH,
    GROUP_ORGNIZATION_DIR,
    DATASET_JSON_PATH,
    INSTRUMENTS_JSON_PATH,
)

# ロガー設定
logger = logging.getLogger(__name__)


@dataclass
class GroupPayload:
    path: Path
    included: List[dict]
    group_ids: Set[str]
    display_name: str


@dataclass
class SummaryExportJob:
    label: str
    output_path: str
    included_items: List[dict]
    allowed_group_ids: Optional[Set[str]] = None


def _load_group_payloads(
    additional_paths: Optional[Sequence[str]] = None,
    selected_default_files: Optional[Sequence[str]] = None,
) -> List[GroupPayload]:
    payloads: List[GroupPayload] = []
    candidate_paths: List[Path] = []

    group_dir = Path(GROUP_ORGNIZATION_DIR)
    selected_file_map: Optional[Set[str]] = None
    if selected_default_files:
        selected_file_map = {Path(path).name.lower() for path in selected_default_files if path}

    if group_dir.exists():
        for file_path in sorted(group_dir.glob("*.json")):
            if selected_file_map and file_path.name.lower() not in selected_file_map:
                continue
            candidate_paths.append(file_path)

    if additional_paths:
        for raw_path in additional_paths:
            if not raw_path:
                continue
            candidate_paths.append(Path(raw_path))

    if not candidate_paths:
        return payloads

    seen_paths: Set[str] = set()
    for json_path in candidate_paths:
        normalized = os.path.abspath(str(json_path))
        key = normalized.lower()
        if key in seen_paths:
            continue
        seen_paths.add(key)
        if not os.path.exists(normalized):
            logger.warning("[XLSX] プロジェクトファイルが見つかりません: %s", normalized)
            continue
        data = load_json(normalized)
        if not data:
            continue
        included = data.get("included", []) or []
        group_ids = {
            item.get("id")
            for item in included
            if isinstance(item, dict) and item.get("type") == "group" and item.get("id")
        }
        display_name = next(
            (
                item.get("attributes", {}).get("name")
                for item in included
                if isinstance(item, dict)
                and item.get("type") == "group"
                and item.get("attributes", {}).get("name")
            ),
            Path(normalized).stem,
        )
        payloads.append(
            GroupPayload(
                path=Path(normalized),
                included=list(included),
                group_ids=set(group_ids),
                display_name=display_name,
            )
        )
    return payloads


def _combine_included(payloads: Sequence[GroupPayload]) -> List[dict]:
    combined: List[dict] = []
    for payload in payloads:
        combined.extend(payload.included)
    return combined


def _build_summary_jobs(
    options: SummaryExportOptions,
    default_output_path: str,
    payloads: Sequence[GroupPayload],
) -> List[SummaryExportJob]:
    jobs: List[SummaryExportJob] = []
    used_labels: Set[str] = set()

    def _make_output_path(base_label: str) -> str:
        label = base_label
        counter = 1
        lowered = label.lower()
        while lowered in used_labels:
            label = f"{base_label}_{counter}"
            lowered = label.lower()
            counter += 1
        used_labels.add(lowered)
        return os.path.abspath(os.path.join(OUTPUT_DIR, f"{label}.xlsx"))

    if not payloads:
        jobs.append(
            SummaryExportJob(
                label="summary",
                output_path=default_output_path,
                included_items=[],
                allowed_group_ids=None,
            )
        )
        return jobs

    mode = options.mode
    if mode == SummaryExportMode.PER_FILE:
        for payload in payloads:
            suffix = SummaryExportOptions.sanitize_suffix(payload.display_name) or payload.path.stem
            base_label = f"summary_{suffix}" if suffix else f"summary_{payload.path.stem}"
            jobs.append(
                SummaryExportJob(
                    label=base_label,
                    output_path=_make_output_path(base_label),
                    included_items=list(payload.included),
                    allowed_group_ids=None,
                )
            )
        return jobs

    if mode == SummaryExportMode.CUSTOM_SELECTION:
        target_ids = [gid for gid in options.selected_group_ids if gid]
        if target_ids:
            target_set = set(target_ids)
            relevant_payloads = [payload for payload in payloads if payload.group_ids & target_set]
            if relevant_payloads:
                suffix = options.custom_suffix or "selection"
                suffix = SummaryExportOptions.sanitize_suffix(suffix) or "selection"
                base_label = f"summary_{suffix}"
                jobs.append(
                    SummaryExportJob(
                        label=base_label,
                        output_path=_make_output_path(base_label),
                        included_items=_combine_included(relevant_payloads),
                        allowed_group_ids=target_set,
                    )
                )
                return jobs
        # Fallback to merged output if selection is empty or not found.
        mode = SummaryExportMode.MERGED

    if mode == SummaryExportMode.MERGED:
        jobs.append(
            SummaryExportJob(
                label="summary",
                output_path=_make_output_path("summary"),
                included_items=_combine_included(payloads),
                allowed_group_ids=None,
            )
        )
        return jobs

    # Safety fallback
    jobs.append(
        SummaryExportJob(
            label="summary",
            output_path=default_output_path,
            included_items=_combine_included(payloads),
            allowed_group_ids=None,
        )
    )
    return jobs


def _ensure_summary_workbook(abs_path: str):
    if os.path.exists(abs_path):
        return
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "概要"
    wb.create_sheet("データ")
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    wb.save(abs_path)


def _is_xlsx_writable(path: str) -> bool:
    try:
        with open(path, "a"):
            pass
        return True
    except PermissionError:
        return False
    except Exception:
        # Treat other errors (e.g., file missing) as writable so caller can handle.
        return True


def _wrap_job_progress(job_index: int, job_total: int, callback):
    if not callback:
        return None

    def _wrapped(current: int, total: int, message: str):
        prefixed = f"[{job_index}/{job_total}] {message}"
        return callback(current, total, prefixed)

    return _wrapped


def _run_single_summary_job(
    job: SummaryExportJob,
    parent,
    dataset_data: List[dict],
    instruments_data: List[dict],
    progress_callback,
):
    from qt_compat.widgets import QMessageBox

    abs_xlsx = job.output_path
    _ensure_summary_workbook(abs_xlsx)

    def emit_progress(current: int, total: int, message: str) -> bool:
        if progress_callback:
            return progress_callback(current, total, message)
        return True

    def show_retry_dialog_sync():
        msg = (
            "Excelファイルが他で開かれているため書き込みできません。\n"
            "Excel等を閉じてから[再開]を押してください。\n"
            f"対象: {abs_xlsx}"
        )
        mbox = QMessageBox(parent)
        mbox.setIcon(QMessageBox.Warning)
        mbox.setWindowTitle("書き込みエラー")
        mbox.setText(msg)
        retry_btn = mbox.addButton("再開", QMessageBox.AcceptRole)
        cancel_btn = mbox.addButton("キャンセル", QMessageBox.RejectRole)
        mbox.setDefaultButton(retry_btn)
        mbox.exec()
        return "cancel" if mbox.clickedButton() == cancel_btn else "retry"

    while True:
        if not os.path.exists(abs_xlsx) or _is_xlsx_writable(abs_xlsx):
            try:
                wb = openpyxl.load_workbook(abs_xlsx)
                summary_context = {
                    "subGroup_included": job.included_items,
                    "dataset_data": dataset_data,
                    "instruments_data": instruments_data,
                    "allowed_group_ids": job.allowed_group_ids,
                }
                write_summary_sheet(
                    wb,
                    parent,
                    False,
                    progress_callback if progress_callback else None,
                    summary_context=summary_context,
                )

                if not emit_progress(0, 1, "各シートの書き出しを実行中..."):
                    return False

                write_members_sheet(wb, parent)
                write_organization_sheet(wb, parent)
                write_instrumentType_sheet(wb, parent)
                write_datasets_sheet(wb, parent)
                write_subgroups_sheet(wb, parent)
                write_groupDetail_sheet(wb, parent)
                write_templates_sheet(wb, parent)
                write_instruments_sheet(wb, parent)
                write_licenses_sheet(wb, parent)
                write_entries_sheet(wb, parent)

                if not emit_progress(0, 1, "ファイル保存中..."):
                    return False
                wb.save(abs_xlsx)
                emit_progress(1, 1, "XLSX書き出し完了")
                return True
            except PermissionError:
                msg = (
                    "Excelファイルが他で開かれているため書き込みできません。\n"
                    "Excelを閉じてから再実行してください。\n"
                    f"対象: {abs_xlsx}"
                )
                QMessageBox.critical(parent, "書き込みエラー", msg)
            except Exception as e:  # noqa: BLE001
                import traceback

                tb = traceback.format_exc()
                msg = f"XLSX書き込み時に予期せぬエラーが発生しました:\n{e}\n{tb}"
                QMessageBox.critical(parent, "書き込みエラー", msg)
                return False

        result = show_retry_dialog_sync()
        if result == "cancel":
            QMessageBox.critical(parent, "書き込みエラー", "書き込みを中止しました。")
            return False

def load_json(path):
    import json
    abs_path = os.path.abspath(path)
    if not os.path.exists(abs_path):
        logger.error("%sが存在しません: %s", path, abs_path)
        return None
    with open(abs_path, "r", encoding="utf-8") as f:
        logger.info("[XLSX] JSONロード成功: %s", abs_path)
        return json.load(f)
def apply_basic_info_to_Xlsx_logic(bearer_token, parent=None, webview=None, ui_callback=None):
    """
    各種JSONを読み込み、XLSXの対応シートに反映（責務分離構造）
    """
    XLSX_PATH = os.path.join(INPUT_DIR, "data.xlsm")
    abs_xlsx = os.path.abspath(XLSX_PATH)
    if not os.path.exists(abs_xlsx):
        msg = f"XLSXファイルが存在しません: {abs_xlsx}"
        print(msg)
        return
    try:
        wb = openpyxl.load_workbook(abs_xlsx, keep_vba=True)
        # ... ここに各種JSONを読み込んでシートに反映する処理 ...
        wb.save(abs_xlsx)
        logger.info("XLSX書き出しに成功: %s", abs_xlsx)
    except PermissionError:
        msg = (
            f"Excelファイルが他で開かれているため書き込みできません。\n"
            f"他のExcelウィンドウが開いていないか確認し、すべて閉じてから再実行してください。\n"
            f"対象: {abs_xlsx}"
        )
        try:
            from qt_compat.widgets import QMessageBox
            QMessageBox.information(None, "Excel書き込みエラー", msg)
        except Exception:
            print(msg)
        raise
    except Exception as e:
        msg = f"XLSX書き込み時に予期せぬエラーが発生しました: {e}"
        try:
            from qt_compat.widgets import QMessageBox
            QMessageBox.information(None, "Excel書き込みエラー", msg)
        except Exception:
            print(msg)
        raise


def summary_basic_info_to_Xlsx_logic(
    bearer_token,
    parent=None,
    webview=None,
    ui_callback=None,
    progress_callback=None,
    export_options=None,
):
    """Generate summary.xlsx files based on the selected export mode."""

    from qt_compat.widgets import QMessageBox

    options = SummaryExportOptions.from_payload(export_options).with_sanitized_suffix()

    default_xlsx = os.path.abspath(os.path.join(OUTPUT_DIR, "summary.xlsx"))
    group_payloads = _load_group_payloads(
        additional_paths=options.extra_project_files,
        selected_default_files=options.project_files or None,
    )
    jobs = _build_summary_jobs(options, default_xlsx, group_payloads)
    total_jobs = len(jobs)
    if total_jobs == 0:
        jobs = _build_summary_jobs(SummaryExportOptions(), default_xlsx, group_payloads)
        total_jobs = len(jobs)

    dataset_json = load_json(DATASET_JSON_PATH)
    instruments_json = load_json(INSTRUMENTS_JSON_PATH)
    if not dataset_json or not instruments_json:
        msg = "必要なJSONデータが不足しています。基本情報を取得してから再実行してください。"
        QMessageBox.critical(parent, "データ不足", msg)
        return False

    dataset_data = dataset_json.get("data", [])
    instruments_data = instruments_json.get("data", [])

    if total_jobs == 0:
        QMessageBox.information(parent, "処理対象なし", "出力対象となるサブグループが見つかりませんでした。")
        return False

    results: List[str] = []
    for idx, job in enumerate(jobs, start=1):
        job_progress = _wrap_job_progress(idx, total_jobs, progress_callback)
        if job_progress:
            if not job_progress(0, 1, f"{job.label} | XLSX書き出しを開始しています..."):
                return "キャンセルされました"

        success = _run_single_summary_job(
            job,
            parent,
            dataset_data,
            instruments_data,
            job_progress,
        )
        if not success:
            return False
        results.append(job.output_path)

    if results:
        display_paths = []
        for path in results:
            try:
                display_paths.append(os.path.relpath(path, OUTPUT_DIR))
            except ValueError:
                display_paths.append(path)
        return "出力完了: " + ", ".join(display_paths)

    return True


def write_summary_sheet(wb, parent, load_data_entry_json=False, progress_callback=None, summary_context=None):
    import json, os
    logger.info("write_summary_sheet called start")
    load_data_entry_json=False
    summary_context = summary_context or {}
    context_subgroups = summary_context.get("subGroup_included") if isinstance(summary_context, dict) else None
    context_dataset = summary_context.get("dataset_data") if isinstance(summary_context, dict) else None
    context_instruments = summary_context.get("instruments_data") if isinstance(summary_context, dict) else None
    allowed_group_ids = summary_context.get("allowed_group_ids") if isinstance(summary_context, dict) else None
    if allowed_group_ids:
        allowed_group_ids = set(allowed_group_ids)
    else:
        allowed_group_ids = None
    
    if progress_callback:
        if not progress_callback(0, 1, "JSONファイル読み込み中..."):
            return False
    def to_ymd(date_str):
        if not date_str:
            return ""
        try:
            dt = parse_datetime(date_str)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return date_str

    def get_dataset_related_info(dataset_datum):
        attr = dataset_datum.get("attributes", {})
        rel = dataset_datum.get("relationships", {})
        return {
            "id": dataset_datum.get("id", ""),
            "manager_id": rel.get("manager", {}).get("data", {}).get("id", ""),
            "owners": rel.get("dataOwners", {}).get("data", []),
            "applicant_id": rel.get("applicant", {}).get("data", {}).get("id", ""),
            "template_id": rel.get("template", {}).get("data", {}).get("id", ""),
            "instrument_id": rel.get("instruments", {}).get("data", [{}])[0].get("id", "") if rel.get("instruments", {}).get("data") else "",
            "embargoDate": to_ymd(attr.get("embargoDate", "")),
            "isAnonymized": attr.get("isAnonymized", ""),
            "description": attr.get("description", ""),
            "relatedLinks_str": "\n".join([link.get("url", "") for link in attr.get("relatedLinks", []) if isinstance(link, dict)]),
            "relatedDatasets_urls_str": "\n".join([f"https://rde.nims.go.jp/datasets/rde/{rd.get('id', '')}" for rd in rel.get("relatedDatasets", {}).get("data", []) if isinstance(rd, dict)]),
            "grantNumber": attr.get("grantNumber", ""),
            "name": attr.get("name", ""),
            "title": attr.get("subjectTitle", ""),
        }

    def write_row(value_dict, row_idx):
        dataset_id = value_dict.get("datasetId", "")
        data_entry_id = value_dict.get("dataEntryId", "")
        if data_entry_id and ("dataEntryId", data_entry_id) in manual_data_map:
            manual_restore = manual_data_map[("dataEntryId", data_entry_id)]
        elif dataset_id and ("datasetId", dataset_id) in manual_data_map:
            manual_restore = manual_data_map[("datasetId", dataset_id)]
        else:
            manual_restore = {}
        for id_ in header_ids:
            col = id_to_col[id_]
            if id_ in value_dict:
                ws.cell(row=row_idx, column=col, value=value_dict[id_])
            elif id_ in manual_restore:
                ws.cell(row=row_idx, column=col, value=manual_restore[id_])
        for id_ in value_dict:
            if id_ not in header_ids:
                header_ids.append(id_)
                col = len(header_ids)
                ws.cell(row=1, column=col, value=id_)
                ws.cell(row=2, column=col, value=id_to_label.get(id_, id_))
                ws.cell(row=row_idx, column=col, value=value_dict[id_])
                id_to_col[id_] = col

    import json, os
    logger.debug("[XLSX] write_summary_sheet called progress")

    # ワークシート取得
    ws = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else wb.active

    # 必要なデータのロード（parentから渡される想定。なければ適宜修正）
    # 例: subGroup_included, dataset_data, user_id_to_name, instrument_id_to_name, instrument_id_to_localid, manual_data_map, id_to_label
    # ここでは仮にparentから取得する例
    subGroup_included = context_subgroups if context_subgroups is not None else getattr(parent, 'subGroup_included', [])
    dataset_data = context_dataset if context_dataset is not None else getattr(parent, 'dataset_data', [])
    user_id_to_name = getattr(parent, 'user_id_to_name', {})
    instruments_data = context_instruments if context_instruments is not None else getattr(parent, 'instruments_data', [])
    instrument_id_to_name = getattr(parent, 'instrument_id_to_name', {})
    instrument_id_to_localid = getattr(parent, 'instrument_id_to_localid', {})
    manual_data_map = getattr(parent, 'manual_data_map', {})
    id_to_label = getattr(parent, 'id_to_label', {})

    header_ids = [
        "subGroupName", "dataset_manager_name", "dataset_applicant_name", "dataset_owner_names_str",
        "grantNumber", "title", "datasetName", "instrument_name", "instrument_local_id", "template_id",
        "datasetId", "dataEntryName", "dataEntryId", "number_of_files", "number_of_image_files",
        "date_of_dataEntry_creation", "total_file_size_MB", "dataset_embargoDate", "dataset_isAnonymized",
        "dataset_description", "dataset_relatedLinks", "dataset_relatedDatasets"
    ]
    id_to_col = {id_: idx+1 for idx, id_ in enumerate(header_ids)}

    def to_ymd(date_str):
        if not date_str:
            return ""
        try:
            dt = parse_datetime(date_str)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return date_str

    def get_dataset_related_info(dataset_datum):
        attr = dataset_datum.get("attributes", {})
        rel = dataset_datum.get("relationships", {})
        return {
            "id": dataset_datum.get("id", ""),
            "manager_id": rel.get("manager", {}).get("data", {}).get("id", ""),
            "owners": rel.get("dataOwners", {}).get("data", []),
            "applicant_id": rel.get("applicant", {}).get("data", {}).get("id", ""),
            "template_id": rel.get("template", {}).get("data", {}).get("id", ""),
            "instrument_id": rel.get("instruments", {}).get("data", [{}])[0].get("id", "") if rel.get("instruments", {}).get("data") else "",
            "embargoDate": to_ymd(attr.get("embargoDate", "")),
            "isAnonymized": attr.get("isAnonymized", ""),
            "description": attr.get("description", ""),
            "relatedLinks_str": "\n".join([link.get("url", "") for link in attr.get("relatedLinks", []) if isinstance(link, dict)]),
            "relatedDatasets_urls_str": "\n".join([f"https://rde.nims.go.jp/datasets/rde/{rd.get('id', '')}" for rd in rel.get("relatedDatasets", {}).get("data", []) if isinstance(rd, dict)]),
            "grantNumber": attr.get("grantNumber", ""),
            "name": attr.get("name", ""),
            "title": attr.get("subjectTitle", ""),
        }

    def write_row(value_dict, row_idx):
        dataset_id = value_dict.get("datasetId", "")
        data_entry_id = value_dict.get("dataEntryId", "")
        if data_entry_id and ("dataEntryId", data_entry_id) in manual_data_map:
            manual_restore = manual_data_map[("dataEntryId", data_entry_id)]
        elif dataset_id and ("datasetId", dataset_id) in manual_data_map:
            manual_restore = manual_data_map[("datasetId", dataset_id)]
        else:
            manual_restore = {}
        for id_ in header_ids:
            col = id_to_col[id_]
            if id_ in value_dict:
                ws.cell(row=row_idx, column=col, value=value_dict[id_])
            elif id_ in manual_restore:
                ws.cell(row=row_idx, column=col, value=manual_restore[id_])
        for id_ in value_dict:
            if id_ not in header_ids:
                header_ids.append(id_)
                col = len(header_ids)
                ws.cell(row=1, column=col, value=id_)
                ws.cell(row=2, column=col, value=id_to_label.get(id_, id_))
                ws.cell(row=row_idx, column=col, value=value_dict[id_])
                id_to_col[id_] = col

    # ヘッダー行の書き込み
    for idx, id_ in enumerate(header_ids):
        ws.cell(row=1, column=idx+1, value=id_)
        ws.cell(row=2, column=idx+1, value=id_to_label.get(id_, id_))

    row_idx = 3
    # --- 進捗集計用 ---
    # 事前に全データセットのdataEntry JSONファイル数・エントリ数を集計
    entry_json_files = []
    entry_total_count = 0
    for dataset in dataset_data:
        ds_info = get_dataset_related_info(dataset)
        dataEntry_path = os.path.join(OUTPUT_RDE_DIR, "data", "dataEntry", f"{ds_info['id']}.json")
        if load_data_entry_json and os.path.exists(dataEntry_path):
            entry_json_files.append(dataEntry_path)
            try:
                entry_json = load_json(dataEntry_path)
                entry_count = len(entry_json.get("data", []))
            except Exception:
                entry_count = 0
            entry_total_count += entry_count
        else:
            entry_total_count += 1
    logger.debug("[PROGRESS] 全データセット: %s、dataEntry JSONファイル: %s、エントリ総数: %s", len(dataset_data), len(entry_json_files), entry_total_count)
    processed_entries = 0
    start_time = time.time()
    for subGroup in subGroup_included:
        if subGroup.get("type") != "group":
            continue
        if allowed_group_ids and subGroup.get("id") not in allowed_group_ids:
            continue
        subGroup_attr = subGroup.get("attributes", {})
        subGroup_name = subGroup_attr.get("name", "")
        subGroup_subjects = subGroup_attr.get("subjects", {})
        for subject in subGroup_subjects:
            grantNumber = subject.get("grantNumber", "") if isinstance(subject, dict) else ""
            title = subject.get("title", "") if isinstance(subject, dict) else ""
            for dataset in dataset_data:
                ds_info = get_dataset_related_info(dataset)
                if ds_info["grantNumber"] != grantNumber:
                    continue
                manager_name = user_id_to_name.get(ds_info["manager_id"], "未設定" if ds_info["manager_id"] in [None, ""] else "")
                applicant_name = user_id_to_name.get(ds_info["applicant_id"], "")
                owner_names = [user_id_to_name.get(owner.get("id", ""), "") for owner in ds_info["owners"] if owner.get("id", "")]
                owner_names_str = "\n".join([n for n in owner_names if n])
                instrument_name = instrument_id_to_name.get(ds_info["instrument_id"], "")
                instrument_local_id = instrument_id_to_localid.get(ds_info["instrument_id"], "")
                dataset_url = f"https://rde.nims.go.jp/datasets/rde/{ds_info['id']}"
                total_file_size_MB = ""
                dataEntry_data = []
                dataEntry_included = []
                if load_data_entry_json == True:
                    logger.debug("[XLSX] dataEntry JSONロード for subGroup: %s", ds_info['id'])
                    dataEntry_path = os.path.join(OUTPUT_RDE_DIR, "data", "dataEntry", f"{ds_info['id']}.json")
                    dataEntry_json = load_json(dataEntry_path)
                    if not dataEntry_json:
                        logger.error("dataEntry JSONが存在しません: %s for dataset_id=%s", dataEntry_path, ds_info['id'])
                        continue
                    dataEntry_data = dataEntry_json.get("data", [])
                    dataEntry_included = dataEntry_json.get("included", [])
                    total_file_size = sum(
                        inc.get("attributes", {}).get("fileSize", 0)
                        for inc in dataEntry_included if inc.get("type") == "file"
                    )
                    total_file_size_MB = total_file_size / (1024 * 1024) if total_file_size else 0
                if dataEntry_data:
                    for entry in dataEntry_data:
                        entry_attr = entry.get("attributes", {})
                        dataEntry_name = entry_attr.get("name", "")
                        dataEntry_id = entry.get("id", "")
                        number_of_files = entry_attr.get("numberOfFiles", "")
                        number_of_image_files = entry_attr.get("numberOfImageFiles", "")
                        date_of_dataEntry_creation = entry_attr.get("created", "")
                        value_dict = {
                            "subGroupName": subGroup_name,
                            "dataset_manager_name": manager_name,
                            "dataset_applicant_name": applicant_name,
                            "dataset_owner_names_str": owner_names_str,
                            "grantNumber": grantNumber,
                            "title": title,
                            "datasetName": ds_info["name"],
                            "instrument_name": instrument_name,
                            "instrument_local_id": instrument_local_id,
                            "template_id": ds_info["template_id"],
                            "datasetId": dataset_url,
                            "dataEntryName": dataEntry_name,
                            "dataEntryId": dataEntry_id,
                            "number_of_files": number_of_files,
                            "number_of_image_files": number_of_image_files,
                            "date_of_dataEntry_creation": to_ymd(date_of_dataEntry_creation),
                            "total_file_size_MB": total_file_size_MB,
                            "dataset_embargoDate": ds_info["embargoDate"],
                            "dataset_isAnonymized": ds_info["isAnonymized"],
                            "dataset_description": ds_info["description"],
                            "dataset_relatedLinks": ds_info["relatedLinks_str"],
                            "dataset_relatedDatasets": ds_info["relatedDatasets_urls_str"],
                            # --- ファイルタイプ集計関連 ---
                            "filetype_MAIN_IMAGE_count": 0,
                            "filetype_MAIN_IMAGE_size": 0,
                            "filetype_STRUCTURED_count": 0,
                            "filetype_STRUCTURED_size": 0,
                            "filetype_THUMBNAIL_count": 0,
                            "filetype_THUMBNAIL_size": 0,
                            "filetype_META_count": 0,
                            "filetype_META_size": 0,
                            "filetype_OTHER_count": 0,
                            "filetype_OTHER_size": 0,
                            "filetype_total_count": 0,
                            "filetype_total_size": 0,
                        }
                        # ...existing code...
                        processed_entries += 1
                        if processed_entries % 10 == 0 or processed_entries == entry_total_count:
                            elapsed = time.time() - start_time
                            avg_time = elapsed / processed_entries if processed_entries else 0
                            remaining = entry_total_count - processed_entries
                            est_remaining = avg_time * remaining
                            logger.debug("[PROGRESS] %s/%s entries processed (%.1fs elapsed, 残り推定 %.1fs)", processed_entries, entry_total_count, elapsed, est_remaining)
                            # プログレスバー更新（ProgressWorker対応：3引数形式）
                            if progress_callback:
                                message = f"処理中: {processed_entries}/{entry_total_count} エントリ (残り推定: {est_remaining:.1f}秒)"
                                if not progress_callback(processed_entries, entry_total_count, message):
                                    logger.info("XLSX書き出し処理がキャンセルされました")
                                    return "キャンセルされました"
    logger.info("[PROGRESS] 完了: %s/%s entries (%.1fs)", processed_entries, entry_total_count, time.time()-start_time)
    # 完了時にプログレスバーを100%に
    elapsed = time.time() - start_time
    if progress_callback:
        message = f"完了: {processed_entries}/{entry_total_count} エントリ処理完了 (経過時間: {elapsed:.1f}秒)"
        progress_callback(entry_total_count, entry_total_count, message)
    
    #logger.info("XLSX書き出し処理完了")
    #return "書き出し完了"

# --- 各シート出力関数 ---
    import json, os
    logger.debug("[XLSX] write_summary_sheet called")

    def load_json(path):
        abs_path = os.path.abspath(path)
        if not os.path.exists(abs_path):
            logger.error("%sが存在しません: %s", path, abs_path)
            return None
        with open(abs_path, "r", encoding="utf-8") as f:
            logger.info("[XLSX] JSONロード成功: %s", abs_path)
            return json.load(f)

    # groupOrgnizationsフォルダ内の全ファイルを読み込んで統合
    if not subGroup_included:
        project_groups_dir = Path(GROUP_ORGNIZATION_DIR)
        subGroup_included = []
        if project_groups_dir.exists():
            logger.info(f"[v2.1.17] groupOrgnizationsフォルダからサブグループ情報を読み込み: {project_groups_dir}")
            subgroup_files = list(project_groups_dir.glob("*.json"))
            logger.info(f"[v2.1.17] 検出されたサブグループファイル数: {len(subgroup_files)}件")
            
            for subgroup_file in subgroup_files:
                try:
                    with open(subgroup_file, "r", encoding="utf-8") as f:
                        subgroup_data = json.load(f)
                        # included配列を統合
                        included_items = subgroup_data.get("included", [])
                        subGroup_included.extend(included_items)
                        logger.info(f"[v2.1.17] {subgroup_file.name}: {len(included_items)}件のアイテムを読み込み")
                except Exception as e:
                    logger.error(f"[v2.1.17] サブグループファイル読み込みエラー: {subgroup_file.name} - {e}")
            
            logger.info(f"[v2.1.17] 統合後のsubGroup_included総数: {len(subGroup_included)}件")
        else:
            logger.warning(f"[v2.1.17] groupOrgnizationsフォルダが存在しません: {project_groups_dir}")
    
    if not dataset_data:
        dataset_json = load_json(DATASET_JSON_PATH)
        if not dataset_json:
            return
        dataset_data = dataset_json.get("data", [])

    if not instruments_data:
        instruments_json = load_json(INSTRUMENTS_JSON_PATH)
        if not instruments_json:
            return
        instruments_data = instruments_json.get("data", [])

    # --- 3層構造ヘッダ定義 ---
    HEADER_DEF = [
        {"id": "subGroupName", "label": "サブグループ名"},
        {"id": "dataset_manager_name", "label": "管理者名"},
        {"id": "dataset_applicant_name", "label": "申請者名"},
        {"id": "dataset_owner_names_str", "label": "オーナー名リスト"},
        {"id": "grantNumber", "label": "課題番号"},
        {"id": "title", "label": "課題名"},
        {"id": "datasetName", "label": "データセット名"},
        {"id": "instrument_name", "label": "装置名"},
        {"id": "instrument_local_id", "label": "装置 ID"},
        {"id": "template_id", "label": "テンプレートID"},
        {"id": "datasetId", "label": "データセットID"},
        {"id": "dataEntryName", "label": "データエントリ名"},
        {"id": "dataEntryId", "label": "データエントリID"},
        {"id": "number_of_files", "label": "ファイル数"},
        {"id": "number_of_image_files", "label": "画像ファイル数"},
        {"id": "date_of_dataEntry_creation", "label": "データエントリ作成日"},
        {"id": "total_file_size_MB", "label": "ファイル合計サイズ(MB)"},
        {"id": "dataset_embargoDate", "label": "エンバーゴ日"},
        {"id": "dataset_isAnonymized", "label": "匿名化"},
        {"id": "dataset_description", "label": "データセット説明"},
        {"id": "dataset_relatedLinks", "label": "関連リンク"},
        {"id": "dataset_relatedDatasets", "label": "関連データセット"},
        # --- ファイルタイプ集計ヘッダ ---
        {"id": "filetype_MAIN_IMAGE_count", "label": "MAIN_IMAGEファイル数"},
        {"id": "filetype_MAIN_IMAGE_size", "label": "MAIN_IMAGE合計サイズ"},
        {"id": "filetype_STRUCTURED_count", "label": "STRUCTUREDファイル数"},
        {"id": "filetype_STRUCTURED_size", "label": "STRUCTURED合計サイズ"},
        {"id": "filetype_THUMBNAIL_count", "label": "THUMBNAILファイル数"},
        {"id": "filetype_THUMBNAIL_size", "label": "THUMBNAIL合計サイズ"},
        {"id": "filetype_META_count", "label": "METAファイル数"},
        {"id": "filetype_META_size", "label": "META合計サイズ"},
        {"id": "filetype_OTHER_count", "label": "OTHERファイル数"},
        {"id": "filetype_OTHER_size", "label": "OTHER合計サイズ"},
        {"id": "filetype_total_count", "label": "ファイル総数"},
        {"id": "filetype_total_size", "label": "ファイル総サイズ"},
    ]
    # instrument_local_id列にはinstruments.jsonのattributes.programs[].localIdを出力
    logger.debug("[XLSX] instruments_dataからinstrument_id_to_localidを作成")
    instrument_id_to_localid = {}
    for inst in instruments_data:
        inst_id = inst.get("id")
        programs = inst.get("attributes", {}).get("programs", [])
        # 複数programsがある場合はカンマ区切りで連結
        local_ids = [prog.get("localId", "") for prog in programs if prog.get("localId")]
        if inst_id and local_ids:
            instrument_id_to_localid[inst_id] = ",".join(local_ids)
    SHEET_NAME = "summary"
    logger.debug("[XLSX] シート名: %s", SHEET_NAME)
    
    if SHEET_NAME in wb.sheetnames:
        logger.debug("[XLSX] シートが既に存在: %s", SHEET_NAME)
        ws = wb[SHEET_NAME]
        # 既存ヘッダー行（1行目）を取得
        existing_id_row = [cell.value for cell in ws[1]] if ws.max_row >= 1 else []
    else:
        logger.debug("[XLSX] シートが存在しないため新規作成: %s", SHEET_NAME)
        ws = wb.create_sheet(SHEET_NAME)
        existing_id_row = []
    logger.debug("[XLSX] 既存ヘッダー行: %s", existing_id_row)
    # 既存ID列の順番を優先し、なければHEADER_DEF順で追加（空文字列やNoneは除外）
    header_ids = []
    id_to_label = {coldef["id"]: coldef["label"] for coldef in HEADER_DEF}
    if existing_id_row and any(existing_id_row):
        # 既存ヘッダーの順番（空文字列やNoneは除外）
        header_ids = [id_ for id_ in existing_id_row if id_ not in (None, "") and str(id_).strip() != ""]
        # HEADER_DEFにあるが既存ヘッダーにないものを追加
        for coldef in HEADER_DEF:
            if coldef["id"] not in header_ids:
                header_ids.append(coldef["id"])
    else:
        header_ids = [coldef["id"] for coldef in HEADER_DEF]
    # 1行目:ID（空値列は除外済みだが念のため）
    for col_idx, id_ in enumerate(header_ids, 1):
        if id_ not in (None, "") and str(id_).strip() != "":
            ws.cell(row=1, column=col_idx, value=id_)
    # 2行目:ラベル
    for col_idx, id_ in enumerate(header_ids, 1):
        if id_ not in (None, "") and str(id_).strip() != "":
            ws.cell(row=2, column=col_idx, value=id_to_label.get(id_, id_))
    id_to_col = {id_: idx+1 for idx, id_ in enumerate(header_ids) if id_ not in (None, "") and str(id_).strip() != ""}
    logger.debug("[XLSX] ヘッダーID: %s", header_ids)   
    # 既存データの保存（3行目以降）
    # datasetId, dataEntryId をキーに、手動列（HEADER_DEFにない列）の値を保存
    manual_col_ids = [id_ for id_ in header_ids if id_ not in [coldef["id"] for coldef in HEADER_DEF] and id_ not in (None, "") and str(id_).strip() != ""]
    manual_data_map = {}  # key: ("dataEntryId", id) or ("datasetId", id) -> {manual_col: value, ...}
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=len(header_ids)):
        # header_idsとrowの長さが異なる場合も安全にペア化
        row_dict = {id_: cell.value for id_, cell in zip(header_ids, row) if id_ not in (None, "") and str(id_).strip() != ""}
        dataset_id = row_dict.get("datasetId", "")
        data_entry_id = row_dict.get("dataEntryId", "")
        if data_entry_id:
            manual_data_map[("dataEntryId", data_entry_id)] = {col: row_dict.get(col, None) for col in manual_col_ids if col not in (None, "") and str(col).strip() != ""}
        elif dataset_id:
            manual_data_map[("datasetId", dataset_id)] = {col: row_dict.get(col, None) for col in manual_col_ids if col not in (None, "") and str(col).strip() != ""}
    # 既存データを一旦全削除（3行目以降）
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    logger.debug("[XLSX] 既存データを削除: 3行目以降")
    # ユーザーID→名前辞書
    user_id_to_name = {user.get("id"): user.get("attributes", {}).get("userName", "") for user in subGroup_included if user.get("type") == "user"}
    instrument_id_to_name = {inst.get("id"): inst.get("attributes", {}).get("nameJa", "") for inst in instruments_data}

    from dateutil.parser import parse as parse_datetime
    def to_ymd(date_str):
        if not date_str:
            return ""
        try:
            dt = parse_datetime(date_str)
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return date_str

    def get_dataset_related_info(dataset_datum):
        attr = dataset_datum.get("attributes", {})
        rel = dataset_datum.get("relationships", {})
        return {
            "id": dataset_datum.get("id", ""),
            "manager_id": rel.get("manager", {}).get("data", {}).get("id", ""),
            "owners": rel.get("dataOwners", {}).get("data", []),
            "applicant_id": rel.get("applicant", {}).get("data", {}).get("id", ""),
            "template_id": rel.get("template", {}).get("data", {}).get("id", ""),
            "instrument_id": rel.get("instruments", {}).get("data", [{}])[0].get("id", "") if rel.get("instruments", {}).get("data") else "",
            "embargoDate": to_ymd(attr.get("embargoDate", "")),
            "isAnonymized": attr.get("isAnonymized", ""),
            "description": attr.get("description", ""),
            "relatedLinks_str": "\n".join([link.get("url", "") for link in attr.get("relatedLinks", []) if isinstance(link, dict)]),
            "relatedDatasets_urls_str": "\n".join([f"https://rde.nims.go.jp/datasets/rde/{rd.get('id', '')}" for rd in rel.get("relatedDatasets", {}).get("data", []) if isinstance(rd, dict)]),
            "grantNumber": attr.get("grantNumber", ""),
            "name": attr.get("name", ""),
            "title": attr.get("subjectTitle", ""),
            
        }
    
    row_idx = 3
    logger.info("サブグループとデータセットの関連情報を処理開始")
    
    # プログレス計算用
    total_subgroups = len([sg for sg in subGroup_included if sg.get("type") == "group"])
    total_datasets = len(dataset_data)
    processed_items = 0
    estimated_total = 0
    
    # 概算の処理件数を計算
    for subGroup in subGroup_included:
        if subGroup.get("type") != "group":
            continue
        subGroup_attr = subGroup.get("attributes", {})
        subGroup_subjects = subGroup_attr.get("subjects", {})
        for subject in subGroup_subjects:
            estimated_total += len(dataset_data)
    
    if progress_callback:
        if not progress_callback(0, estimated_total if estimated_total > 0 else 1, f"データ行処理開始 (推定処理数: {estimated_total})"):
            return False
    
    subgroup_idx = 0
    for subGroup in subGroup_included:
        if subGroup.get("type") != "group":
            continue
            
        subgroup_idx += 1
        subGroup_attr = subGroup.get("attributes", {})
        subGroup_name = subGroup_attr.get("name", "")
        subGroup_subjects = subGroup_attr.get("subjects", {})
        
        subject_idx = 0
        for subject in subGroup_subjects:
            subject_idx += 1
            grantNumber = subject.get("grantNumber", "") if isinstance(subject, dict) else ""
            title = subject.get("title", "") if isinstance(subject, dict) else ""
            
            # プログレス更新
            if progress_callback:
                message = f"処理中... サブグループ {subgroup_idx}/{total_subgroups}, 課題 {subject_idx}, 行 {row_idx - 2}"
                if not progress_callback(processed_items, estimated_total if estimated_total > 0 else 1, message):
                    return False
            
            for dataset in dataset_data:
                processed_items += 1
                ds_info = get_dataset_related_info(dataset)
                if ds_info["grantNumber"] != grantNumber:
                    continue
                manager_name = user_id_to_name.get(ds_info["manager_id"], "未設定" if ds_info["manager_id"] in [None, ""] else "")
                applicant_name = user_id_to_name.get(ds_info["applicant_id"], "")
                owner_names = [user_id_to_name.get(owner.get("id", ""), "") for owner in ds_info["owners"] if owner.get("id", "")]
                owner_names_str = "\n".join([n for n in owner_names if n])
                instrument_name = instrument_id_to_name.get(ds_info["instrument_id"], "")
                instrument_local_id = instrument_id_to_localid.get(ds_info["instrument_id"], "")
                dataset_url = f"https://rde.nims.go.jp/datasets/rde/{ds_info['id']}"

                dataEntry_path = os.path.join(OUTPUT_RDE_DIR, "data", "dataEntry", f"{ds_info['id']}.json")
                
                dataEntry_json = load_json(dataEntry_path)
                logger.debug("[XLSX] dataEntry JSONロード for dataset: %s", ds_info['id'])

                if not dataEntry_json:
                    print(f"[ERROR] dataEntry JSONが存在しません: {dataEntry_path} for dataset_id={ds_info['id']}" )
                    continue
                dataEntry_data = dataEntry_json.get("data", [])

                dataEntry_included = dataEntry_json.get("included", [])
                # --- ファイルタイプごとの合計サイズ・ファイル数集計 ---
                # プリセットファイルタイプ
                PRESET_FILETYPES = [
                    "MAIN_IMAGE", "STRUCTURED", "THUMBNAIL", "META"
                ]
                filetype_stats = {ftype: {"count": 0, "size": 0} for ftype in PRESET_FILETYPES}
                filetype_stats["OTHER"] = {"count": 0, "size": 0}
                total_size = 0
                total_count = 0
                for inc in dataEntry_included:
                    if inc.get("type") != "file":
                        continue
                    attr = inc.get("attributes", {})
                    ftype = attr.get("fileType", "OTHER")
                    fsize = attr.get("fileSize", 0)
                    if ftype not in PRESET_FILETYPES:
                        ftype = "OTHER"
                    filetype_stats[ftype]["count"] += 1
                    filetype_stats[ftype]["size"] += fsize
                    total_size += fsize
                    total_count += 1
                filetype_stats["total"] = {"count": total_count, "size": total_size}
                # 既存のtotal_file_size_MBも維持
                total_file_size = total_size
                total_file_size_MB = total_file_size / (1024 * 1024) if total_file_size else 0
                # 必要ならwrite_rowのvalue_dictにfiletype_statsを追加可能

                def write_row(value_dict):
                    # 既存列にデータがなければ既存値を維持
                    # datasetId, dataEntryIdで手動列データを復元
                    dataset_id = value_dict.get("datasetId", "")
                    data_entry_id = value_dict.get("dataEntryId", "")
                    if data_entry_id and ("dataEntryId", data_entry_id) in manual_data_map:
                        manual_restore = manual_data_map[("dataEntryId", data_entry_id)]
                    elif dataset_id and ("datasetId", dataset_id) in manual_data_map:
                        manual_restore = manual_data_map[("datasetId", dataset_id)]
                    else:
                        manual_restore = {}
                    for id_ in header_ids:
                        col = id_to_col[id_]
                        if id_ in value_dict:
                            ws.cell(row=row_idx, column=col, value=value_dict[id_])
                        elif id_ in manual_restore:
                            ws.cell(row=row_idx, column=col, value=manual_restore[id_])
                        else:
                            # 既存値維持（openpyxlは新規行はNoneなので何もしない）
                            pass
                    # value_dictにのみ存在する新規IDは末尾に追加
                    for id_ in value_dict:
                        if id_ not in header_ids:
                            header_ids.append(id_)
                            col = len(header_ids)
                            ws.cell(row=1, column=col, value=id_)
                            ws.cell(row=2, column=col, value=id_to_label.get(id_, id_))
                            ws.cell(row=row_idx, column=col, value=value_dict[id_])
                            id_to_col[id_] = col

                # 複数データエントリ対応
                if dataEntry_data:
                    for entry in dataEntry_data:
                        entry_attr = entry.get("attributes", {})
                        dataEntry_name = entry_attr.get("name", "")
                        dataEntry_id = entry.get("id", "")
                        number_of_files = entry_attr.get("numberOfFiles", "")
                        number_of_image_files = entry_attr.get("numberOfImageFiles", "")
                        date_of_dataEntry_creation = entry_attr.get("created", "")
                        # entry_idに紐づくファイルのみ集計
                        PRESET_FILETYPES = ["MAIN_IMAGE", "STRUCTURED", "THUMBNAIL", "META"]
                        entry_filetype_stats = {ftype: {"count": 0, "size": 0} for ftype in PRESET_FILETYPES}
                        entry_filetype_stats["OTHER"] = {"count": 0, "size": 0}
                        total_size = 0
                        total_count = 0
                        for inc in dataEntry_included:
                            if inc.get("type") != "file":
                                continue
                            attr = inc.get("attributes", {})
                            ftype = attr.get("fileType", "OTHER")
                            fsize = attr.get("fileSize", 0)
                            file_id = inc.get("id", "")
                            # relationships.files.data の id リストに含まれるファイルのみ集計
                            entry_file_ids = [f.get("id", "") for f in entry.get("relationships", {}).get("files", {}).get("data", [])]
                            if file_id not in entry_file_ids:
                                continue
                            if ftype not in PRESET_FILETYPES:
                                ftype = "OTHER"
                            entry_filetype_stats[ftype]["count"] += 1
                            entry_filetype_stats[ftype]["size"] += fsize
                            total_size += fsize
                            total_count += 1
                        entry_filetype_stats["total"] = {"count": total_count, "size": total_size}
                        value_dict = {
                            "subGroupName": subGroup_name,
                            "dataset_manager_name": manager_name,
                            "dataset_applicant_name": applicant_name,
                            "dataset_owner_names_str": owner_names_str,
                            "grantNumber": grantNumber,
                            "title": title,
                            "datasetName": ds_info["name"],
                            "instrument_name": instrument_name,
                            "instrument_local_id": instrument_local_id,
                            "template_id": ds_info["template_id"],
                            "datasetId": dataset_url,
                            "dataEntryName": dataEntry_name,
                            "dataEntryId": dataEntry_id,
                            "number_of_files": number_of_files,
                            "number_of_image_files": number_of_image_files,
                            "date_of_dataEntry_creation": to_ymd(date_of_dataEntry_creation),
                            "total_file_size_MB": total_size / (1024 * 1024) if total_size else 0,
                            "dataset_embargoDate": ds_info["embargoDate"],
                            "dataset_isAnonymized": ds_info["isAnonymized"],
                            "dataset_description": ds_info["description"],
                            "dataset_relatedLinks": ds_info["relatedLinks_str"],
                            "dataset_relatedDatasets": ds_info["relatedDatasets_urls_str"],
                            # ファイルタイプごとの集計値を追加
                            "filetype_MAIN_IMAGE_count": entry_filetype_stats["MAIN_IMAGE"]["count"],
                            "filetype_MAIN_IMAGE_size": entry_filetype_stats["MAIN_IMAGE"]["size"],
                            "filetype_STRUCTURED_count": entry_filetype_stats["STRUCTURED"]["count"],
                            "filetype_STRUCTURED_size": entry_filetype_stats["STRUCTURED"]["size"],
                            "filetype_THUMBNAIL_count": entry_filetype_stats["THUMBNAIL"]["count"],
                            "filetype_THUMBNAIL_size": entry_filetype_stats["THUMBNAIL"]["size"],
                            "filetype_META_count": entry_filetype_stats["META"]["count"],
                            "filetype_META_size": entry_filetype_stats["META"]["size"],
                            "filetype_OTHER_count": entry_filetype_stats["OTHER"]["count"],
                            "filetype_OTHER_size": entry_filetype_stats["OTHER"]["size"],
                            "filetype_total_count": entry_filetype_stats["total"]["count"],
                            "filetype_total_size": entry_filetype_stats["total"]["size"],
                        }
                        write_row(value_dict)
                        row_idx += 1
                else:
                    # データエントリがない場合も空で1行出す
                    value_dict = {
                        "subGroupName": subGroup_name,
                        "dataset_manager_name": manager_name,
                        "dataset_applicant_name": applicant_name,
                        "dataset_owner_names_str": owner_names_str,
                        "grantNumber": grantNumber,
                        "title": title,
                        "datasetName": ds_info["name"],
                        "instrument_name": instrument_name,
                        "instrument_local_id": instrument_local_id,
                        "template_id": ds_info["template_id"],
                        "datasetId": dataset_url,
                        "dataEntryName": "",
                        "dataEntryId": "",
                        "number_of_files": "",
                        "number_of_image_files": "",
                        "date_of_dataEntry_creation": "",
                        "total_file_size_MB": total_file_size_MB,
                        "dataset_embargoDate": ds_info["embargoDate"],
                        "dataset_isAnonymized": ds_info["isAnonymized"],
                        "dataset_description": ds_info["description"],
                        "dataset_relatedLinks": ds_info["relatedLinks_str"],
                        "dataset_relatedDatasets": ds_info["relatedDatasets_urls_str"],
                        # ファイルタイプごとの集計値を追加
                        "filetype_MAIN_IMAGE_count": filetype_stats["MAIN_IMAGE"]["count"],
                        "filetype_MAIN_IMAGE_size": filetype_stats["MAIN_IMAGE"]["size"],
                        "filetype_STRUCTURED_count": filetype_stats["STRUCTURED"]["count"],
                        "filetype_STRUCTURED_size": filetype_stats["STRUCTURED"]["size"],
                        "filetype_THUMBNAIL_count": filetype_stats["THUMBNAIL"]["count"],
                        "filetype_THUMBNAIL_size": filetype_stats["THUMBNAIL"]["size"],
                        "filetype_META_count": filetype_stats["META"]["count"],
                        "filetype_META_size": filetype_stats["META"]["size"],
                        "filetype_OTHER_count": filetype_stats["OTHER"]["count"],
                        "filetype_OTHER_size": filetype_stats["OTHER"]["size"],
                        "filetype_total_count": filetype_stats["total"]["count"],
                        "filetype_total_size": filetype_stats["total"]["size"],
                    }
                    write_row(value_dict)
                    row_idx += 1



# --- 各シート出力関数 ---
def write_members_sheet(wb, parent):
    import json, os
    from pathlib import Path
    
    # groupOrgnizationsフォルダ内の全ファイルから全ユーザーを統合
    project_groups_dir = Path(GROUP_ORGNIZATION_DIR)
    users = []
    
    if project_groups_dir.exists():
        subgroup_files = list(project_groups_dir.glob("*.json"))
        logger.info(f"[v2.1.17] member シート: {len(subgroup_files)}件のサブグループファイルから読み込み")
        
        for subgroup_file in subgroup_files:
            try:
                with open(subgroup_file, "r", encoding="utf-8") as f:
                    sub_group = json.load(f)
                    file_users = [item for item in sub_group.get("included", []) if item.get("type") == "user"]
                    users.extend(file_users)
            except Exception as e:
                logger.error(f"[v2.1.17] member シート: ファイル読み込みエラー - {subgroup_file.name}: {e}")
    else:
        logger.warning(f"[v2.1.17] member シート: groupOrgnizationsフォルダが存在しません")
        return
    
    HEADER_ROW = ["userId", "userName", "familyName", "givenName", "familyNameKanji", "givenNameKanji", "organizationName", "emailAddress", "isDeleted"]
    if "member" in wb.sheetnames:
        ws = wb["member"]
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER_ROW)
    else:
        ws = wb.create_sheet("member")
        ws.append(HEADER_ROW)
    for user in users:
        attr = user.get("attributes", {})
        ws.append([
            user.get("id", ""),
            attr.get("userName", ""),
            attr.get("familyName", ""),
            attr.get("givenName", ""),
            attr.get("familyNameKanji", ""),
            attr.get("givenNameKanji", ""),
            attr.get("organizationName", ""),
            attr.get("emailAddress", ""),
            attr.get("isDeleted", False)
        ])

# 今後拡張用の関数枠
def write_datasets_sheet(wb, parent):
    import json, os
    JSON_PATH = os.path.join(OUTPUT_RDE_DIR, "data", "dataset.json")
    abs_json = os.path.abspath(JSON_PATH)
    logger.debug("write_datasets_sheet: JSON_PATH=%s, abs_json=%s", JSON_PATH, abs_json)
    if not os.path.exists(abs_json):
        return
    with open(abs_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    datasets = data.get("data", [])
    HEADER_ROW = [
        "id", "datasetType", "grantNumber", "subjectTitle", "name", "description", "contact",
        "dataListingType", "usesInstrument", "isOpen", "openAt", "embargoDate", "isAnonymized",
        "created", "modified", "managerId", "ownerId", "templateId", "instrumentId"]
    SHEET_NAME = "datasets"
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER_ROW)
    else:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADER_ROW)
    import re
    def safe_str(val):
        try:
            if val is None:
                return "!!NULL!!"
            s = str(val)
            # 改行・制御文字（ASCII 0-31, 127）を除去
            s = re.sub(r"[\x00-\x1F\x7F]", " ", s)
            # 連続する空白を1つに
            s = re.sub(r" +", " ", s)
            # 先頭・末尾の空白除去
            return s.strip()
        except Exception:
            return "!!ERROR!!"

    for ds in datasets:
        try:
            attr = ds.get("attributes", {})
            rel = ds.get("relationships", {})
           

            templateId = ""
            template_data = rel.get("template", {}).get("data")
            if isinstance(template_data, dict):
                templateId = template_data.get("id", "")

            instrumentId = ""
            instruments = rel.get("instruments", {}).get("data", [])
            if isinstance(instruments, list) and instruments:
                instrument = instruments[0]
                if isinstance(instrument, dict):
                    instrumentId = instrument.get("id", "")

            # managerId（managerのid)
            managerId = "!!"
            m = rel.get("manager", {}).get("data")
            if isinstance(m, dict):
                managerId = m.get("id", "")

            # ownerId（dataOwners配列の最初のid）
            ownerId = "!!"
            owners = rel.get("dataOwners", {}).get("data", [])
            if isinstance(owners, list) and owners:
                owner = owners[0]
                if isinstance(owner, dict):
                    ownerId = owner.get("id", "")
            row = [
                safe_str(ds.get("id", "")),
                safe_str(attr.get("datasetType", "")),
                safe_str(attr.get("grantNumber", "")),
                safe_str(attr.get("subjectTitle", "")),
                safe_str(attr.get("name", "")),
                safe_str(attr.get("description", "")),
                safe_str(attr.get("contact", "")),
                safe_str(attr.get("dataListingType", "")),
                safe_str(attr.get("usesInstrument", "")),
                safe_str(attr.get("isOpen", "")),
                safe_str(attr.get("openAt", "")),
                safe_str(attr.get("embargoDate", "")),
                safe_str(attr.get("isAnonymized", "")),
                safe_str(attr.get("created", "")),
                safe_str(attr.get("modified", "")),
                
                safe_str(managerId),
                safe_str(ownerId),
                safe_str(templateId),
                safe_str(instrumentId)
            ]
            ws.append(row)
        except Exception as e:
            logger.error("datasetsシート出力時エラー: id=%s : %s", ds.get('id',''), e)
            ws.append(["" for _ in range(len(HEADER_ROW))])
def write_templates_sheet(wb, parent):
    import json, os
    JSON_PATH = os.path.join(OUTPUT_RDE_DIR, "data", "template.json")
    abs_json = os.path.abspath(JSON_PATH)
    if not os.path.exists(abs_json):
        return
    with open(abs_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    templates = data.get("data", [])
    HEADER_ROW = [
        "id", "nameJa", "nameEn", "version", "datasetType", "description", "isPrivate", "workflowEnabled", "usesInstrument", "created"
    ]
    SHEET_NAME = "templates"
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER_ROW)
    else:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADER_ROW)
    for tpl in templates:
        attr = tpl.get("attributes", {})
        ws.append([
            tpl.get("id", ""),
            attr.get("nameJa", ""),
            attr.get("nameEn", ""),
            attr.get("version", ""),
            attr.get("datasetType", ""),
            attr.get("description", ""),
            attr.get("isPrivate", ""),
            attr.get("workflowEnabled", ""),
            attr.get("usesInstrument", ""),
            attr.get("created", "")
        ])
    import json, os
    JSON_PATH = os.path.join(OUTPUT_RDE_DIR, "data", "instruments.json")
    abs_json = os.path.abspath(JSON_PATH)
    if not os.path.exists(abs_json):
        return
    with open(abs_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    instruments = data.get("data", [])
    HEADER_ROW = [
        "id", "organizationId", "organizationNameJa", "organizationNameEn", "nameJa", "nameEn", "modelNumber", "manufacturerJa", "manufacturerEn"
    ]
    SHEET_NAME = "instruments"
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER_ROW)
    else:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADER_ROW)
    for inst in instruments:
        attr = inst.get("attributes", {})
        ws.append([
            inst.get("id", ""),
            attr.get("organizationId", ""),
            attr.get("organizationNameJa", ""),
            attr.get("organizationNameEn", ""),
            attr.get("nameJa", ""),
            attr.get("nameEn", ""),
            attr.get("modelNumber", ""),
            attr.get("manufacturerJa", ""),
            attr.get("manufacturerEn", "")
        ])
def write_group_sheet(wb, parent):
    pass
def write_instrumentType_sheet(wb, parent):
    import json, os
    JSON_PATH = os.path.join(OUTPUT_RDE_DIR, "data", "instrumentType.json")
    abs_json = os.path.abspath(JSON_PATH)
    if not os.path.exists(abs_json):
        return
    with open(abs_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    type_terms = data.get("typeTerms", [])
    HEADER_ROW = ["termId", "termNameJa", "termNameEn", "narrowerTerms"]
    SHEET_NAME = "instrumentType"
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER_ROW)
    else:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADER_ROW)
    for term in type_terms:
        narrower = term.get("narrowerTerms", [])
        # Noneを除外しstr化
        narrower_str = ", ".join([str(x) for x in narrower if x is not None])
        ws.append([
            term.get("termId", ""),
            term.get("termNameJa", ""),
            term.get("termNameEn", ""),
            narrower_str
        ])
def write_groupDetail_sheet(wb, parent):
    pass
def write_organization_sheet(wb, parent):
    import json, os
    JSON_PATH = os.path.join(OUTPUT_RDE_DIR, "data", "organization.json")
    abs_json = os.path.abspath(JSON_PATH)
    if not os.path.exists(abs_json):
        return
    with open(abs_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    orgs = data.get("data", [])
    HEADER_ROW = ["id", "nameJa", "nameEn"]
    SHEET_NAME = "organization"
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER_ROW)
    else:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADER_ROW)
    for org in orgs:
        attr = org.get("attributes", {})
        ws.append([
            org.get("id", ""),
            attr.get("nameJa", ""),
            attr.get("nameEn", "")
        ])

def write_instruments_sheet(wb, parent):
    import json, os
    JSON_PATH = os.path.join(OUTPUT_RDE_DIR, "data", "instruments.json")
    abs_json = os.path.abspath(JSON_PATH)
    if not os.path.exists(abs_json):
        return
    with open(abs_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    instruments = data.get("data", [])
    HEADER_ROW = [
        "id", "organizationId", "organizationNameJa", "organizationNameEn", "nameJa", "nameEn", "modelNumber", "manufacturerJa", "manufacturerEn"
    ]
    SHEET_NAME = "instruments"
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER_ROW)
    else:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADER_ROW)
    for inst in instruments:
        attr = inst.get("attributes", {})
        ws.append([
            inst.get("id", ""),
            attr.get("organizationId", ""),
            attr.get("organizationNameJa", ""),
            attr.get("organizationNameEn", ""),
            attr.get("nameJa", ""),
            attr.get("nameEn", ""),
            attr.get("modelNumber", ""),
            attr.get("manufacturerJa", ""),
            attr.get("manufacturerEn", "")
        ])

def write_licenses_sheet(wb, parent):
    """利用ライセンス情報をlicensesシートに出力"""
    import json, os
    JSON_PATH = os.path.join(OUTPUT_RDE_DIR, "data", "licenses.json")
    abs_json = os.path.abspath(JSON_PATH)
    if not os.path.exists(abs_json):
        return
    with open(abs_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    licenses = data.get("data", [])
    HEADER_ROW = ["id", "fullName", "url"]
    SHEET_NAME = "licenses"
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER_ROW)
    else:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADER_ROW)
    for license_item in licenses:
        attr = license_item.get("attributes", {})
        ws.append([
            license_item.get("id", ""),
            attr.get("fullName", ""),
            attr.get("url", "")
        ])


def write_subgroups_sheet(wb, parent):
    import json, os
    from pathlib import Path
    
    # groupOrgnizationsフォルダ内の全ファイルから全サブグループを統合
    project_groups_dir = Path(GROUP_ORGNIZATION_DIR)
    subgroups = []
    
    if project_groups_dir.exists():
        subgroup_files = list(project_groups_dir.glob("*.json"))
        logger.info(f"[v2.1.17] subgroups シート: {len(subgroup_files)}件のサブグループファイルから読み込み")
        
        for subgroup_file in subgroup_files:
            try:
                with open(subgroup_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    included = data.get("included", [])
                    file_subgroups = [item for item in included if item.get("type") == "group"]
                    subgroups.extend(file_subgroups)
            except Exception as e:
                logger.error(f"[v2.1.17] subgroups シート: ファイル読み込みエラー - {subgroup_file.name}: {e}")
    else:
        logger.warning(f"[v2.1.17] subgroups シート: groupOrgnizationsフォルダが存在しません")
        return
    
    HEADER_ROW = ["groupId", "name", "groupType", "description", "ownerId"]
    SHEET_NAME = "subgroups"
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER_ROW)
    else:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADER_ROW)
    for group in subgroups:
        attr = group.get("attributes", {})
        owner_id = ""
        for role in attr.get("roles", []):
            if role.get("role") == "OWNER":
                owner_id = role.get("userId", "")
                break
        ws.append([
            group.get("id", ""),
            attr.get("name", ""),
            attr.get("groupType", ""),
            attr.get("description", ""),
            owner_id
        ])

def write_groupDetail_sheet(wb, parent):
    import json, os
    JSON_PATH = os.path.join(OUTPUT_RDE_DIR, "data", "groupDetail.json")
    abs_json = os.path.abspath(JSON_PATH)
    if not os.path.exists(abs_json):
        return
    with open(abs_json, "r", encoding="utf-8") as f:
        data = json.load(f)
    # data本体
    d = data.get("data", {})
    attr = d.get("attributes", {})
    HEADER_ROW = ["id", "groupType", "name", "description"]
    SHEET_NAME = "groupDetail"
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER_ROW)
    else:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADER_ROW)
    ws.append([
        d.get("id", ""),
        attr.get("groupType", ""),
        attr.get("name", ""),
        attr.get("description", "")
    ])
    # included配列のgroupも出力
    included = data.get("included", [])
    for group in included:
        if group.get("type") == "group":
            attr = group.get("attributes", {})
            ws.append([
                group.get("id", ""),
                attr.get("groupType", ""),
                attr.get("name", ""),
                attr.get("description", "")
            ])        

def write_entries_sheet(wb, parent):
    import os, json, glob
    DATA_ENTRY_DIR = os.path.join(OUTPUT_RDE_DIR, "data", "dataEntry")
    files = glob.glob(os.path.join(DATA_ENTRY_DIR, "*.json"))
    HEADER_ROW = [
        "entryId", "datasetId", "dataNumber", "name", "description", "experimentId","numberOfFiles", "numberOfImageFiles",
        "instrument.name", "instrument.organization", "invoice.basic.data_owner", "invoice.basic.date_submitted", "sample.name"
    ]
    SHEET_NAME = "entries"
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER_ROW)
    else:
        ws = wb.create_sheet(SHEET_NAME)
        ws.append(HEADER_ROW)
    for file in files:
        dataset_id = os.path.splitext(os.path.basename(file))[0]
        with open(file, "r", encoding="utf-8") as f:
            data = json.load(f)
        for entry in data.get("data", []):
            attr = entry.get("attributes", {})
            meta = attr.get("metadata", {})
            def get_meta(key):
                v = meta.get(key, {}).get("value")
                if isinstance(v, list):
                    return v[0] if v else ""
                return v if v is not None else ""
            ws.append([
                entry.get("id", ""),
                dataset_id,
                attr.get("dataNumber", ""),
                attr.get("name", ""),
                attr.get("description", ""),
                attr.get("experimentId", ""),
                attr.get("numberOfFiles", ""),
                attr.get("numberOfImageFiles", ""),
                get_meta("instrument.name"),
                get_meta("instrument.organization"),
                get_meta("invoice.basic.data_owner"),
                get_meta("invoice.basic.date_submitted"),
                get_meta("sample.name")
            ])


# ========================================
# UIController用ラッパー関数
# ========================================

def apply_basic_info_to_xlsx(ui_controller):
    """XLSX反映 - UIController用ラッパー（プログレス表示対応）"""
    try:
        from ..ui.ui_basic_info import apply_basic_info_to_Xlsx as ui_apply_basic_info_to_xlsx
        ui_apply_basic_info_to_xlsx(ui_controller)
    except Exception as e:
        ui_controller.show_error(f"XLSX反映でエラーが発生しました: {e}")
        logger.error(f"apply_basic_info_to_xlsx エラー: {e}")


def summary_basic_info_to_xlsx(ui_controller):
    """まとめXLSX作成 - UIController用ラッパー（プログレス表示対応）"""
    try:
        from ..ui.ui_basic_info import summary_basic_info_to_Xlsx as ui_summary_basic_info_to_xlsx
        ui_summary_basic_info_to_xlsx(ui_controller)
    except Exception as e:
        ui_controller.show_error(f"まとめXLSX作成でエラーが発生しました: {e}")
        logger.error(f"summary_basic_info_to_xlsx エラー: {e}")
