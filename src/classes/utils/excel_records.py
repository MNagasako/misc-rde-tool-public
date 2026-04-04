from __future__ import annotations

import math
from typing import Any, Dict, List, Sequence, Tuple

from openpyxl import load_workbook


class EmptyExcelError(ValueError):
    """Raised when an Excel file does not contain a usable header row."""


def is_missing_value(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, float):
        try:
            return math.isnan(value)
        except Exception:
            return False
    return False


def has_meaningful_value(value: object) -> bool:
    if is_missing_value(value):
        return False
    if isinstance(value, str):
        normalized = value.strip()
        return bool(normalized) and normalized.lower() not in {"nan", "none"}
    try:
        normalized = str(value).strip()
    except Exception:
        return False
    return bool(normalized) and normalized.lower() not in {"nan", "none"}


def get_record_headers(records: Sequence[Dict[str, Any]]) -> List[str]:
    headers: List[str] = []
    seen = set()
    for record in records:
        for key in record.keys():
            key_text = str(key)
            if key_text in seen:
                continue
            seen.add(key_text)
            headers.append(key_text)
    return headers


def ensure_alias_column(records: Sequence[Dict[str, Any]], source_key: str, target_key: str) -> None:
    for record in records:
        if target_key in record and has_meaningful_value(record.get(target_key)):
            continue
        record[target_key] = record.get(source_key)


def load_excel_records(file_path: str, *, sheet_name: str | None = None) -> Tuple[List[str], List[Dict[str, Any]]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active
    rows_iter = worksheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise EmptyExcelError(f"Excel file has no header row: {file_path}")

    headers = [str(value).strip() if value is not None else "" for value in header_row]
    if not any(headers):
        raise EmptyExcelError(f"Excel file has no usable header row: {file_path}")

    records: List[Dict[str, Any]] = []
    header_count = len(headers)
    for row in rows_iter:
        values = list(row[:header_count])
        if len(values) < header_count:
            values.extend([None] * (header_count - len(values)))
        if not any(has_meaningful_value(value) or (value is not None and not is_missing_value(value)) for value in values):
            continue

        record: Dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = None if is_missing_value(values[index]) else values[index]
        records.append(record)

    return headers, records