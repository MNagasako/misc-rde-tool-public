"""
報告書データ変換モジュール

ARIM-extracted2フォーマットから標準フォーマットへのExcel変換を実施。
カラムマッピング、設備列の変換、チャンク処理、レジューム対応を含む。
"""

import ast
import json
import os
import re
from typing import Optional, Dict, Any, List, Callable, cast

from dataclasses import dataclass
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet._read_only import ReadOnlyWorksheet
from openpyxl.worksheet.worksheet import Worksheet

from classes.utils.excel_records import load_excel_records


@dataclass
class ConversionResult:
    """変換結果"""
    success: bool
    output_path: Optional[str] = None
    row_count: int = 0
    error: Optional[str] = None


class ReportConverter:
    """報告書データ変換クラス
    
    ARIM-extracted2フォーマットのExcelファイルを標準フォーマットに変換。
    """
    
    # カラムマッピング定義
    COLUMN_MAPPING = {
        'ARIM-extracted2 Columns': [
            'ARIMNO', 'コード', '年度', '機関コード', '課題番号（下4桁）', 
            '機関外・機関内の利用', '利用形態（主）', '利用形態（副）', 
            '利用課題名', '利用者名', '所属名', 
            '横断技術領域（主）', '横断技術領域（副）', 
            '重要技術領域（主）', '重要技術領域（副）', 
            '利用した主な設備1', '利用した主な設備2', '利用した主な設備3', 
            '利用した主な設備4', '利用した主な設備5', 
            'キーワード【横断技術領域】（主）', 'キーワード【横断技術領域】（副）', 
            'キーワード【重要技術領域】（主）', 'キーワード【重要技術領域】（副）', 
            'キーワード【自由記述】', '概要（目的・実施内容）', '実験', '結果と考察'
        ],
        'Output Columns': [
            '課題番号 / Project Issue Number', 'code', 'np.nan', 'np.nan', 'np.nan',
            '機関外・機関内の利用 / External or Internal Use', 
            '利用形態・主', '利用形態・副', 
            '利用課題名 / Title', 
            '利用者名（課題申請者）/ User Name (Project Applicant)', 
            '所属名 / Affiliation', 
            '横断技術領域・主', '横断技術領域・副', 
            '重要技術領域・主', '重要技術領域・副', 
            'np.nan', 'np.nan', 'np.nan', 'np.nan', 'np.nan', 
            'np.nan', 'np.nan', 'np.nan', 'np.nan', 
            'キーワード / Keywords', 
            '概要（目的・用途・実施内容）/ Abstract (Aim, Use Applications and Contents)', 
            '実験 / Experimental', 
            '結果と考察 / Results and Discussion', 
            '利用した実施機関 / Support Institute', 
            '共同利用者氏名 / Names of Collaborators in Other Institutes Than Hub and Spoke Institutes', 
            'ARIM実施機関支援担当者 / Names of Collaborators in The Hub and Spoke Institutes', 
            '利用した主な設備 / Equipment Used in This Project', 
            'その他・特記事項（参考文献・謝辞等） / Remarks(References and Acknowledgements)', 
            '論文・プロシーディング（DOIのあるもの） / DOI (Publication and Proceedings)', 
            '口頭発表、ポスター発表および、その他の論文 / Oral Presentations etc.', 
            '特許出願件数', '特許登録件数', 'key'
        ]
    }
    
    # ARIM-extracted2カラム名
    ARIM_COLUMNS = [
        'ARIMNO', 'コード', '年度', '機関コード', '課題番号（下4桁）', 
        '機関外・機関内の利用', '利用形態（主）', '利用形態（副）', 
        '利用課題名', '利用者名', '所属名', 
        '横断技術領域（主）', '横断技術領域（副）', 
        '重要技術領域（主）', '重要技術領域（副）', 
        '利用した主な設備1', '利用した主な設備2', '利用した主な設備3', 
        '利用した主な設備4', '利用した主な設備5', 
        'キーワード【横断技術領域】（主）', 'キーワード【横断技術領域】（副）', 
        'キーワード【重要技術領域】（主）', 'キーワード【重要技術領域】（副）', 
        'キーワード【自由記述】', '概要（目的・実施内容）', '実験', '結果と考察'
    ]
    
    # 分割処理が必要なカラム(日本語/英語)
    COLUMNS_TO_SPLIT = [
        '機関外・機関内の利用', '利用形態（主）', '利用形態（副）', 
        '横断技術領域（主）', '横断技術領域（副）', 
        '重要技術領域（主）', '重要技術領域（副）'
    ]
    
    # ダッシュ変換が必要なカラム
    COLUMNS_TO_DASH_CONVERT = [
        '利用形態（主）', '利用形態（副）', 
        '横断技術領域（主）', '横断技術領域（副）', 
        '重要技術領域（主）', '重要技術領域（副）'
    ]
    
    CHUNK_SIZE = 100  # 100件ごとに保存
    EQUIPMENT_COLUMN_NAME = '利用した主な設備 / Equipment Used in This Project'
    OUTPUT_COLUMNS = ARIM_COLUMNS + ['key']
    
    def __init__(self, progress_callback: Optional[Callable[[str], None]] = None):
        """
        Args:
            progress_callback: プログレスコールバック関数
        """
        self.progress_callback = progress_callback
    
    def _log(self, message: str):
        """ログ出力"""
        if self.progress_callback:
            self.progress_callback(message)
    
    @staticmethod
    def safely_evaluate_literal(literal_str: Any) -> List:
        """文字列をPythonリテラル(リスト)として安全に評価
        
        Args:
            literal_str: 評価する文字列
            
        Returns:
            評価結果のリスト（失敗時は空リスト）
        """
        if not isinstance(literal_str, str):
            return []

        candidate = literal_str.strip()
        if not candidate:
            return []

        try:
            return ast.literal_eval(candidate)
        except (ValueError, SyntaxError):
            # フォールバック: カンマまたは改行で区切られた文字列をリスト化
            if "," in candidate or "\n" in candidate:
                parts = [part.strip() for part in re.split(r"[,\n]", candidate)]
                return [part for part in parts if part]
            return [candidate]
    
    def _transform_chunk(self, chunk_records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """データチャンクの変換処理"""
        transformed_records: List[Dict[str, Any]] = []

        for record in chunk_records:
            transformed: Dict[str, Any] = {column: None for column in self.OUTPUT_COLUMNS}

            for arim_col, output_col in zip(
                self.COLUMN_MAPPING['ARIM-extracted2 Columns'],
                self.COLUMN_MAPPING['Output Columns'],
            ):
                if output_col != 'np.nan':
                    value = record.get(output_col)
                    if arim_col in self.COLUMNS_TO_SPLIT and isinstance(value, str):
                        value = value.split('/')[0].strip()
                    if arim_col in self.COLUMNS_TO_DASH_CONVERT and value == '-':
                        value = '----'
                    transformed[arim_col] = value
                    continue

                if '利用した主な設備' not in arim_col:
                    continue
                equipment_num = int(arim_col[-1])
                equipment_list = record.get(self.EQUIPMENT_COLUMN_NAME, [])
                transformed[arim_col] = (
                    equipment_list[equipment_num - 1]
                    if isinstance(equipment_list, list) and len(equipment_list) >= equipment_num
                    else None
                )

            arimno = f"JPMXP12{transformed.get('ARIMNO', '')}"
            transformed['ARIMNO'] = arimno
            transformed['年度'] = f"20{arimno[7:9]}" if len(arimno) >= 9 else None
            transformed['機関コード'] = arimno[9:11] if len(arimno) >= 11 else None
            suffix = arimno[-4:] if len(arimno) >= 4 else ''
            try:
                transformed['課題番号（下4桁）'] = int(suffix)
            except ValueError:
                transformed['課題番号（下4桁）'] = suffix
            transformed['key'] = record.get(self.COLUMN_MAPPING['Output Columns'][-1])
            transformed_records.append(transformed)

        return transformed_records

    def _save_workbook_atomically(self, workbook: Workbook, output_path: str) -> None:
        """ワークブックを一時ファイル経由で原子的に保存する。"""
        staging_path = output_path + ".writing"
        try:
            workbook.save(staging_path)
            os.replace(staging_path, output_path)
        finally:
            if os.path.exists(staging_path):
                try:
                    os.remove(staging_path)
                except OSError:
                    pass

    def _write_chunk(self, output_path: str, records: List[Dict[str, Any]], *, append: bool) -> None:
        if append and os.path.exists(output_path):
            workbook = load_workbook(output_path)
            worksheet = cast(Worksheet, workbook.active)
        else:
            workbook = Workbook()
            worksheet = cast(Worksheet, workbook.active)
            worksheet.append(self.OUTPUT_COLUMNS)

        try:
            for record in records:
                worksheet.append([record.get(column) for column in self.OUTPUT_COLUMNS])

            self._save_workbook_atomically(workbook, output_path)
        finally:
            close = getattr(workbook, "close", None)
            if callable(close):
                close()

    def _inspect_resume_workbook(self, tmp_output_file: str) -> Optional[int]:
        """レジューム用ワークブックを検証し、保持済みデータ行数を返す。"""
        try:
            workbook = load_workbook(tmp_output_file, read_only=True, data_only=True)
        except Exception:
            return None

        try:
            worksheet = cast(ReadOnlyWorksheet, workbook.active)
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if header_row is None:
                return None
            normalized_headers = ["" if value is None else str(value) for value in header_row]
            if normalized_headers[:len(self.OUTPUT_COLUMNS)] != self.OUTPUT_COLUMNS:
                return None
            max_row = worksheet.max_row or 1
            return max(max_row - 1, 0)
        except Exception:
            return None
        finally:
            close = getattr(workbook, "close", None)
            if callable(close):
                close()

    def _clear_resume_state(self, progress_file: str, tmp_output_file: str, reason: str) -> None:
        """破損・不整合なレジューム状態を削除する。"""
        for stale_path in (progress_file, tmp_output_file):
            if not os.path.exists(stale_path):
                continue
            try:
                os.remove(stale_path)
            except OSError:
                pass
        self._log(f"レジューム情報を破棄して先頭から再開: {reason}")

    def _resolve_resume_start_index(
        self,
        *,
        resume: bool,
        progress_file: str,
        tmp_output_file: str,
    ) -> int:
        """レジューム可能なら開始インデックスを返し、不整合なら安全にリセットする。"""
        if not resume or not os.path.exists(progress_file):
            return 0

        try:
            with open(progress_file, 'r', encoding='utf-8') as f:
                progress = json.load(f)
            progress_index = int(progress.get('last_index', 0) or 0)
        except Exception as exc:
            self._clear_resume_state(progress_file, tmp_output_file, f"進捗ファイルを読めません ({exc})")
            return 0

        resume_rows = self._inspect_resume_workbook(tmp_output_file)
        if resume_rows is None:
            self._clear_resume_state(progress_file, tmp_output_file, "一時Excelが破損または不正です")
            return 0

        if resume_rows != progress_index:
            self._log(
                f"レジューム情報を補正: progress={progress_index}行, tmp={resume_rows}行"
            )

        start_idx = resume_rows
        self._log(f"レジューム: {start_idx}行目から再開")
        return start_idx
    
    def convert_report_data(
        self, 
        input_path: str, 
        output_path: str,
        resume: bool = True
    ) -> ConversionResult:
        """報告書データの変換
        
        Args:
            input_path: 入力Excelファイルパス
            output_path: 出力Excelファイルパス
            resume: レジューム機能を有効化（デフォルト: True）
            
        Returns:
            ConversionResult
        """
        try:
            # 入力ファイル存在確認
            if not os.path.exists(input_path):
                return ConversionResult(
                    success=False,
                    error=f"入力ファイルが見つかりません: {input_path}"
                )
            
            self._log(f"入力ファイル読み込み: {input_path}")
            
            _headers, output_records = load_excel_records(input_path)

            # 設備列の変換（文字列→リスト）
            for record in output_records:
                value = record.get(self.EQUIPMENT_COLUMN_NAME)
                record[self.EQUIPMENT_COLUMN_NAME] = (
                    self.safely_evaluate_literal(value) if isinstance(value, str) else []
                )
            
            # 進捗管理用ファイル
            output_dir = os.path.dirname(output_path)
            progress_file = os.path.join(output_dir, 'convert_progress.json')
            tmp_output_file = output_path + '.tmp.xlsx'
            
            # レジューム対応
            start_idx = self._resolve_resume_start_index(
                resume=resume,
                progress_file=progress_file,
                tmp_output_file=tmp_output_file,
            )
            
            # チャンク処理
            num_rows = len(output_records)
            if start_idx > num_rows:
                self._log(
                    f"レジューム位置を補正: {start_idx}行 -> {num_rows}行 (入力総行数を超過)"
                )
                start_idx = num_rows
            self._log(f"変換開始: {num_rows}行 ({start_idx}行目から)")
            
            for chunk_start in range(start_idx, num_rows, self.CHUNK_SIZE):
                chunk_end = min(chunk_start + self.CHUNK_SIZE, num_rows)
                chunk_records = output_records[chunk_start:chunk_end]
                
                # チャンク変換
                chunk_transformed = self._transform_chunk(chunk_records)
                
                # 保存
                self._write_chunk(tmp_output_file, chunk_transformed, append=(chunk_start != 0))
                
                # 進捗保存
                with open(progress_file, 'w', encoding='utf-8') as f:
                    json.dump({'last_index': chunk_end}, f)
                
                progress_pct = (chunk_end / num_rows) * 100
                self._log(f"変換中: {chunk_end}/{num_rows}行 ({progress_pct:.1f}%)")
            
            # 完了時にリネーム&進捗ファイル削除
            os.replace(tmp_output_file, output_path)
            if os.path.exists(progress_file):
                os.remove(progress_file)
            
            self._log(f"変換完了: {output_path}")
            
            return ConversionResult(
                success=True,
                output_path=output_path,
                row_count=num_rows
            )
            
        except Exception as e:
            return ConversionResult(
                success=False,
                error=f"変換エラー: {str(e)}"
            )
