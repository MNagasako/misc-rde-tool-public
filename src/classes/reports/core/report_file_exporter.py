"""
報告書機能 - ファイル出力モジュール

Excel/JSON形式での報告書データ出力機能を提供します。

"""

import os
import json
import logging
import shutil
from typing import List, Dict, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook

from config.common import OUTPUT_DIR
from ..conf.field_definitions import EXCEL_COLUMNS
from .report_cache_manager import ReportCacheManager


logger = logging.getLogger(__name__)


class ReportFileExporter:
    """
    報告書ファイル出力クラス
    
    Excel/JSON形式でデータを出力し、バックアップ管理を行います。
    """
    
    def __init__(self, output_dir: Optional[str] = None):
        """
        初期化
        
        Args:
            output_dir: 出力ディレクトリパス（Noneの場合はデフォルト）
        """
        if output_dir is None:
            # デフォルトはOUTPUT_DIR/arim-site/reports
            output_dir = os.path.join(OUTPUT_DIR, "arim-site", "reports")
        
        # ディレクトリが存在しない場合は作成
        os.makedirs(output_dir, exist_ok=True)
        
        self.output_dir = output_dir
        logger.info(f"ReportFileExporter初期化: output_dir={output_dir}")
    
    def export_excel(
        self,
        reports: List[Dict],
        filename: str = "output.xlsx",
        append: bool = False
    ) -> str:
        """
        Excel形式で出力
        
        Args:
            reports: 報告書データリスト
            filename: ファイル名
            append: Trueの場合は既存ファイルに追記
        
        Returns:
            出力ファイルパス
        
        Examples:
            >>> exporter = ReportFileExporter()
            >>> path = exporter.export_excel(reports, "output.xlsx")
            >>> print(f"保存完了: {path}")
        """
        output_path = os.path.join(self.output_dir, filename)
        
        # DataFrameを作成
        df = pd.DataFrame(reports, columns=EXCEL_COLUMNS)
        
        try:
            if append and os.path.exists(output_path):
                # 既存ファイルに追記
                logger.info(f"既存ファイルに追記: {output_path}")
                with pd.ExcelWriter(output_path, mode='a', if_sheet_exists='overlay', engine='openpyxl') as writer:
                    # 既存シートの最後の行を取得
                    workbook = load_workbook(output_path)
                    sheet = workbook.active
                    startrow = sheet.max_row
                    
                    # ヘッダーなしで追記
                    df.to_excel(writer, index=False, header=False, startrow=startrow)
            else:
                # 新規ファイル作成
                logger.info(f"新規ファイル作成: {output_path}")
                with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
                    df.to_excel(writer, index=False, header=True)
            
            logger.info(f"Excel保存成功: {output_path} ({len(reports)}件)")
            
        except PermissionError:
            # ファイルロック対策: 別名で保存
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            alt_filename = f"{os.path.splitext(filename)[0]}_{timestamp}.xlsx"
            alt_path = os.path.join(self.output_dir, alt_filename)
            
            with pd.ExcelWriter(alt_path, engine='xlsxwriter') as writer:
                df.to_excel(writer, index=False, header=True)
            
            logger.warning(f"ファイルロック検出、別名で保存: {alt_path}")
            output_path = alt_path
        
        return output_path
    
    def export_json(
        self,
        reports: List[Dict],
        filename: str = "output.json"
    ) -> str:
        """
        JSON形式で出力（全データ）
        
        Args:
            reports: 報告書データリスト
            filename: ファイル名
        
        Returns:
            出力ファイルパス
        """
        output_path = os.path.join(self.output_dir, filename)
        
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(reports, f, ensure_ascii=False, indent=2)
            
            logger.info(f"JSON保存成功: {output_path} ({len(reports)}件)")
            
        except Exception as e:
            logger.error(f"JSON保存失敗: {e}")
            raise
        
        return output_path
    
    def export_json_entries(
        self,
        reports: List[Dict],
        entries_dir: Optional[str] = None
    ) -> str:
        """
        個別JSONエントリとして出力
        
        Args:
            reports: 報告書データリスト
            entries_dir: JSONエントリディレクトリ（Noneの場合はデフォルト）
        
        Returns:
            JSONエントリディレクトリパス
        
        Note:
            各報告書を個別のJSONファイルとして保存します。
            ファイル名: {code}.json
        """
        if entries_dir is None:
            entries_dir = os.path.join(self.output_dir, "json_data")
        
        # ディレクトリ作成
        os.makedirs(entries_dir, exist_ok=True)
        
        # 各報告書を個別ファイルとして保存
        saved_count = 0
        for report in reports:
            code = report.get('code', '')
            if not code:
                logger.warning("code欠如のため個別JSON保存スキップ")
                continue
            
            entry_path = os.path.join(entries_dir, f"{code}.json")
            try:
                with open(entry_path, 'w', encoding='utf-8') as f:
                    json.dump(report, f, ensure_ascii=False, indent=2)
                saved_count += 1
            except Exception as e:
                logger.error(f"個別JSON保存失敗 ({code}): {e}")
        
        logger.info(f"個別JSONエントリ保存完了: {entries_dir} ({saved_count}件)")
        return entries_dir
    
    def export_with_backup(
        self,
        reports: List[Dict],
        base_filename: str = "output"
    ) -> Dict[str, str]:
        """
        最新版とタイムスタンプ付きバックアップの両方を作成
        
        Args:
            reports: 報告書データリスト
            base_filename: ベースファイル名（拡張子なし）
        
        Returns:
            作成されたファイルパスの辞書:
            {
                "latest_excel": "...",
                "latest_json": "...",
                "backup_dir": "...",
                "backup_excel": "...",
                "backup_json": "...",
                "json_entries_dir": "..."
            }
        
        Note:
            設備タブのパターンを踏襲した出力形式です。
            - output/arim-site/reports/ に最新版
            - output/arim-site/reports/backups/YYYYMMDD_HHMMSS/ にバックアップ
        """
        logger.info(f"バックアップ付き出力開始: {len(reports)}件")
        
        # タイムスタンプ
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # ========================================
        # 1. 最新版を facilities/ に保存
        # ========================================
        latest_excel_filename = f"{base_filename}.xlsx"
        latest_json_filename = f"{base_filename}.json"
        
        latest_excel_path = self.export_excel(reports, latest_excel_filename)
        latest_json_path = self.export_json(reports, latest_json_filename)
        
        logger.info(f"最新版Excel作成: {latest_excel_path}")
        logger.info(f"最新版JSON作成: {latest_json_path}")
        
        # ========================================
        # 2. バックアップディレクトリ作成
        # ========================================
        backup_base_dir = os.path.join(OUTPUT_DIR, "arim-site", "reports", "backups")
        backup_dir = os.path.join(backup_base_dir, timestamp)
        os.makedirs(backup_dir, exist_ok=True)
        
        # ========================================
        # 3. バックアップにコピー
        # ========================================
        backup_excel_path = os.path.join(backup_dir, latest_excel_filename)
        backup_json_path = os.path.join(backup_dir, latest_json_filename)
        
        shutil.copy2(latest_excel_path, backup_excel_path)
        shutil.copy2(latest_json_path, backup_json_path)
        
        logger.info(f"バックアップExcel作成: {backup_excel_path}")
        logger.info(f"バックアップJSON作成: {backup_json_path}")
        
        # ========================================
        # 4. 個別JSONエントリ（バックアップ内）
        # ========================================
        json_entries_dir = self.export_json_entries(
            reports,
            entries_dir=os.path.join(backup_dir, "json_data")
        )
        
        logger.info(f"個別JSONエントリ作成: {json_entries_dir}")
        
        # ========================================
        # 5. キャッシュにも反映
        # ========================================
        cache_manager = ReportCacheManager()
        cache_manager.save_entries(reports)

        # ========================================
        # 6. 結果を返す
        # ========================================
        result = {
            "latest_excel": latest_excel_path,
            "latest_json": latest_json_path,
            "backup_dir": backup_dir,
            "backup_excel": backup_excel_path,
            "backup_json": backup_json_path,
            "json_entries_dir": json_entries_dir,
        }
        
        logger.info(f"バックアップ付き出力完了")
        logger.info(f"  最新Excel: {latest_excel_path}")
        logger.info(f"  バックアップ: {backup_dir}")
        logger.info(f"  個別JSONエントリ: {json_entries_dir}")
        
        return result
