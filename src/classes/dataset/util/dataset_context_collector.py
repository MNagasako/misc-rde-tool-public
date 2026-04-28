"""
データセット関連データ読み込みロジック
AIプロンプト生成のためのデータセット関連情報を取得
ARIM課題データ（実験データ・拡張情報・実験データ）も統合対応
"""

import os
import json
import csv
from datetime import date, datetime
from typing import Dict, List, Optional, Any
from config.common import get_dynamic_file_path
from .arim_data_collector import get_arim_data_collector
from .arim_data_collector import get_arim_data_collector

import logging

# ロガー設定
logger = logging.getLogger(__name__)

AI_DOWNLOAD_STORAGE_MODE_TEMP = "temp"
AI_DOWNLOAD_STORAGE_MODE_DATAFILES = "dataFiles"
AI_DOWNLOAD_STORAGE_MODE_DATAFILES_AI = "dataFilesAI"
AI_DOWNLOAD_STORAGE_MODE_DEFAULT = AI_DOWNLOAD_STORAGE_MODE_TEMP
_VALID_AI_DOWNLOAD_STORAGE_MODES = {
    AI_DOWNLOAD_STORAGE_MODE_TEMP,
    AI_DOWNLOAD_STORAGE_MODE_DATAFILES,
    AI_DOWNLOAD_STORAGE_MODE_DATAFILES_AI,
}


class DatasetContextCollector:
    """データセット関連のコンテキストデータ収集クラス"""
    
    def __init__(self):
        self.cache = {}

    def _get_ai_download_storage_mode(self) -> str:
        """AI自動取得ファイルの保存モードを取得する。"""
        try:
            from classes.managers.app_config_manager import AppConfigManager

            manager = AppConfigManager()
            mode = str(
                manager.get(
                    "file_text_extraction.ai_download_storage_mode",
                    AI_DOWNLOAD_STORAGE_MODE_DEFAULT,
                )
                or AI_DOWNLOAD_STORAGE_MODE_DEFAULT
            )
        except Exception as exc:
            logger.debug("AIダウンロード保存モード取得失敗: %s", exc)
            mode = AI_DOWNLOAD_STORAGE_MODE_DEFAULT

        if mode not in _VALID_AI_DOWNLOAD_STORAGE_MODES:
            return AI_DOWNLOAD_STORAGE_MODE_DEFAULT
        return mode

    def _get_persistent_download_base_dir(self, storage_mode: str) -> Optional[str]:
        """恒久保存モード時のベースディレクトリを返す。"""
        from config.common import get_dynamic_file_path

        if storage_mode == AI_DOWNLOAD_STORAGE_MODE_DATAFILES:
            return get_dynamic_file_path("output/rde/data/dataFiles")
        if storage_mode == AI_DOWNLOAD_STORAGE_MODE_DATAFILES_AI:
            return get_dynamic_file_path("output/rde/data/dataFilesAI")
        return None

    def _load_dataset_identity(
        self,
        dataset_id: str,
        data_items: Optional[List[Dict[str, Any]]] = None,
        bearer_token: Optional[str] = None,
        proxy_get=None,
    ) -> Dict[str, str]:
        """保存先決定に必要なデータセット名/課題番号を可能な範囲で解決する。"""
        from config.common import get_dynamic_file_path

        dataset_name = ""
        grant_number = ""

        def _extract_attrs(payload: Any) -> Dict[str, Any]:
            if isinstance(payload, dict):
                if isinstance(payload.get("data"), dict):
                    return payload.get("data", {}).get("attributes", {}) or {}
                return payload.get("attributes", {}) or {}
            return {}

        try:
            dataset_detail_path = get_dynamic_file_path(f"output/rde/data/datasets/{dataset_id}.json")
            if os.path.exists(dataset_detail_path):
                with open(dataset_detail_path, "r", encoding="utf-8") as fp:
                    attrs = _extract_attrs(json.load(fp))
                dataset_name = str(attrs.get("name") or "").strip()
                grant_number = str(attrs.get("grantNumber") or "").strip()
        except Exception as exc:
            logger.debug("dataset detail identity load failed (%s): %s", dataset_id, exc)

        if not dataset_name or not grant_number:
            try:
                dataset_listing_path = get_dynamic_file_path("output/rde/data/dataset.json")
                if os.path.exists(dataset_listing_path):
                    with open(dataset_listing_path, "r", encoding="utf-8") as fp:
                        listing = json.load(fp)
                    for item in list(listing.get("data") or []):
                        if str(item.get("id") or "").strip() != str(dataset_id or "").strip():
                            continue
                        attrs = item.get("attributes", {}) or {}
                        dataset_name = dataset_name or str(attrs.get("name") or "").strip()
                        grant_number = grant_number or str(attrs.get("grantNumber") or "").strip()
                        break
            except Exception as exc:
                logger.debug("dataset listing identity load failed (%s): %s", dataset_id, exc)

        if data_items:
            for item in data_items:
                attrs = item.get("attributes", {}) or {}
                dataset_name = dataset_name or str(attrs.get("datasetName") or "").strip()
                grant_number = grant_number or str(attrs.get("grantNumber") or "").strip()
                if dataset_name and grant_number:
                    break

        if (not dataset_name or not grant_number) and dataset_id and bearer_token and proxy_get:
            try:
                dataset_detail_url = f"https://rde-api.nims.go.jp/datasets/{dataset_id}"
                detail_headers = {
                    "Accept": "application/vnd.api+json",
                    "Authorization": f"Bearer {bearer_token}",
                    "Host": "rde-api.nims.go.jp",
                    "Origin": "https://rde.nims.go.jp",
                    "Referer": "https://rde.nims.go.jp/",
                }
                response = proxy_get(dataset_detail_url, headers=detail_headers)
                if getattr(response, "status_code", None) == 200:
                    attrs = _extract_attrs(response.json())
                    dataset_name = dataset_name or str(attrs.get("name") or "").strip()
                    grant_number = grant_number or str(attrs.get("grantNumber") or "").strip()
            except Exception as exc:
                logger.debug("dataset detail api identity load failed (%s): %s", dataset_id, exc)

        return {
            "dataset_name": dataset_name,
            "grant_number": grant_number,
        }

    def _get_existing_dataset_dirs(
        self,
        dataset_id: str,
        storage_mode: str,
        data_items: Optional[List[Dict[str, Any]]] = None,
        bearer_token: Optional[str] = None,
        proxy_get=None,
    ) -> List[str]:
        """既存ファイル再利用候補のディレクトリを優先順で返す。"""
        from config.common import get_dynamic_file_path
        from classes.data_fetch2.core.logic.fetch2_filelist_logic import replace_invalid_path_chars

        candidates: List[str] = []
        identity = self._load_dataset_identity(
            dataset_id,
            data_items,
            bearer_token=bearer_token,
            proxy_get=proxy_get,
        )
        dataset_name = str(identity.get("dataset_name") or "").strip()
        grant_number = str(identity.get("grant_number") or "").strip()

        if dataset_name and grant_number:
            preferred_base = self._get_persistent_download_base_dir(storage_mode)
            base_dirs = []
            if preferred_base:
                base_dirs.append(preferred_base)
            elif storage_mode == AI_DOWNLOAD_STORAGE_MODE_TEMP:
                base_dirs.append(get_dynamic_file_path("output/rde/data/dataFiles"))
                base_dirs.append(get_dynamic_file_path("output/rde/data/dataFilesAI"))

            safe_dataset_name = replace_invalid_path_chars(dataset_name)
            for base_dir in base_dirs:
                if not base_dir:
                    continue
                candidates.append(os.path.join(base_dir, grant_number, safe_dataset_name))

        # 旧来の dataFiles/{dataset_id} 構造も後方互換で探索する
        if storage_mode == AI_DOWNLOAD_STORAGE_MODE_DATAFILES:
            candidates.append(get_dynamic_file_path(f"output/rde/data/dataFiles/{dataset_id}"))
        elif storage_mode == AI_DOWNLOAD_STORAGE_MODE_DATAFILES_AI:
            candidates.append(get_dynamic_file_path(f"output/rde/data/dataFilesAI/{dataset_id}"))
        else:
            candidates.append(get_dynamic_file_path(f"output/rde/data/dataFiles/{dataset_id}"))
            candidates.append(get_dynamic_file_path(f"output/rde/data/dataFilesAI/{dataset_id}"))

        deduped: List[str] = []
        seen: set[str] = set()
        for path in candidates:
            norm = os.path.normcase(os.path.normpath(path))
            if norm in seen:
                continue
            seen.add(norm)
            deduped.append(path)
        return deduped

    def _load_cached_entry_files_payload(self, entry_id: str) -> Optional[Dict[str, Any]]:
        """dataFiles/<entry_id>.json キャッシュを読み込む。"""
        from config.common import get_dynamic_file_path

        if not entry_id:
            return None

        candidates = (
            get_dynamic_file_path(f"output/rde/data/dataFiles/{entry_id}.json"),
            get_dynamic_file_path(f"output/rde/data/dataFiles/sub/{entry_id}.json"),
        )
        for path in candidates:
            try:
                if not os.path.exists(path):
                    continue
                with open(path, "r", encoding="utf-8") as fh:
                    payload = json.load(fh)
                if isinstance(payload, dict) and isinstance(payload.get("data"), list):
                    return payload
            except Exception as exc:
                logger.debug("cached entry files load failed (%s): %s", path, exc)
        return None

    def _collect_structured_files_from_local_cache(
        self,
        dataset_id: str,
        bearer_token: Optional[str] = None,
        proxy_get=None,
    ) -> List[Dict[str, Any]]:
        """dataEntry / dataFiles キャッシュから STRUCTURED ファイル情報を組み立てる。"""
        from config.common import get_dynamic_file_path

        entry_path = get_dynamic_file_path(f"output/rde/data/dataEntry/{dataset_id}.json")
        if not os.path.exists(entry_path):
            return []

        try:
            with open(entry_path, "r", encoding="utf-8") as fh:
                payload = json.load(fh)
        except Exception as exc:
            logger.debug("dataEntry cache load failed (%s): %s", dataset_id, exc)
            return []

        entries = payload.get("data") or []
        dataset_identity = self._load_dataset_identity(
            dataset_id,
            entries if isinstance(entries, list) else None,
            bearer_token=bearer_token,
            proxy_get=proxy_get,
        )
        structured_files: List[Dict[str, Any]] = []

        if not isinstance(entries, list):
            return structured_files

        for entry in entries:
            if not isinstance(entry, dict):
                continue
            entry_id = str(entry.get("id") or "").strip()
            entry_attrs = entry.get("attributes", {}) or {}
            cached_files = self._load_cached_entry_files_payload(entry_id)
            if not cached_files:
                continue
            for item in list(cached_files.get("data") or []):
                if not isinstance(item, dict):
                    continue
                file_attrs = item.get("attributes", {}) or {}
                if str(file_attrs.get("fileType") or "") != "STRUCTURED":
                    continue
                structured_files.append(
                    {
                        "id": str(item.get("id") or "").strip(),
                        "name": file_attrs.get("fileName", ""),
                        "size": file_attrs.get("fileSize", 0),
                        "media_type": file_attrs.get("mediaType", ""),
                        "tile_name": entry_attrs.get("name", ""),
                        "tile_number": entry_attrs.get("dataNumber", ""),
                        "dataset_name": dataset_identity.get("dataset_name") or entry_attrs.get("datasetName", ""),
                        "grant_number": dataset_identity.get("grant_number") or entry_attrs.get("grantNumber", ""),
                    }
                )

        return structured_files

    def _download_structured_file(
        self,
        *,
        file_id: str,
        file_name: str,
        bearer_token: Optional[str],
        storage_mode: str,
        file_info: Dict[str, Any],
        proxy_get,
    ) -> Optional[str]:
        """設定に応じて一時/恒久保存でSTRUCTUREDファイルを取得する。"""
        if storage_mode == AI_DOWNLOAD_STORAGE_MODE_TEMP:
            import tempfile

            file_download_url = f"https://rde-api.nims.go.jp/files/{file_id}?isDownload=true"
            download_headers = {
                "Authorization": f"Bearer {bearer_token}",
                "Host": "rde-api.nims.go.jp",
                "Origin": "https://rde.nims.go.jp",
                "Referer": "https://rde.nims.go.jp/",
                "Accept": "*/*",
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            }

            file_response = proxy_get(file_download_url, headers=download_headers, stream=True)
            if file_response.status_code != 200:
                logger.warning("ファイルダウンロード失敗: %s (HTTP %s)", file_name, file_response.status_code)
                return None

            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file_name)[1]) as tmp_file:
                for chunk in file_response.iter_content(chunk_size=8192):
                    if chunk:
                        tmp_file.write(chunk)
                file_path = tmp_file.name
            logger.info("AI動的ダウンロード完了(一時): %s -> %s", file_name, file_path)
            return file_path

        base_dir = self._get_persistent_download_base_dir(storage_mode)
        if not base_dir:
            logger.warning("未対応のAI保存モード: %s", storage_mode)
            return None

        if not file_id:
            logger.warning("file_id が無いため恒久保存ダウンロードできません: %s", file_name)
            return None

        from classes.data_fetch2.core.logic.fetch2_filelist_logic import download_file_for_data_id

        saved_path = download_file_for_data_id(
            data_id=file_id,
            bearer_token=bearer_token,
            save_dir_base=base_dir,
            file_name=file_name,
            grantNumber=file_info.get("grant_number"),
            dataset_name=file_info.get("dataset_name"),
            tile_name=file_info.get("tile_name"),
            tile_number=file_info.get("tile_number"),
            parent=None,
        )
        if not saved_path:
            logger.warning("AI恒久保存ダウンロード失敗: %s", file_name)
            return None

        logger.info("AI動的ダウンロード完了(恒久保存:%s): %s -> %s", storage_mode, file_name, saved_path)
        return str(saved_path)
        
    def collect_full_context(self, dataset_id: Optional[str] = None, **form_data) -> Dict[str, Any]:
        """
        データセットの完全なコンテキストデータを収集
        ARIM課題データ（実験データ・拡張情報・実験データ）も統合
        
        Args:
            dataset_id: データセットID（修正時に使用）
            **form_data: フォームから取得した基本データ
            
        Returns:
            収集されたコンテキストデータ
        """
        context = {}
        
        # フォームデータを基本情報として設定
        context.update(form_data)
        
        # 既存の説明文を取得
        context['existing_description'] = form_data.get('description', '')
        
        # 課題番号を取得
        grant_number = form_data.get('grant_number', '').strip()
        
        # 追加データを収集
        if dataset_id:
            # データセットIDが指定されている場合（修正時）
            logger.debug("データセット詳細情報を取得開始: dataset_id=%s", dataset_id)
            details = self._collect_dataset_details(dataset_id)
            context.update(details)
            
            logger.debug("取得した詳細情報のキー: %s", list(details.keys()))
            if 'file_info' in details:
                logger.debug("file_info の長さ: %s 文字", len(details['file_info']))
                logger.debug("file_info の先頭100文字: %s", details['file_info'][:100])
                logger.debug("file_info の型: %s", type(details['file_info']))
                logger.debug("file_info が空: %s", not details['file_info'])
            
            # ファイルツリー情報を個別キーとしても設定（プロンプトテンプレート用）
            if 'file_info' in details and details['file_info']:
                context['file_tree'] = details['file_info']
                logger.debug("[OK] file_tree をセット: %s 文字", len(context['file_tree']))
            else:
                # file_infoが空の場合もメッセージを設定
                context['file_tree'] = details.get('file_info', '（ファイルツリー情報の取得に失敗しました）')
                logger.debug("file_info が空またはFalsy - file_tree = '%s...'", context['file_tree'][:50])
            
            logger.debug("最終的な context['file_tree'] の長さ: %s 文字", len(context.get('file_tree', '')))
            
            # STRUCTUREDファイルのテキスト内容を個別キーとして設定（プロンプトテンプレート用）
            if 'file_contents' in details and details['file_contents']:
                context['text_from_structured_files'] = details['file_contents']
                logger.debug("[OK] text_from_structured_files をセット: %s 文字", len(context['text_from_structured_files']))
            else:
                context['text_from_structured_files'] = details.get('file_contents', '（STRUCTUREDファイルのテキスト抽出に失敗しました）')
                logger.debug("file_contents が空またはFalsy - text_from_structured_files = '%s...'", context['text_from_structured_files'][:50])

            json_payload = details.get('file_contents_json')
            if json_payload:
                context['json_from_structured_files'] = json_payload
            else:
                fallback_json = self._build_structured_json_payload(None, context['text_from_structured_files'])
                context['json_from_structured_files'] = fallback_json
                logger.debug("file_contents_json が空 - json_from_structured_files にフォールバックメッセージを設定")
        else:
            # 新規作成時はフォームデータのみ
            context.update(self._collect_general_data())
            context['file_tree'] = '（新規作成のためファイルツリー情報なし）'
            context['text_from_structured_files'] = '（新規作成のためSTRUCTUREDファイル情報なし）'
            context['json_from_structured_files'] = self._build_structured_json_payload(
                None,
                '（新規作成のためSTRUCTUREDファイル情報なし）'
            )
            
        # ARIM課題データを収集（課題番号が存在する場合）
        if grant_number:
            arim_data = self._collect_arim_data(grant_number, context)
            context.update(arim_data)
            
        return context
        
    def _collect_arim_data(self, grant_number: str, form_data: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
        """
        課題番号からARIM課題データを収集
        AIテスト機能と同じデータソースを使用
        
        Args:
            grant_number: 課題番号
            form_data: フォームデータ（データセット名など）
            
        Returns:
            ARIM課題データ（プロンプトテンプレート用）
        """
        arim_formatted: Dict[str, Any] = {}
        
        try:
            logger.info("ARIM課題データ収集開始: %s", grant_number)
            
            # ARIMデータコレクターを使用してデータを取得
            arim_collector = get_arim_data_collector()
            arim_data = arim_collector.collect_arim_data_by_grant_number(grant_number)
            
            # プロンプトテンプレート用にフォーマット
            arim_formatted = arim_collector.format_for_prompt_template(arim_data)
            
            # 設備IDを抽出してコンテキストに追加
            equipment_ids = arim_collector.extract_equipment_ids(grant_number)
            
            # データセット名からも設備IDを抽出（フォームデータから）
            if form_data:
                dataset_name = form_data.get('name', '')
                if dataset_name:
                    dataset_equipment_ids = arim_collector._extract_equipment_from_dataset_name(dataset_name)
                    for eq_id in dataset_equipment_ids:
                        if eq_id not in equipment_ids:
                            equipment_ids.append(eq_id)
                            logger.info("データセット名から設備ID抽出: %s from %s", eq_id, dataset_name)
                
                # データセットJSONデータからも設備IDを抽出（データセットテンプレートIDを含む）
                dataset_json_data = form_data.get('dataset_json_data')
                if dataset_json_data:
                    json_equipment_ids = arim_collector.extract_equipment_from_dataset_json(dataset_json_data)
                    for eq_id in json_equipment_ids:
                        if eq_id not in equipment_ids:
                            equipment_ids.append(eq_id)
            
            arim_formatted['equipment_ids'] = equipment_ids
            
            logger.info("ARIM課題データ収集完了: %s", arim_formatted.get('collection_summary', 'N/A'))
            logger.info("設備ID抽出結果: %s", equipment_ids)
            
            # 詳細ログ追加
            for key, value in arim_formatted.items():
                if key not in ['collection_summary', 'equipment_ids']:
                    has_data = '[ARIM拡張情報なし]' not in str(value) and '[ARIM課題データ取得エラー]' not in str(value)
                    status = "○" if has_data else "×"
                    logger.debug("%s: %s (%s文字)", key, status, len(str(value)))
            
        except Exception as e:
            logger.warning("ARIM課題データ収集エラー: %s", e)
            # エラー時のフォールバック
            arim_formatted = {
                'dataset_existing_info': '[ARIM課題データ取得エラー]',
                'arim_extension_data': '[ARIM課題データ取得エラー]',
                'arim_experiment_data': '[ARIM課題データ取得エラー]',
                'collection_summary': 'エラーのため取得失敗'
            }
            
        return arim_formatted
        
    def _collect_dataset_details(self, dataset_id: str) -> Dict[str, Any]:
        """
        特定のデータセットの詳細情報を収集
        
        Args:
            dataset_id: データセットID
            
        Returns:
            データセット詳細情報
        """
        details = {}
        
        try:
            # TODO: 実際のデータセット詳細取得ロジックを実装
            # 現在はダミーデータを返す
            
            # ファイル情報を取得
            details['file_info'] = self._get_file_info(dataset_id)
            
            # メタデータを取得
            details['metadata'] = self._get_metadata(dataset_id)
            
            # 関連データセットを取得
            details['related_datasets'] = self._get_related_datasets(dataset_id)
            
            # STRUCTUREDファイルのテキスト内容を取得
            file_contents_result = self._get_file_contents(dataset_id, include_json=True)
            if isinstance(file_contents_result, tuple):
                details['file_contents'], details['file_contents_json'] = file_contents_result
            else:
                details['file_contents'] = file_contents_result
                details['file_contents_json'] = self._build_structured_json_payload(None, file_contents_result)
            
        except Exception as e:
            logger.warning("データセット詳細取得エラー: %s", e)
            details['file_info'] = f'（データセット詳細取得中にエラーが発生しました: {str(e)}）'
            details['metadata'] = '（メタデータ取得失敗）'
            details['related_datasets'] = '（関連データセット取得失敗）'
            details['file_contents'] = '（ファイル内容取得失敗）'
            details['file_contents_json'] = self._build_structured_json_payload(None, details['file_contents'])
            
        return details
        
    def _collect_general_data(self) -> Dict[str, Any]:
        """
        一般的なデータセット情報を収集
        
        Returns:
            一般的なデータセット情報
        """
        general = {}
        
        try:
            # TODO: 一般的なデータセット統計や傾向を取得
            # 現在はダミーデータを返す
            
            general['file_info'] = 'ファイル情報は作成時に自動設定されます'
            general['metadata'] = 'メタデータは登録後に生成されます'
            general['related_datasets'] = '関連データセットは自動検出されます'
            
        except Exception as e:
            logger.warning("一般データ取得エラー: %s", e)
            general['file_info'] = ''
            general['metadata'] = ''
            general['related_datasets'] = ''
            
        return general
        
    def _get_file_info(self, dataset_id: str) -> str:
        """
        データセットのファイル情報を取得（RDE API使用）
        
        Args:
            dataset_id: データセットID
            
        Returns:
            ファイル情報の文字列（プロンプトテンプレート用にフォーマット済み）
        """
        try:
            logger.debug("ファイル情報取得開始: dataset_id=%s", dataset_id)
            from core.bearer_token_manager import BearerTokenManager
            from net.http_helpers import proxy_get
            
            # Bearer Token取得
            bearer_token = BearerTokenManager.get_token_with_relogin_prompt()
            if not bearer_token:
                logger.warning("Bearer Token取得失敗 - ログインが必要です")
                return '（ファイルツリー情報の取得にはログインが必要です。RDEシステムにログインしてから再試行してください）'
            
            logger.debug("Bearer Token取得成功: %s...", bearer_token[:20])
            
            # RDE APIでデータ情報を取得
            api_url = f"https://rde-api.nims.go.jp/data?filter%5Bdataset.id%5D={dataset_id}&sort=-created&page%5Boffset%5D=0&page%5Blimit%5D=100&include=owner%2Csample%2CthumbnailFile%2Cfiles"
            
            logger.debug("API URL: %s", api_url)
            
            headers = {
                "Accept": "application/vnd.api+json",
                "Authorization": f"Bearer {bearer_token}",
                "Host": "rde-api.nims.go.jp",
                "Origin": "https://rde.nims.go.jp",
                "Referer": "https://rde.nims.go.jp/"
            }
            
            response = proxy_get(api_url, headers=headers)
            
            logger.debug("API Response Status: %s", response.status_code)
            
            if response.status_code != 200:
                error_msg = f"（データ情報取得API失敗: HTTP {response.status_code}）"
                logger.warning("%s", error_msg)
                if response.status_code == 401:
                    error_msg += "\n認証エラー - 再ログインが必要です"
                elif response.status_code == 403:
                    error_msg += "\nアクセス権限がありません"
                elif response.status_code == 404:
                    error_msg += "\n指定されたデータセットが見つかりません"
                return error_msg
            
            data = response.json()
            
            logger.debug("API Response Data: %s タイル取得", len(data.get('data', [])))
            
            # ファイルツリー情報をフォーマット
            formatted_result = self._format_file_tree(data)
            logger.debug("フォーマット結果の長さ: %s 文字", len(formatted_result))
            return formatted_result
            
        except Exception as e:
            error_msg = f"（ファイル情報取得中にエラーが発生: {str(e)}）"
            logger.warning("%s", error_msg)
            import traceback
            traceback.print_exc()
            return error_msg
    
    def _format_file_tree(self, api_data: Dict[str, Any]) -> str:
        """
        APIレスポンスからファイルツリー情報をフォーマット
        RDEシステムではファイルはフラット配置（階層構造なし）
        
        Args:
            api_data: RDE API /data のレスポンス
            
        Returns:
            プロンプトテンプレート用にフォーマットされた文字列
        """
        try:
            data_list = api_data.get('data', [])
            included = api_data.get('included', [])
            
            if not data_list:
                return '（このデータセットにはまだデータ（タイル）が登録されていません。データ登録後にファイルツリー情報が利用可能になります）'
            
            # includedをIDでインデックス化
            file_dict = {}
            for item in included:
                if item.get('type') == 'file':
                    file_dict[item['id']] = item.get('attributes', {})
            
            # 各タイル（データ詳細）の情報を整形
            tile_info_list = []
            
            # ファイルタイプの日本語名マッピング
            file_type_names = {
                'MAIN_IMAGE': '主要画像',
                'STRUCTURED': 'データファイル',
                'META': 'メタデータ',
                'ATTACHMENT': '添付ファイル',
                'THUMBNAIL': 'サムネイル'
            }
            
            for idx, data_item in enumerate(data_list, 1):
                attributes = data_item.get('attributes', {})
                data_name = attributes.get('name', '名前なし')
                data_number = attributes.get('dataNumber', idx)
                description = attributes.get('description', '')
                num_files = attributes.get('numberOfFiles', 0)
                num_images = attributes.get('numberOfImageFiles', 0)
                experiment_id = attributes.get('experimentId', '')
                
                # メタデータから主要情報を抽出
                metadata = attributes.get('metadata', {})
                instrument_name = metadata.get('instrument.name', {}).get('value', '')
                
                # ファイル情報を取得
                relationships = data_item.get('relationships', {})
                file_ids = [f['id'] for f in relationships.get('files', {}).get('data', [])]
                
                # タイル基本情報
                tile_info = f"■ タイル#{data_number}: {data_name}"
                if description and description != data_name:
                    tile_info += f"\n  説明: {description}"
                if experiment_id:
                    tile_info += f"\n  実験ID: {experiment_id}"
                if instrument_name:
                    tile_info += f"\n  使用装置: {instrument_name}"
                tile_info += f"\n  ファイル統計: 全{num_files}件 (画像ファイル: {num_images}件)"
                
                # ファイル詳細情報
                if file_ids:
                    file_types = {
                        'MAIN_IMAGE': [],
                        'STRUCTURED': [],
                        'META': [],
                        'ATTACHMENT': [],
                        'THUMBNAIL': []
                    }
                    
                    for file_id in file_ids:
                        if file_id in file_dict:
                            file_attr = file_dict[file_id]
                            file_name = file_attr.get('fileName', '')
                            file_type = file_attr.get('fileType', 'UNKNOWN')
                            file_size = file_attr.get('fileSize', 0)
                            media_type = file_attr.get('mediaType', '')
                            
                            # ファイルサイズを読みやすく変換
                            if file_size < 1024:
                                size_str = f"{file_size}B"
                            elif file_size < 1024 * 1024:
                                size_str = f"{file_size / 1024:.1f}KB"
                            else:
                                size_str = f"{file_size / (1024 * 1024):.1f}MB"
                            
                            file_info_item = {
                                'name': file_name,
                                'size': size_str,
                                'type': media_type
                            }
                            
                            if file_type in file_types:
                                file_types[file_type].append(file_info_item)
                            else:
                                file_types['STRUCTURED'].append(file_info_item)
                    
                    # ファイルタイプごとに整形して表示（意味のある順序で）
                    for ftype in ['MAIN_IMAGE', 'STRUCTURED', 'ATTACHMENT', 'META', 'THUMBNAIL']:
                        if file_types[ftype]:
                            type_name = file_type_names.get(ftype, ftype)
                            tile_info += f"\n  【{type_name}】"
                            for f_item in file_types[ftype]:
                                tile_info += f"\n    - {f_item['name']} ({f_item['size']}, {f_item['type']})"
                
                tile_info_list.append(tile_info)
            
            # サマリー情報
            total_tiles = len(data_list)
            total_files = sum(d.get('attributes', {}).get('numberOfFiles', 0) for d in data_list)
            total_images = sum(d.get('attributes', {}).get('numberOfImageFiles', 0) for d in data_list)
            
            result = f"【データセット内ファイル構成】\n"
            result += f"タイル数: {total_tiles}件 / 全ファイル数: {total_files}件 (画像: {total_images}件)\n"
            result += f"※ RDEシステムではファイルはフラット配置（階層構造なし）\n\n"
            result += "\n\n".join(tile_info_list)
            
            logger.info("ファイルツリー情報を生成しました: %sタイル, %sファイル", total_tiles, total_files)
            return result
            
        except Exception as e:
            logger.error("ファイルツリーフォーマットエラー: %s", e)
            import traceback
            traceback.print_exc()
            return f'（ファイルツリーのフォーマット中にエラーが発生しました: {str(e)}）'
            
    def _get_metadata(self, dataset_id: str) -> str:
        """
        データセットのメタデータを取得
        
        Args:
            dataset_id: データセットID
            
        Returns:
            メタデータの文字列
        """
        try:
            # TODO: 実際のメタデータ取得ロジック
            # 例: RDE API呼び出し、データベースクエリなど
            
            # ダミー実装
            metadata_items = [
                f"作成日: 2024-01-01",
                f"最終更新: 2024-02-01", 
                f"データタイプ: 実験データ",
                f"分野: 材料科学"
            ]
            
            return '; '.join(metadata_items)
            
        except Exception as e:
            logger.warning("メタデータ取得エラー: %s", e)
            return ''
            
    def _get_related_datasets(self, dataset_id: str) -> str:
        """
        関連するデータセットを取得
        
        Args:
            dataset_id: データセットID
            
        Returns:
            関連データセットの文字列
        """
        try:
            # TODO: 実際の関連データセット検索ロジック
            # 例: 同じ課題番号、同じ研究者、類似タグなど
            
            # ダミー実装
            related_list = [
                f"同一課題の関連データセット: 3件",
                f"同じ研究者による関連データセット: 5件",
                f"類似分野のデータセット: 10件"
            ]
            
            return '; '.join(related_list)
            
        except Exception as e:
            logger.warning("関連データセット取得エラー: %s", e)
            return ''
    
    def _get_file_contents(self, dataset_id: str, include_json: bool = False):
        """
        データセット内のSTRUCTUREDファイルからテキスト内容を抽出
        
        Args:
            dataset_id: データセットID
            include_json: Trueの場合はJSON文字列も同時に返す
            
        Returns:
            抽出されたテキスト内容、または (テキスト, JSON文字列) のタプル
        """
        def _finalize_response(
            text_value: str,
            contents: Optional[Dict[str, str]] = None,
            json_ready: Optional[Dict[str, Any]] = None,
        ):
            """テキストとJSONの返却形式を整形"""
            if include_json:
                payload_source = json_ready or contents
                json_payload = self._build_structured_json_payload(payload_source, text_value)
                return text_value, json_payload
            return text_value

        try:
            logger.debug("ファイル内容抽出開始: dataset_id=%s", dataset_id)
            from core.bearer_token_manager import BearerTokenManager
            from net.http_helpers import proxy_get
            from classes.dataset.util.file_text_extractor import get_file_text_extractor, format_extracted_files_for_prompt
            import tempfile
            storage_mode = self._get_ai_download_storage_mode()

            # Bearer Token取得（既存ファイル使用時は不要だが、API呼び出し用に試行）
            bearer_token = BearerTokenManager.get_token_with_relogin_prompt()
            if not bearer_token:
                logger.warning("Bearer Token取得失敗 - 既存ダウンロード済みファイルのみ使用します")
                # トークンがなくても既存ファイルからの抽出は試みる
            
            # RDE APIでデータ情報を取得（トークンがある場合のみ）
            structured_files = []
            data_list: List[Dict[str, Any]] = []
            
            if bearer_token:
                api_url = f"https://rde-api.nims.go.jp/data?filter%5Bdataset.id%5D={dataset_id}&sort=-created&page%5Boffset%5D=0&page%5Blimit%5D=100&include=owner%2Csample%2CthumbnailFile%2Cfiles"
                
                headers = {
                    "Accept": "application/vnd.api+json",
                    "Authorization": f"Bearer {bearer_token}",
                    "Host": "rde-api.nims.go.jp",
                    "Origin": "https://rde.nims.go.jp",
                    "Referer": "https://rde.nims.go.jp/"
                }
                
                response = proxy_get(api_url, headers=headers)
                
                if response.status_code == 200:
                    data = response.json()
                    data_list = data.get('data', [])
                    included = data.get('included', [])
                    
                    logger.debug(f"API応答: data件数={len(data_list)}, included件数={len(included)}")
                    
                    if data_list:
                        # includedからファイル情報を抽出
                        file_dict = {}
                        file_types_found = {}
                        for item in included:
                            if item.get('type') == 'file':
                                file_dict[item['id']] = item.get('attributes', {})
                                # デバッグ: fileTypeを集計
                                ft = item.get('attributes', {}).get('fileType', 'UNKNOWN')
                                file_types_found[ft] = file_types_found.get(ft, 0) + 1
                        
                        logger.debug(f"ファイル情報取得: {len(file_dict)}件, タイプ別={file_types_found}")
                        
                        # STRUCTUREDファイルのみをフィルタリング
                        for data_item in data_list:
                            data_attrs = data_item.get('attributes', {}) or {}
                            relationships = data_item.get('relationships', {})
                            file_ids = [f['id'] for f in relationships.get('files', {}).get('data', [])]
                            
                            for file_id in file_ids:
                                if file_id in file_dict:
                                    file_attr = file_dict[file_id]
                                    file_type = file_attr.get('fileType', '')
                                    
                                    if file_type == 'STRUCTURED':
                                        structured_files.append({
                                            'id': file_id,
                                            'name': file_attr.get('fileName', ''),
                                            'size': file_attr.get('fileSize', 0),
                                            'media_type': file_attr.get('mediaType', ''),
                                            'tile_name': data_attrs.get('name', ''),
                                            'tile_number': data_attrs.get('dataNumber', ''),
                                            'dataset_name': data_attrs.get('datasetName', ''),
                                            'grant_number': data_attrs.get('grantNumber', ''),
                                        })
                        
                        logger.info(f"API経由でSTRUCTUREDファイル検出: {len(structured_files)}件")
                else:
                    logger.warning(f"ファイル情報取得API失敗: HTTP {response.status_code}")
            
            # API情報がない場合は、既存ダウンロード済みファイルから直接検索
            import glob
            
            existing_dataset_dirs = self._get_existing_dataset_dirs(
                dataset_id,
                storage_mode,
                data_list,
                bearer_token=bearer_token,
                proxy_get=proxy_get,
            )

            if not structured_files:
                structured_files = self._collect_structured_files_from_local_cache(
                    dataset_id,
                    bearer_token=bearer_token,
                    proxy_get=proxy_get,
                )
                if structured_files:
                    logger.info("ローカルキャッシュ経由でSTRUCTUREDファイル検出: %s件", len(structured_files))

            if not structured_files:
                extractor_temp = get_file_text_extractor()
                for dataset_files_dir in existing_dataset_dirs:
                    if not os.path.exists(dataset_files_dir):
                        continue
                    logger.info("API情報なし - 既存ファイルから直接検索: %s", dataset_files_dir)
                    all_files = glob.glob(os.path.join(dataset_files_dir, '**', '*'), recursive=True)
                    for file_path in all_files:
                        if os.path.isfile(file_path):
                            file_name = os.path.basename(file_path)

                            # 書誌情報JSONファイルを除外（UUIDパターン + .json）
                            # 例: 4a932435-495b-4394-999a-d42136066b04.json, *_anonymized.json
                            if file_name.endswith('_anonymized.json'):
                                continue
                            # UUID形式のJSONファイルを除外（8-4-4-4-12の形式）
                            import re
                            if re.match(r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}\.json$', file_name, re.IGNORECASE):
                                continue

                            # 抽出可能なファイルのみをstructured_filesに追加
                            if extractor_temp.is_extractable(file_name):
                                structured_files.append({
                                    'id': '',  # file_idは不明
                                    'name': file_name,
                                    'size': os.path.getsize(file_path),
                                    'media_type': '',
                                    'local_path': file_path,  # 既存ファイルのパスを保持
                                })
                logger.info("既存ファイルから抽出可能ファイル検出: %s件", len(structured_files))
            
            if not structured_files:
                no_structured_msg = '（このデータセットにはSTRUCTUREDタイプのファイルが含まれていません。また、既存のダウンロード済みファイルも見つかりませんでした）'
                return _finalize_response(no_structured_msg)
            
            # テキスト抽出器を取得
            extractor = get_file_text_extractor()
            
            # 各ファイルをダウンロードしてテキスト抽出（最大10ファイルまで）
            extracted_contents = {}
            json_ready_contents = {}
            max_files = 10
            dataset_identity = self._load_dataset_identity(
                dataset_id,
                data_list,
                bearer_token=bearer_token,
                proxy_get=proxy_get,
            )
            
            for idx, file_info in enumerate(structured_files[:max_files], 1):
                file_name = file_info['name']
                file_id = file_info.get('id', '')
                
                # 画像ファイルは除外（拡張子ベース）
                if not extractor.is_extractable(file_name):
                    logger.debug(f"テキスト抽出非対応ファイルをスキップ: {file_name}")
                    continue
                
                try:
                    file_path = None
                    
                    # 0. API情報取得時に既存ファイルパスが判明している場合はそれを使用
                    if 'local_path' in file_info:
                        file_path = file_info['local_path']
                        logger.info(f"既存ファイル（ダイレクト）を使用: {file_name} ({file_path})")
                    
                    # 1. 既存のダウンロード済みファイルを検索
                    if not file_path:
                        for dataset_files_dir in existing_dataset_dirs:
                            if not os.path.exists(dataset_files_dir):
                                continue
                            search_pattern = os.path.join(dataset_files_dir, '**', file_name)
                            matching_files = glob.glob(search_pattern, recursive=True)
                            if matching_files:
                                file_path = matching_files[0]
                                logger.info(f"既存ファイルを使用: {file_name} ({file_path})")
                                break
                    
                    # 2. 既存ファイルが見つからない場合は動的ダウンロード
                    if not file_path:
                        logger.debug("既存ファイルなし、動的ダウンロード開始: %s (mode=%s)", file_name, storage_mode)
                        enriched_file_info = dict(file_info)
                        enriched_file_info["dataset_name"] = str(
                            file_info.get("dataset_name") or dataset_identity.get("dataset_name") or ""
                        )
                        enriched_file_info["grant_number"] = str(
                            file_info.get("grant_number") or dataset_identity.get("grant_number") or ""
                        )
                        file_path = self._download_structured_file(
                            file_id=file_id,
                            file_name=file_name,
                            bearer_token=bearer_token,
                            storage_mode=storage_mode,
                            file_info=enriched_file_info,
                            proxy_get=proxy_get,
                        )
                        if not file_path:
                            continue
                    
                    # 3. テキスト抽出（既存ファイルまたはダウンロードしたファイル）
                    is_temp_file = file_path and file_path.startswith(tempfile.gettempdir())
                    
                    try:
                        extracted_text = extractor.extract_text(file_path, file_name)
                        json_ready_contents[file_name] = self._create_json_ready_entry(
                            file_path,
                            file_name,
                            extracted_text,
                            extractor
                        )
                        
                        if extracted_text:
                            extracted_contents[file_name] = extracted_text
                            logger.info(f"テキスト抽出成功: {file_name} ({len(extracted_text)}文字)")
                        else:
                            logger.debug(f"テキスト抽出失敗: {file_name}")
                    finally:
                        # 一時ファイル（動的ダウンロードしたファイル）のみ削除
                        if is_temp_file:
                            try:
                                os.unlink(file_path)
                            except:
                                pass
                
                except Exception as e:
                    logger.warning(f"ファイル処理エラー ({file_name}): {e}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            # 抽出結果をフォーマット
            formatted_result = format_extracted_files_for_prompt(extracted_contents)
            logger.info(f"ファイル内容抽出完了: {len(extracted_contents)}件のファイルから {len(formatted_result)}文字を抽出")
            
            return _finalize_response(formatted_result, extracted_contents, json_ready_contents)
            
        except Exception as e:
            error_msg = f"（ファイル内容取得中にエラーが発生: {str(e)}）"
            logger.warning("%s", error_msg)
            import traceback
            traceback.print_exc()
            return _finalize_response(error_msg)

    def _create_json_ready_entry(self, file_path: Optional[str], file_name: str, extracted_text: Optional[str], extractor) -> Any:
        """構造化JSON用の値を生成（CSV/XLSXは配列化、それ以外はテキスト）"""
        try:
            if not file_path or not os.path.exists(file_path):
                return extracted_text or ''
            _, ext = os.path.splitext(file_path)
            ext_lower = ext.lower()
            if ext_lower in {'.csv', '.tsv'}:
                delimiter = ',' if ext_lower == '.csv' else '\t'
                rows = self._load_csv_rows_for_json(file_path, delimiter, getattr(extractor, 'excel_max_rows', 1000))
                if rows:
                    return rows
            if ext_lower in {'.xlsx', '.xls', '.xlsm'}:
                sheets = self._load_excel_rows_for_json(file_path, extractor)
                if sheets:
                    return sheets
        except Exception as exc:
            logger.warning("構造化JSON変換エラー (%s): %s", file_name, exc)
        return extracted_text or ''

    def _load_csv_rows_for_json(self, file_path: str, delimiter: str, max_rows: int) -> List[List[str]]:
        rows: List[List[str]] = []
        for encoding in self._get_text_encodings():
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as fp:
                    reader = csv.reader(fp, delimiter=delimiter)
                    for row in reader:
                        rows.append(row)
                        if len(rows) >= max_rows:
                            break
                if rows:
                    return rows
            except UnicodeDecodeError:
                continue
            except Exception as exc:
                logger.debug("CSV構造化読み込みエラー (%s, %s): %s", file_path, encoding, exc)
                break
        return rows

    def _load_excel_rows_for_json(self, file_path: str, extractor) -> List[Dict[str, Any]]:
        sheets: List[Dict[str, Any]] = []
        try:
            import warnings
            import openpyxl
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            try:
                process_all = getattr(extractor, 'excel_all_sheets', True)
                sheet_names = wb.sheetnames if process_all else [wb.sheetnames[0]] if wb.sheetnames else []
                max_rows = getattr(extractor, 'excel_max_rows', 1000)
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]
                    rows: List[List[Any]] = []
                    for row in sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
                        if not any(row):
                            continue
                        rows.append([self._convert_cell_value_for_json(cell) for cell in row])
                        if len(rows) >= max_rows:
                            break
                    if rows:
                        sheets.append({'name': sheet_name, 'rows': rows})
            finally:
                wb.close()
        except ImportError:
            logger.error("openpyxlがインストールされていません (構造化JSON)")
        except Exception as exc:
            logger.warning("Excel構造化JSON変換エラー (%s): %s", file_path, exc)
        return sheets

    def _convert_cell_value_for_json(self, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, (int, float, bool)):
            return value
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return str(value)

    def _get_text_encodings(self) -> List[str]:
        return ['utf-8', 'utf-8-sig', 'cp932', 'shift-jis', 'euc-jp', 'iso-2022-jp']

    def _build_structured_json_payload(self, file_contents: Optional[Dict[str, Any]], fallback_message: str) -> str:
        """STRUCTUREDファイル内容をJSON文字列として整形"""
        if file_contents:
            return json.dumps(file_contents, ensure_ascii=False, indent=2)
        safe_message = fallback_message or '（STRUCTUREDファイルのテキスト抽出に失敗しました）'
        return json.dumps({'message': safe_message}, ensure_ascii=False)


# グローバルインスタンス
_dataset_context_collector = None

def get_dataset_context_collector() -> DatasetContextCollector:
    """データセットコンテキストコレクターのシングルトンインスタンスを取得"""
    global _dataset_context_collector
    if _dataset_context_collector is None:
        _dataset_context_collector = DatasetContextCollector()
    return _dataset_context_collector
