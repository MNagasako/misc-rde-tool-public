"""
AI拡張設定管理モジュール
AI拡張機能のボタン設定とプロンプトファイルの管理を行う
"""

import os
import json
import re
import copy
from dataclasses import dataclass
from functools import lru_cache
from typing import Dict, Iterable, List, Optional, Tuple
from config.common import get_dynamic_file_path
from classes.ai.util.prompt_assembly import build_prompt

import logging

# ロガー設定
logger = logging.getLogger(__name__)

_PROMPT_PLACEHOLDER_PATTERN = re.compile(r"{([A-Za-z0-9_]+)}")
_DATAPORTAL_MASTER_CACHE: Dict[str, str] = {}
_STATIC_MATERIAL_INDEX_CACHE: Optional[str] = None
_BUTTON_PROMPT_ASSEMBLY_FIELD_MAP = {
    'prompt_assembly_mode': 'mode',
    'prompt_assembly_fallback_behavior': 'fallback_behavior',
    'prompt_assembly_min_candidates': 'min_candidates',
    'prompt_assembly_max_candidates': 'max_candidates',
    'prompt_assembly_max_chars': 'max_chars',
    'prompt_assembly_max_token_estimate': 'max_token_estimate',
    'prompt_assembly_debug_save_enabled': 'debug_save_enabled',
}


def infer_ai_suggest_target_kind(button_config: Dict) -> str:
        """Infer target kind for an AI suggest button.

        Backward compatible default:
            - JSON重要技術領域3 (id: json_suggest_important_tech_area3) is for report
            - everything else is for dataset (AI拡張/従来)

        Returns:
            "dataset" or "report"
        """
        if not isinstance(button_config, dict):
                return "dataset"
        raw = (button_config.get('target_kind') or '').strip().lower()
        if raw in {"dataset", "report"}:
                return raw

        button_id = (button_config.get('id') or '').strip()
        label = (button_config.get('label') or '').strip()
        if button_id == 'json_suggest_important_tech_area3' or label == 'JSON重要技術領域3':
                return "report"
        return "dataset"


@dataclass(frozen=True)
class TemplatePlaceholder:
    name: str
    description: str
    example: str = ""
    source: str = ""


def _safe_str(value) -> str:
    if value is None:
        return ""
    try:
        return str(value)
    except Exception:
        return ""


def _sanitize_placeholder_key(text: str) -> str:
    """Convert arbitrary column/header text into a safe placeholder key.

    Note: Japanese headers may become empty; in that case we simply skip alias.
    """
    if text is None:
        return ""
    s = _safe_str(text).strip().lower()
    if not s:
        return ""
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    if not s:
        return ""
    if s[0].isdigit():
        s = f"col_{s}"
    return s


def _read_excel_headers(file_path: str) -> List[str]:
    """Read header row (column names) from an Excel file."""
    if not file_path or not os.path.exists(file_path):
        return []
    # Prefer pandas if available (fast for headers), else fallback to openpyxl.
    try:
        import pandas as pd  # type: ignore

        df = pd.read_excel(file_path, nrows=0)
        cols = [c for c in df.columns if c is not None]
        return [str(c) for c in cols]
    except Exception:
        pass

    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not row:
            return []
        headers = [h for h in row if h is not None]
        return [str(h) for h in headers]
    except Exception:
        return []


@lru_cache(maxsize=32)
def _load_converted_xlsx_records_cached(file_path: str, mtime: float) -> List[Dict]:
    """Load converted.xlsx into list-of-dicts. Cached by path+mtime."""
    try:
        import pandas as pd  # type: ignore

        df = pd.read_excel(file_path)
        return df.to_dict('records')
    except Exception as e:
        logger.debug("converted.xlsx 読み込みに失敗: %s", e)
        return []


def _load_converted_xlsx_records(file_path: str) -> List[Dict]:
    if not file_path or not os.path.exists(file_path):
        return []
    try:
        mtime = os.path.getmtime(file_path)
    except Exception:
        mtime = 0.0
    return _load_converted_xlsx_records_cached(file_path, mtime)


def _find_matching_record_by_grant_number(grant_number: str, records: List[Dict]) -> Optional[Dict]:
    if not grant_number or not records:
        return None
    target = str(grant_number).strip()
    if not target:
        return None

    candidate_keys = [
        '課題番号', 'ARIMNO', 'ARIM ID', 'ARIMID', 'grant_number', 'GrantNumber',
    ]

    for key in candidate_keys:
        for rec in records:
            try:
                v = rec.get(key)
            except Exception:
                continue
            if v is None:
                continue
            if str(v).strip() == target:
                return rec
    return None


def load_converted_xlsx_report_entries() -> List[Dict]:
    """Load converted.xlsx as list-of-dicts for UI usage.

    This is intended for the AISuggestionDialog "報告書" tab.
    It returns raw records as read from Excel (column names preserved).
    """
    report_path = get_dynamic_file_path("output/arim-site/reports/converted.xlsx")
    if not os.path.exists(report_path):
        return []
    return _load_converted_xlsx_records(report_path)


def placeholders_from_converted_xlsx_record(record: Dict) -> Dict[str, str]:
    """Convert a converted.xlsx record (row dict) into placeholder dict.

    - Keeps raw column names: {<column>}
    - Adds ASCII-safe aliases: {converted_xlsx_<sanitized>}
    - Derives arim_report_* (fallback) and report_* aliases
    """
    if not record:
        return {}

    placeholders: Dict[str, str] = {}

    for col, value in record.items():
        if col is None:
            continue
        col_name = str(col)
        placeholders[col_name] = _safe_str(value)

        alias = _sanitize_placeholder_key(col_name)
        if alias:
            placeholders[f"converted_xlsx_{alias}"] = _safe_str(value)

    # converted.xlsx 由来の arim_report_* を補完
    try:
        derived_arim = derive_arim_report_placeholders_from_converted(placeholders)
        for k, v in derived_arim.items():
            placeholders.setdefault(k, _safe_str(v))
        derived_report = derive_report_aliases_from_arim_report({**placeholders, **derived_arim})
        for k, v in derived_report.items():
            placeholders.setdefault(k, _safe_str(v))
    except Exception:
        pass

    return placeholders


def load_converted_xlsx_placeholders(grant_number: str) -> Dict[str, str]:
    """Load output-side converted.xlsx row data and expose each column as placeholders.

    - Raw column names are available as placeholders: {<column>}
    - Additionally, ASCII-safe aliases are added: {converted_xlsx_<sanitized>}

    Existing implementation is preserved; this is additive.
    """
    placeholders: Dict[str, str] = {}

    # User requested path
    report_path = get_dynamic_file_path("output/arim-site/reports/converted.xlsx")
    if not os.path.exists(report_path):
        return placeholders

    records = _load_converted_xlsx_records(report_path)
    if not records:
        return placeholders

    record = _find_matching_record_by_grant_number(grant_number, records)
    if record is None:
        # If only one record exists, treat it as the target.
        if len(records) == 1:
            record = records[0]
        else:
            return placeholders

    for col, value in record.items():
        if col is None:
            continue
        col_name = str(col)
        placeholders[col_name] = _safe_str(value)

        alias = _sanitize_placeholder_key(col_name)
        if alias:
            placeholders[f"converted_xlsx_{alias}"] = _safe_str(value)

    return placeholders


def _first_non_empty(values: Iterable[str]) -> str:
    for v in values:
        if v is None:
            continue
        s = _safe_str(v).strip()
        if s:
            return s
    return ""


def derive_arim_report_placeholders_from_converted(converted_placeholders: Dict[str, str]) -> Dict[str, str]:
    """converted.xlsx の列（2バイト列名/標準列名混在）から arim_report_* を導出する。

    目的: ネット取得(arim_report_fetcher)が失敗/未実行でも、converted.xlsx があれば
    同じ arim_report_* キーでテンプレートが動くようにする。
    """
    if not converted_placeholders:
        return {}

    def pick(*candidates: str) -> str:
        return _first_non_empty(converted_placeholders.get(k, "") for k in candidates)

    derived: Dict[str, str] = {}

    derived['arim_report_project_number'] = pick(
        '課題番号',
        '課題番号 / Project Issue Number',
        'ARIMNO',
    )
    derived['arim_report_title'] = pick(
        '利用課題名',
        '利用課題名 / Title',
    )
    derived['arim_report_institute'] = pick(
        '利用した実施機関',
        '利用した実施機関 / Support Institute',
    )
    derived['arim_report_usage_type'] = pick(
        '機関外・機関内の利用',
        '機関外・機関内の利用 / External or Internal Use',
    )
    derived['arim_report_keywords'] = pick(
        'キーワード',
        'キーワード / Keywords',
    )

    # 技術領域（主/副）
    derived['arim_report_cross_tech_main'] = pick(
        '横断技術領域・主',
        '横断技術領域（主）',
    )
    derived['arim_report_cross_tech_sub'] = pick(
        '横断技術領域・副',
        '横断技術領域（副）',
    )
    derived['arim_report_important_tech_main'] = pick(
        '重要技術領域・主',
        '重要技術領域（主）',
    )
    derived['arim_report_important_tech_sub'] = pick(
        '重要技術領域・副',
        '重要技術領域（副）',
    )

    # 空は落とす
    return {k: v for k, v in derived.items() if _safe_str(v).strip()}


def derive_report_aliases_from_arim_report(arim_report_data: Dict[str, str]) -> Dict[str, str]:
    """arim_report_* から report_*（欧文化エイリアス）を作る。"""
    if not arim_report_data:
        return {}
    aliases: Dict[str, str] = {}
    mapping = {
        'arim_report_project_number': 'report_project_number',
        'arim_report_title': 'report_title',
        'arim_report_institute': 'report_institute',
        'arim_report_usage_type': 'report_usage_type',
        'arim_report_semiconductor': 'report_semiconductor',
        'arim_report_tech_area': 'report_tech_area',
        'arim_report_keywords': 'report_keywords',
        'arim_report_user_name': 'report_user_name',
        'arim_report_affiliation': 'report_affiliation',
        'arim_report_collaborators': 'report_collaborators',
        'arim_report_supporters': 'report_supporters',
        'arim_report_support_type': 'report_support_type',
        'arim_report_abstract': 'report_abstract',
        'arim_report_experimental': 'report_experimental',
        'arim_report_results': 'report_results',
        'arim_report_remarks': 'report_remarks',
        'arim_report_publications': 'report_publications',
        'arim_report_presentations': 'report_presentations',
        'arim_report_patents': 'report_patents',
        'arim_report_cross_tech_main': 'report_cross_tech_main',
        'arim_report_cross_tech_sub': 'report_cross_tech_sub',
        'arim_report_important_tech_main': 'report_important_tech_main',
        'arim_report_important_tech_sub': 'report_important_tech_sub',
    }
    for src, dst in mapping.items():
        v = arim_report_data.get(src)
        if v is None:
            continue
        sv = _safe_str(v).strip()
        if not sv:
            continue
        aliases[dst] = sv
    return aliases


def list_available_placeholders() -> List[TemplatePlaceholder]:
    """Return the full catalog of placeholders supported by prompt editing UI.

    This is intentionally conservative: it reflects placeholders actually implemented
    in code paths used by AI拡張 (dataset suggestion / AI suggest).
    """

    placeholders: List[TemplatePlaceholder] = []

    def add(name: str, description: str, example: str = "", source: str = ""):
        placeholders.append(TemplatePlaceholder(name=name, description=description, example=example, source=source))

    # 基本（データセット）
    add("name", "データセット名", "サンプルデータセット", "dataset")
    add("type", "データセットタイプ（dataset_type のエイリアス）", "experimental", "dataset")
    add("dataset_type", "データセットタイプ", "experimental", "dataset")
    add("grant_number", "課題番号", "JPMXP1234567890", "dataset")
    add("description", "既存の説明文（existing_description のエイリアス）", "説明文...", "dataset")
    add("existing_description", "既存の説明文", "説明文...", "dataset")

    # 構造化ファイル/ファイルツリー
    add("file_tree", "ファイル構成（file_info を簡略化したキー）", "...", "dataset_context")
    add("text_from_structured_files", "STRUCTURED ファイルから抽出したテキスト", "...", "dataset_context")
    add("json_from_structured_files", "STRUCTURED ファイルのJSON表現", "{...}", "dataset_context")

    # ARIM課題データ（ローカルExcel等から）
    add("dataset_existing_info", "ARIM課題データ: データセット既存情報", "...", "arim_data_collector")
    add("arim_extension_data", "ARIM課題データ: 拡張情報（converted.xlsx由来 / 要約済み）", "...", "arim_data_collector")
    add("arim_experiment_data", "ARIM課題データ: 実験情報（arim_exp.xlsx由来）", "...", "arim_data_collector")
    add("arim_detailed_experiment", "ARIM課題データ: 拡張実験情報（拡張情報内の実験/結果と考察）", "...", "arim_data_collector")
    add("experiment_summary", "実験データ件数などのサマリー", "実験データ件数: 0件", "arim_data_collector")
    add("equipment_ids", "抽出された設備ID一覧", "['TU-507']", "arim_data_collector")

    # MI/装置
    add("material_index_data", "マテリアルインデックスデータ（整形済み）", "{...}", "ai_extension")
    add("equipment_data", "装置情報データ（整形済み）", "...", "ai_extension")
    add("static_material_index", "静的マテリアルインデックス (input/ai/MI.json)", "{...JSON...}", "static")

    # データポータルマスタ
    add("dataportal_material_index", "データポータル: マテリアルインデックスマスタ", "{...JSON...}", "dataportal_master")
    add("dataportal_tag", "データポータル: タグマスタ", "{...JSON...}", "dataportal_master")
    add("dataportal_equipment", "データポータル: 装置分類マスタ", "{...JSON...}", "dataportal_master")

    # ARIM利用報告書（ネット取得）
    # 実装は arim_report_fetcher.map_header_to_key に準拠
    add("arim_report_project_number", "ARIM利用報告書: 課題番号", "...", "arim_report")
    add("arim_report_title", "ARIM利用報告書: 利用課題名", "...", "arim_report")
    add("arim_report_institute", "ARIM利用報告書: 利用した実施機関", "...", "arim_report")
    add("arim_report_usage_type", "ARIM利用報告書: 機関外・機関内の利用", "...", "arim_report")
    add("arim_report_semiconductor", "ARIM利用報告書: ARIM半導体基盤PF関連課題", "...", "arim_report")
    add("arim_report_tech_area", "ARIM利用報告書: 技術領域", "...", "arim_report")
    add("arim_report_keywords", "ARIM利用報告書: キーワード", "...", "arim_report")
    add("arim_report_cross_tech_main", "ARIM利用報告書: 横断技術領域（主）", "...", "arim_report")
    add("arim_report_cross_tech_sub", "ARIM利用報告書: 横断技術領域（副）", "...", "arim_report")
    add("arim_report_important_tech_main", "ARIM利用報告書: 重要技術領域（主）", "...", "arim_report")
    add("arim_report_important_tech_sub", "ARIM利用報告書: 重要技術領域（副）", "...", "arim_report")
    add("arim_report_user_name", "ARIM利用報告書: 利用者名（課題申請者）", "...", "arim_report")
    add("arim_report_affiliation", "ARIM利用報告書: 所属名", "...", "arim_report")
    add("arim_report_collaborators", "ARIM利用報告書: 共同利用者氏名", "...", "arim_report")
    add("arim_report_supporters", "ARIM利用報告書: ARIM実施機関支援担当者", "...", "arim_report")
    add("arim_report_support_type", "ARIM利用報告書: 利用形態", "...", "arim_report")
    add("arim_report_abstract", "ARIM利用報告書: 概要（目的・用途・実施内容）", "...", "arim_report")
    add("arim_report_experimental", "ARIM利用報告書: 実験", "...", "arim_report")
    add("arim_report_results", "ARIM利用報告書: 結果と考察", "...", "arim_report")
    add("arim_report_remarks", "ARIM利用報告書: その他・特記事項（参考文献・謝辞等）", "...", "arim_report")
    add("arim_report_publications", "ARIM利用報告書: 論文・プロシーディング", "...", "arim_report")
    add("arim_report_presentations", "ARIM利用報告書: 口頭発表/ポスター発表/その他", "...", "arim_report")
    add("arim_report_patents", "ARIM利用報告書: 特許", "...", "arim_report")

    # arim_report_* の英語エイリアス（欧文化）
    add("report_project_number", "Report: Project number (alias of arim_report_project_number)", "...", "alias")
    add("report_title", "Report: Title (alias of arim_report_title)", "...", "alias")
    add("report_institute", "Report: Support institute (alias of arim_report_institute)", "...", "alias")
    add("report_usage_type", "Report: External/Internal use (alias of arim_report_usage_type)", "...", "alias")
    add("report_semiconductor", "Report: Semiconductor PF (alias of arim_report_semiconductor)", "...", "alias")
    add("report_tech_area", "Report: Technology area (alias of arim_report_tech_area)", "...", "alias")
    add("report_keywords", "Report: Keywords (alias of arim_report_keywords)", "...", "alias")
    add("report_user_name", "Report: User name (alias of arim_report_user_name)", "...", "alias")
    add("report_affiliation", "Report: Affiliation (alias of arim_report_affiliation)", "...", "alias")
    add("report_collaborators", "Report: Collaborators (alias of arim_report_collaborators)", "...", "alias")
    add("report_supporters", "Report: Supporters (alias of arim_report_supporters)", "...", "alias")
    add("report_support_type", "Report: Support type (alias of arim_report_support_type)", "...", "alias")
    add("report_abstract", "Report: Abstract (alias of arim_report_abstract)", "...", "alias")
    add("report_experimental", "Report: Experimental (alias of arim_report_experimental)", "...", "alias")
    add("report_results", "Report: Results (alias of arim_report_results)", "...", "alias")
    add("report_remarks", "Report: Remarks (alias of arim_report_remarks)", "...", "alias")
    add("report_publications", "Report: Publications (alias of arim_report_publications)", "...", "alias")
    add("report_presentations", "Report: Presentations (alias of arim_report_presentations)", "...", "alias")
    add("report_patents", "Report: Patents (alias of arim_report_patents)", "...", "alias")
    add("report_cross_tech_main", "Report: Cross technology area main (alias)", "...", "alias")
    add("report_cross_tech_sub", "Report: Cross technology area sub (alias)", "...", "alias")
    add("report_important_tech_main", "Report: Important technology area main (alias)", "...", "alias")
    add("report_important_tech_sub", "Report: Important technology area sub (alias)", "...", "alias")

    # LLM設定
    add("llm_provider", "LLMプロバイダー名", "gemini", "ai_config")
    add("llm_model", "LLMモデル名", "gemini-2.0-flash", "ai_config")
    add("llm_model_name", "LLM識別子（provider:model）", "gemini:gemini-2.0-flash", "ai_config")

    # converted.xlsx（output側）列
    report_path = get_dynamic_file_path("output/arim-site/reports/converted.xlsx")
    headers = _read_excel_headers(report_path)
    for col in headers:
        # raw column placeholder
        add(col, f"converted.xlsx 列 '{col}'", "...", "converted.xlsx")
        alias = _sanitize_placeholder_key(col)
        if alias:
            add(f"converted_xlsx_{alias}", f"converted.xlsx 列 '{col}'（安全キー）", "...", "converted.xlsx")

    return placeholders

def load_ai_extension_config():
    """AI拡張設定ファイルを読み込む"""
    try:
        config_path = get_dynamic_file_path("input/ai/ai_ext_conf.json")
        
        if os.path.exists(config_path):
            # メモ化（同一プロセス内での重複読み込みを避ける）
            # - UI起動中に複数箇所から参照されるため、ここでI/Oを抑制する
            # - 呼び出し側が辞書を変更しても良いように deepcopy で返す
            global _AI_EXT_CONF_CACHE_PATH, _AI_EXT_CONF_CACHE_MTIME, _AI_EXT_CONF_CACHE_DATA

            try:
                mtime = os.path.getmtime(config_path)
            except Exception:
                mtime = None

            if (
                _AI_EXT_CONF_CACHE_DATA is not None
                and _AI_EXT_CONF_CACHE_PATH == config_path
                and _AI_EXT_CONF_CACHE_MTIME == mtime
            ):
                return copy.deepcopy(_AI_EXT_CONF_CACHE_DATA)

            with open(config_path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            logger.info("AI拡張設定ファイルを読み込みました: %s", config_path)

            # 既定(setup)に存在するキー/ボタンを補完（古い input/ai/ai_ext_conf.json を自動マイグレーション）
            # - 既存ユーザーのカスタム設定を壊さない（欠落分だけ追加）
            # - pytest 実行中はワークスペースを汚さない
            try:
                defaults = get_default_ai_extension_config()
                changed = _merge_missing_ai_extension_config(config, defaults)
                if changed and not os.environ.get("PYTEST_CURRENT_TEST"):
                    try:
                        save_ai_extension_config(config)
                    except Exception:
                        pass
            except Exception:
                pass

            # Backward compatible normalization: add target_kind in-memory
            try:
                for btn in config.get('buttons', []) or []:
                    if isinstance(btn, dict):
                        btn.setdefault('target_kind', infer_ai_suggest_target_kind(btn))
                for btn in config.get('default_buttons', []) or []:
                    if isinstance(btn, dict):
                        btn.setdefault('target_kind', infer_ai_suggest_target_kind(btn))
            except Exception:
                pass

            _AI_EXT_CONF_CACHE_PATH = config_path
            _AI_EXT_CONF_CACHE_MTIME = mtime
            _AI_EXT_CONF_CACHE_DATA = config
            return copy.deepcopy(config)
        else:
            logger.info("AI拡張設定ファイルが見つかりません。デフォルト設定を使用します: %s", config_path)
            config = get_default_ai_extension_config()
            # 初回起動時は setup 側の既定を userdir/input に複製して永続化する
            # pytest 実行中はワークスペースを汚さない
            if not os.environ.get("PYTEST_CURRENT_TEST"):
                try:
                    save_ai_extension_config(config)
                except Exception:
                    pass
            return config

    except Exception as e:
        logger.error("AI拡張設定読み込みエラー: %s", e)
        logger.info("デフォルト設定を使用します")
        return get_default_ai_extension_config()


# load_ai_extension_config() のキャッシュ（mtimeで無効化）
_AI_EXT_CONF_CACHE_PATH: str | None = None
_AI_EXT_CONF_CACHE_MTIME: float | None = None
_AI_EXT_CONF_CACHE_DATA: dict | None = None


def normalize_results_json_keys(keys) -> List[str]:
    """Normalize configured JSON display keys for results table.

    Accepts:
      - list/tuple of strings
      - a single string (comma/newline separated)

    Returns:
      - de-duplicated list (preserving order)
      - empty strings removed
    """
    if keys is None:
        return []

    items: List[str] = []
    if isinstance(keys, (list, tuple)):
        for v in keys:
            if v is None:
                continue
            s = str(v).strip()
            if s:
                items.append(s)
    else:
        s = str(keys)
        # split by newline or comma
        parts = re.split(r"[\n,]+", s)
        for p in parts:
            sp = str(p).strip()
            if sp:
                items.append(sp)

    # de-dup preserve order
    seen = set()
    out: List[str] = []
    for it in items:
        if it in seen:
            continue
        seen.add(it)
        out.append(it)
    return out


def save_ai_extension_config(config: Dict):
    """AI拡張設定ファイルを保存する"""
    try:
        config_path = get_dynamic_file_path("input/ai/ai_ext_conf.json")
        os.makedirs(os.path.dirname(config_path), exist_ok=True)
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        logger.info("AI拡張設定ファイルを保存しました: %s", config_path)
        return True
    except Exception as e:
        logger.error("AI拡張設定保存エラー: %s", e)
        return False

def get_default_ai_extension_config():
    """デフォルトのAI拡張設定を取得"""
    # 既定は setup/input/ai/ai_ext_conf.json を優先（ソース実行時/バイナリ同梱時の両対応）
    candidates: list[str] = []
    try:
        candidates.append(get_dynamic_file_path("setup/input/ai/ai_ext_conf.json"))
    except Exception:
        pass
    try:
        # バイナリ時に _MEIPASS 配下へ同梱されている場合
        from config.common import get_static_resource_path

        candidates.append(get_static_resource_path("setup/input/ai/ai_ext_conf.json"))
    except Exception:
        pass

    for path in candidates:
        try:
            if not path or not os.path.exists(path):
                continue
            with open(path, "r", encoding="utf-8") as f:
                config = json.load(f)
            # Backward compatible normalization: add target_kind in-memory
            try:
                for btn in config.get('buttons', []) or []:
                    if isinstance(btn, dict):
                        btn.setdefault('target_kind', infer_ai_suggest_target_kind(btn))
                for btn in config.get('default_buttons', []) or []:
                    if isinstance(btn, dict):
                        btn.setdefault('target_kind', infer_ai_suggest_target_kind(btn))
            except Exception:
                pass
            return config
        except Exception:
            continue

    # 最終フォールバック（最低限）
    return {
        "version": "1.0.0",
        "description": "デフォルトAI拡張設定（最終フォールバック）",
        "buttons": [
            {
                "id": "default_analysis",
                "label": "総合分析",
                "description": "データセットの総合的な分析を実行",
                "prompt_template": "以下のデータセットについて総合的な分析を行ってください。\n\nデータセット名: {name}\nタイプ: {type}\n課題番号: {grant_number}\n既存説明: {description}\n\n分析項目:\n1. 技術的特徴\n2. 学術的価値\n3. 応用可能性\n4. データ品質\n5. 改善提案\n\n各項目について詳しく分析し、200文字程度で要約してください。",
                "icon": "📊",
                "category": "総合",
                "target_kind": "dataset",
            }
        ],
        "default_buttons": [],
        "ui_settings": {
            "buttons_per_row": 3,
            "button_height": 60,
            "button_width": 140,
            "response_area_height": 400,
            "enable_categories": True,
            "show_icons": True
        }
    }


def _merge_missing_ai_extension_config(config: Dict, defaults: Dict) -> bool:
    """Merge missing keys/buttons from defaults into config.

    - Does NOT overwrite existing user values.
    - Returns True if config is modified.
    """
    if not isinstance(config, dict) or not isinstance(defaults, dict):
        return False

    changed = False

    # Top-level keys: add only if missing
    for key, default_value in defaults.items():
        if key in {"buttons", "default_buttons"}:
            continue
        if key not in config:
            config[key] = copy.deepcopy(default_value)
            changed = True
            continue
        # Shallow merge dicts
        if isinstance(default_value, dict) and isinstance(config.get(key), dict):
            for sub_k, sub_v in default_value.items():
                if sub_k not in config[key]:
                    config[key][sub_k] = copy.deepcopy(sub_v)
                    changed = True

    # default_buttons list: add if missing
    if "default_buttons" not in config and "default_buttons" in defaults:
        config["default_buttons"] = copy.deepcopy(defaults.get("default_buttons") or [])
        changed = True

    # buttons: ensure default buttons exist by id
    config_buttons = config.get("buttons")
    default_buttons = defaults.get("buttons")
    if not isinstance(config_buttons, list):
        config_buttons = []
        config["buttons"] = config_buttons
        changed = True
    if isinstance(default_buttons, list) and default_buttons:
        existing_ids = set()
        for b in config_buttons:
            if isinstance(b, dict):
                bid = (b.get("id") or "").strip()
                if bid:
                    existing_ids.add(bid)
        for b in default_buttons:
            if not isinstance(b, dict):
                continue
            bid = (b.get("id") or "").strip()
            if not bid or bid in existing_ids:
                continue
            config_buttons.append(copy.deepcopy(b))
            existing_ids.add(bid)
            changed = True

    return changed

def load_prompt_file(prompt_file_path):
    """プロンプトファイルを読み込む"""
    try:
        # 絶対パスかチェック
        if os.path.isabs(prompt_file_path):
            full_path = prompt_file_path
        else:
            # 相対パスは動的パスとして解決（バイナリ時はユーザーディレクトリを使用）
            full_path = get_dynamic_file_path(prompt_file_path)
        
        logger.debug("プロンプトファイル読み込み試行: %s", full_path)
        
        if os.path.exists(full_path):
            with open(full_path, 'r', encoding='utf-8') as f:
                content = f.read()
                logger.info("プロンプトファイル読み込み成功: %s", full_path)
                return content
        else:
            logger.warning("プロンプトファイルが見つかりません: %s", full_path)
            return None
            
    except Exception as e:
        logger.error("プロンプトファイル読み込みエラー: %s", e)
        return None

def save_prompt_file(prompt_file_path, content):
    """プロンプトファイルを保存する"""
    try:
        full_path = get_dynamic_file_path(prompt_file_path)
        
        # ディレクトリが存在しない場合は作成
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        
        with open(full_path, 'w', encoding='utf-8') as f:
            f.write(content)
        return True
            
    except Exception as e:
        logger.error("プロンプトファイル保存エラー: %s", e)
        return False

def _extract_prompt_placeholders(prompt_template: str) -> List[str]:
    return list(dict.fromkeys(_PROMPT_PLACEHOLDER_PATTERN.findall(prompt_template or "")))


def _apply_runtime_prompt_assembly_override(
    config: Dict,
    feature_id: str,
    prompt_assembly_override: Optional[Dict],
) -> None:
    if not feature_id or not isinstance(prompt_assembly_override, dict):
        return

    prompt_assembly = config.setdefault('prompt_assembly', {})
    features = prompt_assembly.setdefault('features', {})
    feature_override = features.setdefault(feature_id, {})

    mode_value = prompt_assembly_override.get('mode')
    if mode_value in {'full_embed', 'filtered_embed'}:
        feature_override['mode'] = mode_value

    source_overrides = prompt_assembly_override.get('sources')
    if isinstance(source_overrides, dict):
        merged_sources = feature_override.setdefault('sources', {})
        for placeholder, override in source_overrides.items():
            if not isinstance(override, dict):
                continue
            target_override = merged_sources.setdefault(placeholder, {})
            source_mode = override.get('mode')
            if source_mode in {'full_embed', 'filtered_embed'}:
                target_override['mode'] = source_mode


def _load_prompt_runtime_ai_config(feature_id: str = "", prompt_assembly_override: Optional[Dict] = None) -> Dict:
    try:
        from classes.config.ui.ai_settings_widget import get_ai_config

        config = get_ai_config() or {}
    except Exception:
        config = {}

    try:
        from classes.ai.util.generation_params import normalize_ai_config_inplace

        normalize_ai_config_inplace(config)
    except Exception:
        pass

    try:
        _merge_button_prompt_assembly_override(config, feature_id)
    except Exception:
        logger.debug("button-level prompt assembly override merge failed", exc_info=True)
    try:
        _apply_runtime_prompt_assembly_override(config, feature_id, prompt_assembly_override)
    except Exception:
        logger.debug("runtime prompt assembly override merge failed", exc_info=True)
    return config


def _merge_button_prompt_assembly_override(config: Dict, feature_id: str) -> None:
    if not feature_id:
        return
    ext_conf = load_ai_extension_config() or {}
    buttons = ext_conf.get('buttons', []) or []
    button_config = None
    for button in buttons:
        if (button.get('id') or '').strip() == feature_id:
            button_config = button
            break
    if not isinstance(button_config, dict):
        return

    prompt_assembly = config.setdefault('prompt_assembly', {})
    features = prompt_assembly.setdefault('features', {})
    feature_override = features.setdefault(feature_id, {})

    for source_key, target_key in _BUTTON_PROMPT_ASSEMBLY_FIELD_MAP.items():
        value = button_config.get(source_key)
        if value in (None, ""):
            continue
        feature_override[target_key] = value

    source_overrides = button_config.get('prompt_assembly_sources')
    if isinstance(source_overrides, dict):
        merged_sources = feature_override.setdefault('sources', {})
        for placeholder, override in source_overrides.items():
            if not isinstance(override, dict):
                continue
            target_override = merged_sources.setdefault(placeholder, {})
            mode_value = override.get('mode')
            if mode_value in {'full_embed', 'filtered_embed'}:
                target_override['mode'] = mode_value


def format_prompt_with_context_details(
    prompt_template,
    context_data,
    *,
    feature_id: str = "",
    template_name: str = "",
    template_path: str = "",
    prompt_assembly_override: Optional[Dict] = None,
):
    """プロンプトテンプレートをコンテキストデータで置換し、診断情報も返す。"""
    try:
        placeholders = set(_extract_prompt_placeholders(prompt_template))
        
        # ARIM報告書データを取得・統合
        enhanced_context = context_data.copy()
        # エイリアスと不足キーのフォールバックを事前適用
        try:
            if 'type' not in enhanced_context and 'dataset_type' in enhanced_context:
                enhanced_context['type'] = enhanced_context.get('dataset_type') or ''
            # description と existing_description の相互エイリアス
            if 'existing_description' not in enhanced_context and 'description' in enhanced_context:
                enhanced_context['existing_description'] = enhanced_context.get('description') or ''
            if 'description' not in enhanced_context and 'existing_description' in enhanced_context:
                enhanced_context['description'] = enhanced_context.get('existing_description') or ''
            if 'llm_model_name' not in enhanced_context:
                provider = enhanced_context.get('llm_provider') or ''
                model = enhanced_context.get('llm_model') or ''
                # provider/model が両方空の場合、AIManagerからデフォルト設定を取得
                if not provider and not model:
                    try:
                        from classes.ai.core.ai_manager import AIManager
                        ai_mgr = AIManager()
                        provider = ai_mgr.get_default_provider()
                        model = ai_mgr.get_default_model(provider)
                        logger.debug(f"llm_model_name未設定のため、AIManagerからデフォルト取得: {provider}:{model}")
                    except Exception as e:
                        logger.debug(f"AIManager設定取得失敗、デフォルト値を使用: {e}")
                        provider = 'gemini'
                        model = 'gemini-2.0-flash'
                enhanced_context['llm_model_name'] = f"{provider}:{model}".strip(':')
            # プレースホルダに空文字を入れて未置換を防ぐ
            for k in ['material_index_data', 'equipment_data', 'file_tree', 'text_from_structured_files', 'json_from_structured_files']:
                if k not in enhanced_context:
                    enhanced_context[k] = ''
        except Exception as _alias_err:
            logger.debug("テンプレート置換のエイリアス適用で警告: %s", _alias_err)
        grant_number = context_data.get('grant_number')
        offline_mode = os.environ.get('ARIM_FETCHER_OFFLINE', '').lower() in ('1', 'true', 'yes')

        needs_report_context = any(
            placeholder.startswith('arim_report_') or placeholder.startswith('report_')
            for placeholder in placeholders
        )
        if grant_number and grant_number != "未設定" and not offline_mode and needs_report_context:
            logger.debug("ARIM報告書データ取得開始: %s", grant_number)
            try:
                from classes.dataset.util.arim_report_fetcher import fetch_arim_report_data
                arim_data = fetch_arim_report_data(grant_number)

                if arim_data:
                    enhanced_context.update(arim_data)
                    logger.info("ARIM報告書データを統合: %s項目", len(arim_data))

                    # arim_report_* の英語エイリアス report_* を追加
                    try:
                        enhanced_context.update(derive_report_aliases_from_arim_report(arim_data))
                    except Exception as e:
                        logger.debug("report_* エイリアス生成失敗: %s", e)
                    
                    # デバッグ用：取得したキーを表示
                    for key in arim_data.keys():
                        logger.debug("ARIM データキー: %s", key)
                else:
                    logger.info("ARIM報告書が見つかりませんでした: %s", grant_number)
            except Exception as e:
                logger.warning("ARIM報告書取得でエラー: %s", e)
                # エラーがあってもベースのコンテキストで続行
        elif offline_mode:
            logger.info("ARIM報告書取得をスキップしました（ARIM_FETCHER_OFFLINE モード）")

        # データポータルマスタデータを取得・統合
        try:
            requested_master_keys = [
                key for key in ('dataportal_material_index', 'dataportal_tag', 'dataportal_equipment')
                if key in placeholders
            ]
            master_data = load_dataportal_master_data(requested_master_keys)
            if master_data:
                enhanced_context.update(master_data)
                logger.debug("データポータルマスタデータを統合: %s項目", len(master_data))
        except Exception as e:
            logger.warning("データポータルマスタデータ取得でエラー: %s", e)

        # 静的マテリアルインデックス（MI.json）を取得・統合
        if 'static_material_index' in placeholders:
            try:
                static_mi = load_static_material_index()
                if static_mi:
                    enhanced_context.update(static_mi)
                    logger.debug("静的マテリアルインデックスを統合")
            except Exception as e:
                logger.warning("静的マテリアルインデックス取得でエラー: %s", e)

        # output/arim-site/reports/converted.xlsx の列データを取得・統合（列→プレースホルダ拡張）
        try:
            if grant_number and grant_number != "未設定":
                converted_placeholders = load_converted_xlsx_placeholders(str(grant_number))
                if converted_placeholders:
                    enhanced_context.update(converted_placeholders)
                    logger.debug("converted.xlsx 列プレースホルダを統合: %s項目", len(converted_placeholders))

                    # converted.xlsx から arim_report_* を不足補完（ネット取得より優先しない）
                    try:
                        derived_report = derive_arim_report_placeholders_from_converted(converted_placeholders)
                        for k, v in derived_report.items():
                            if not _safe_str(enhanced_context.get(k, '')).strip():
                                enhanced_context[k] = v
                        # 補完後に report_* の英語エイリアスも作る
                        enhanced_context.update(
                            {
                                k: v
                                for k, v in derive_report_aliases_from_arim_report(enhanced_context).items()
                                if not _safe_str(enhanced_context.get(k, '')).strip()
                            }
                        )
                    except Exception as e:
                        logger.debug("converted.xlsx 由来 arim_report_* 補完失敗: %s", e)
        except Exception as e:
            logger.warning("converted.xlsx プレースホルダ統合でエラー: %s", e)

        result = build_prompt(
            prompt_template,
            enhanced_context,
            ai_config=_load_prompt_runtime_ai_config(feature_id, prompt_assembly_override),
            feature_id=feature_id,
            template_name=template_name,
            template_path=template_path,
        )
        return result

    except Exception as e:
        logger.error("プロンプト置換エラー: %s", e)
        return build_prompt(
            prompt_template,
            context_data or {},
            ai_config=_load_prompt_runtime_ai_config(feature_id, prompt_assembly_override),
            feature_id=feature_id,
            template_name=template_name,
            template_path=template_path,
        )


def format_prompt_with_context(prompt_template, context_data, *, feature_id: str = "", template_name: str = "", template_path: str = "", prompt_assembly_override: Optional[Dict] = None):
    """プロンプトテンプレートをコンテキストデータで置換する（ARIM報告書対応・データポータルマスタ対応）"""
    result = format_prompt_with_context_details(
        prompt_template,
        context_data,
        feature_id=feature_id,
        template_name=template_name,
        template_path=template_path,
        prompt_assembly_override=prompt_assembly_override,
    )
    return result.prompt


def load_dataportal_master_data(requested_keys: Optional[Iterable[str]] = None):
    """データポータルマスタデータを読み込む
    
    Returns:
        dict: プレースホルダ用のマスタデータ辞書
            - dataportal_material_index: マテリアルインデックスマスタ（JSON文字列）
            - dataportal_tag: タグマスタ（JSON文字列）
            - dataportal_equipment: 装置分類マスタ（JSON文字列）
    """
    result = {}

    # マスタデータの定義（ファイル名パターン）
    master_types = [
        ('dataportal_material_index', 'material_index'),
        ('dataportal_tag', 'tag'),
        ('dataportal_equipment', 'equipment')
    ]
    requested = set(requested_keys or [name for name, _prefix in master_types])

    for placeholder_key, file_prefix in master_types:
        if placeholder_key not in requested:
            continue
        try:
            cached = _DATAPORTAL_MASTER_CACHE.get(placeholder_key)
            if cached is not None:
                result[placeholder_key] = cached
                continue
            # production優先、なければtestを使用
            production_path = get_dynamic_file_path(f'input/master_data/{file_prefix}_production.json')
            test_path = get_dynamic_file_path(f'input/master_data/{file_prefix}_test.json')

            target_path = None
            if os.path.exists(production_path):
                target_path = production_path
                logger.debug("マスタデータ読み込み（production）: %s", file_prefix)
            elif os.path.exists(test_path):
                target_path = test_path
                logger.debug("マスタデータ読み込み（test）: %s", file_prefix)
            else:
                logger.warning("マスタデータファイルが見つかりません: %s", file_prefix)
                result[placeholder_key] = "マスタデータなし"
                continue
            
            # JSONファイル読み込み
            with open(target_path, 'r', encoding='utf-8') as f:
                master_json = json.load(f)

            # JSON文字列として格納（整形して見やすく）
            result[placeholder_key] = json.dumps(master_json, ensure_ascii=False, indent=2)
            _DATAPORTAL_MASTER_CACHE[placeholder_key] = result[placeholder_key]
            logger.info("マスタデータ読み込み成功: %s (件数: %s)", file_prefix, master_json.get('count', 'N/A'))

        except Exception as e:
            logger.error("マスタデータ読み込みエラー (%s): %s", file_prefix, e)
            result[placeholder_key] = f"マスタデータ読み込みエラー: {str(e)}"

    return result


def load_static_material_index():
    """静的マテリアルインデックス(MI.json)を読み込み、プレースホルダ提供

    Returns:
        dict: { 'static_material_index': '<JSON文字列>' }
    """
    global _STATIC_MATERIAL_INDEX_CACHE
    try:
        if _STATIC_MATERIAL_INDEX_CACHE is not None:
            return {'static_material_index': _STATIC_MATERIAL_INDEX_CACHE}
        mi_path = get_dynamic_file_path('input/ai/MI.json')
        if not os.path.exists(mi_path):
            logger.info("MI.jsonが見つかりません: %s", mi_path)
            # テストの安定性のため、空配列のJSONを返す
            return {'static_material_index': '[]'}

        with open(mi_path, 'r', encoding='utf-8') as f:
            mi_json = json.load(f)

        mi_str = json.dumps(mi_json, ensure_ascii=False, indent=2)
        _STATIC_MATERIAL_INDEX_CACHE = mi_str
        logger.info("MI.json読み込み成功（カテゴリ数推定）")
        return {'static_material_index': mi_str}

    except Exception as e:
        logger.error("MI.json読み込みエラー: %s", e)
        # エラー時も空配列のJSONを返す
        return {'static_material_index': '[]'}