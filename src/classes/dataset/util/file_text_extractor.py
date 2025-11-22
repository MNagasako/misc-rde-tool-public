"""
ファイルからテキスト内容を抽出するユーティリティモジュール
STRUCTURED ファイルからテキストを抽出し、AIプロンプトに組み込む
"""

import os
import json
import re
import logging
from typing import Optional, Dict, Any, List

logger = logging.getLogger(__name__)


class FileTextExtractor:
    """ファイルからテキスト内容を抽出するクラス"""
    
    # 画像ファイル拡張子（除外対象・デフォルト）
    DEFAULT_IMAGE_EXTENSIONS = {
        '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif',
        '.svg', '.webp', '.ico', '.heic', '.raw', '.cr2', '.nef'
    }
    
    # テキスト抽出対応拡張子（デフォルト）
    DEFAULT_SUPPORTED_EXTENSIONS = {
        '.txt', '.csv', '.json', '.log', '.md', '.yaml', '.yml', '.xml', '.html', '.xlsx', '.xls', '.xlsm'
    }
    
    # 除外パターン（デフォルト）
    DEFAULT_EXCLUDE_PATTERNS = [
        r'.*_anonymized\.json',
        r'[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}\.json'
    ]
    
    def __init__(self):
        # デフォルト設定
        self.max_file_size = 10 * 1024 * 1024  # 10MB制限（デフォルト）
        self.max_text_length = 10000  # 抽出テキストの最大文字数（1ファイルあたり）
        self.max_files = 10  # 処理する最大ファイル数
        self.target_extensions = list(self.DEFAULT_SUPPORTED_EXTENSIONS)
        self.exclude_patterns = list(self.DEFAULT_EXCLUDE_PATTERNS)
        self.excel_all_sheets = True  # 全シートを処理するか
        self.excel_max_rows = 1000  # シートあたり最大行数
        
        # 設定を読み込み
        self._load_config()
    
    def _load_config(self):
        """app_config.jsonから設定を読み込む"""
        try:
            from config.common import get_dynamic_file_path
            config_path = get_dynamic_file_path('config/app_config.json')
            
            if os.path.exists(config_path):
                with open(config_path, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    extraction_config = config.get('file_text_extraction', {})
                    
                    # 設定を適用
                    if 'max_file_size_bytes' in extraction_config:
                        self.max_file_size = extraction_config['max_file_size_bytes']
                    
                    if 'max_chars_per_file' in extraction_config:
                        self.max_text_length = extraction_config['max_chars_per_file']
                    
                    if 'max_files' in extraction_config:
                        self.max_files = extraction_config['max_files']
                    
                    if 'target_extensions' in extraction_config:
                        self.target_extensions = extraction_config['target_extensions']
                    
                    if 'exclude_patterns' in extraction_config:
                        self.exclude_patterns = extraction_config['exclude_patterns']
                    
                    if 'excel_all_sheets' in extraction_config:
                        self.excel_all_sheets = extraction_config['excel_all_sheets']
                    
                    if 'excel_max_rows' in extraction_config:
                        self.excel_max_rows = extraction_config['excel_max_rows']
                    
                    logger.info("ファイル抽出設定を読み込みました: max_files=%d, max_chars=%d", 
                               self.max_files, self.max_text_length)
            else:
                logger.debug("app_config.jsonが存在しないため、デフォルト設定を使用します")
                
        except Exception as e:
            logger.warning("設定読み込みエラー（デフォルト設定を使用）: %s", e)
    
    def is_extractable(self, file_path: str) -> bool:
        """
        ファイルがテキスト抽出可能かどうかを判定
        
        Args:
            file_path: ファイルパスまたはファイル名
            
        Returns:
            bool: 抽出可能ならTrue
        """
        file_name = os.path.basename(file_path)
        
        # 除外パターンチェック
        for pattern in self.exclude_patterns:
            try:
                if re.match(pattern, file_name, re.IGNORECASE):
                    logger.debug(f"除外パターンにマッチ: {file_name} (パターン: {pattern})")
                    return False
            except re.error as e:
                logger.warning(f"正規表現エラー: {pattern} - {e}")
                continue
        
        # ファイルサイズチェック（実際のパスの場合のみ）
        if os.path.exists(file_path):
            try:
                file_size = os.path.getsize(file_path)
                if file_size > self.max_file_size:
                    logger.debug(f"ファイルサイズ超過: {file_path} ({file_size} bytes)")
                    return False
            except OSError:
                return False
        
        # 拡張子チェック
        _, ext = os.path.splitext(file_path)
        ext_lower = ext.lower()
        
        # 画像ファイルは除外
        if ext_lower in self.DEFAULT_IMAGE_EXTENSIONS:
            return False
        
        # サポート対象の拡張子（設定から）
        return ext_lower in self.target_extensions
    
    def extract_text(self, file_path: str, file_name: str = None) -> Optional[str]:
        """
        ファイルからテキストを抽出
        
        Args:
            file_path: ファイルパス
            file_name: ファイル名（ログ出力用、省略可）
            
        Returns:
            str: 抽出されたテキスト（失敗時はNone）
        """
        if not file_name:
            file_name = os.path.basename(file_path)
        
        if not self.is_extractable(file_path):
            logger.debug(f"テキスト抽出非対応: {file_name}")
            return None
        
        _, ext = os.path.splitext(file_path)
        ext_lower = ext.lower()
        
        try:
            # Excelファイル
            if ext_lower in {'.xlsx', '.xls', '.xlsm'}:
                return self._extract_excel_text(file_path, file_name)
            
            # プレーンテキストファイル
            else:
                return self._extract_plain_text(file_path, file_name)
                
        except Exception as e:
            logger.warning(f"テキスト抽出エラー ({file_name}): {e}")
            return None
    
    def _extract_plain_text(self, file_path: str, file_name: str) -> Optional[str]:
        """
        プレーンテキストファイルから内容を読み込み
        
        Args:
            file_path: ファイルパス
            file_name: ファイル名（ログ用）
            
        Returns:
            str: テキスト内容
        """
        try:
            # エンコーディングを自動検出
            encodings = ['utf-8', 'utf-8-sig', 'shift-jis', 'cp932', 'euc-jp', 'iso-2022-jp']
            
            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        text = f.read(self.max_text_length)
                        logger.debug(f"テキスト抽出成功: {file_name} ({encoding})")
                        return text
                except (UnicodeDecodeError, UnicodeError):
                    continue
            
            # 全てのエンコーディングで失敗した場合
            logger.warning(f"エンコーディング判定失敗: {file_name}")
            
            # バイナリモードで読み込み、可能な限りデコード
            with open(file_path, 'rb') as f:
                raw_data = f.read(self.max_text_length)
                text = raw_data.decode('utf-8', errors='ignore')
                logger.debug(f"バイナリモードで抽出: {file_name}")
                return text
                
        except Exception as e:
            logger.warning(f"プレーンテキスト抽出エラー ({file_name}): {e}")
            return None
    
    def _extract_excel_text(self, file_path: str, file_name: str) -> Optional[str]:
        """
        Excelファイルからテキストを抽出
        
        Args:
            file_path: ファイルパス
            file_name: ファイル名（ログ用）
            
        Returns:
            str: 抽出されたテキスト
        """
        try:
            import openpyxl
            import warnings
            
            # openpyxlのData Validation警告を抑制
            with warnings.catch_warnings():
                warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
                wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            extracted_lines = []
            char_count = 0
            
            # シート処理範囲を決定
            sheets_to_process = wb.sheetnames if self.excel_all_sheets else [wb.sheetnames[0]]
            
            for sheet_name in sheets_to_process:
                sheet = wb[sheet_name]
                
                # シート名を追加
                if self.excel_all_sheets:
                    extracted_lines.append(f"[シート: {sheet_name}]")
                    char_count += len(extracted_lines[-1])
                
                # 行を処理（設定された最大行数を使用）
                max_rows = min(sheet.max_row, self.excel_max_rows)
                
                for row in sheet.iter_rows(min_row=1, max_row=max_rows, values_only=True):
                    # 空行をスキップ
                    if not any(row):
                        continue
                    
                    # セルの値を文字列に変換
                    row_text = '\t'.join(str(cell) if cell is not None else '' for cell in row)
                    
                    # 文字数制限チェック
                    if char_count + len(row_text) > self.max_text_length:
                        extracted_lines.append("... (以降省略)")
                        break
                    
                    extracted_lines.append(row_text)
                    char_count += len(row_text)
                
                # 文字数制限に達したら終了
                if char_count >= self.max_text_length:
                    break
                
                if self.excel_all_sheets:
                    extracted_lines.append("")  # シート間の空行
            
            wb.close()
            
            result_text = '\n'.join(extracted_lines)
            logger.info(f"Excel抽出成功: {file_name} ({len(result_text)}文字, シート数: {len(sheets_to_process)})")
            return result_text
            
        except ImportError:
            logger.error("openpyxlがインストールされていません")
            return None
        except Exception as e:
            logger.warning(f"Excel抽出エラー ({file_name}): {e}")
            return None


def format_extracted_files_for_prompt(file_contents: Dict[str, str]) -> str:
    """
    抽出されたファイルコンテンツをプロンプト用にフォーマット
    
    Args:
        file_contents: {ファイル名: テキスト内容} の辞書
        
    Returns:
        str: プロンプト用にフォーマットされた文字列
    """
    if not file_contents:
        return '（データセットにSTRUCTUREDファイルが含まれていないか、テキスト抽出に失敗しました）'
    
    formatted_parts = []
    
    for file_name, content in file_contents.items():
        if not content:
            continue
        
        formatted_parts.append(f"■ ファイル: {file_name}")
        formatted_parts.append("```")
        
        # 内容を適度に省略（1ファイルあたり最大5000文字）
        max_content_length = 5000
        if len(content) > max_content_length:
            formatted_parts.append(content[:max_content_length])
            formatted_parts.append("... (以下省略)")
        else:
            formatted_parts.append(content)
        
        formatted_parts.append("```")
        formatted_parts.append("")  # 空行
    
    result = '\n'.join(formatted_parts)
    
    # 統計情報を追加
    total_files = len(file_contents)
    total_chars = sum(len(c) for c in file_contents.values() if c)
    
    header = f"【データセット内STRUCTUREDファイルのテキスト内容】\n"
    header += f"抽出ファイル数: {total_files}件 / 総文字数: {total_chars:,}文字\n\n"
    
    return header + result


# グローバルインスタンス
_file_text_extractor = None

def get_file_text_extractor() -> FileTextExtractor:
    """FileTextExtractorのシングルトンインスタンスを取得"""
    global _file_text_extractor
    if _file_text_extractor is None:
        _file_text_extractor = FileTextExtractor()
    return _file_text_extractor
